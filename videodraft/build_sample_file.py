"""
Build Sample_Quarterly_Report.xlsx for Video 3 Universal Tools Demo.

This file is intentionally messy — it has all the problems that the universal
tools are designed to fix. Each problem is carefully placed so the demo
tools have something to clean up on camera.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, numbers
import random
import datetime

random.seed(42)  # Reproducible

# --- Data Setup ---
DEPARTMENTS = [
    "Engineering", "Sales", "Marketing", "Finance",
    "Operations", "Customer Success", "Product", "HR"
]

FIRST_NAMES = [
    "Sarah", "Michael", "Jennifer", "David", "Amanda", "Robert", "Emily", "James",
    "Jessica", "Daniel", "Ashley", "Christopher", "Stephanie", "Matthew", "Nicole",
    "Andrew", "Lauren", "Joshua", "Megan", "Brandon", "Rachel", "Ryan", "Samantha",
    "Kevin", "Rebecca", "Justin", "Michelle", "Tyler", "Kimberly", "Nathan",
    "Elizabeth", "Thomas", "Angela", "Jonathan", "Catherine", "Brian", "Heather",
    "Patrick", "Christina", "Gregory", "Melissa", "Scott", "Andrea", "Mark",
    "Donna", "Eric", "Sharon", "Stephen", "Pamela", "Kenneth"
]

LAST_NAMES = [
    "Anderson", "Thompson", "Garcia", "Martinez", "Robinson", "Clark", "Rodriguez",
    "Lewis", "Lee", "Walker", "Hall", "Allen", "Young", "Hernandez", "King",
    "Wright", "Lopez", "Hill", "Scott", "Green", "Adams", "Baker", "Gonzalez",
    "Nelson", "Carter", "Mitchell", "Perez", "Roberts", "Turner", "Phillips",
    "Campbell", "Parker", "Evans", "Edwards", "Collins", "Stewart", "Sanchez",
    "Morris", "Rogers", "Reed", "Cook", "Morgan", "Bell", "Murphy", "Bailey",
    "Rivera", "Cooper", "Richardson", "Cox", "Howard"
]

CATEGORIES = [
    "Software", "Consulting", "Travel", "Equipment",
    "Training", "Subscriptions", "Facilities", "Marketing Spend"
]

STATUSES = ["Approved", "Pending", "Reviewed", "Flagged", "Complete"]

DATE_FORMATS_TEXT = [
    lambda d: d.strftime("%m/%d/%Y"),        # MM/DD/YYYY
    lambda d: d.strftime("%d-%b-%Y"),         # DD-MMM-YYYY
    lambda d: d.strftime("%Y-%m-%d"),         # YYYY-MM-DD
    lambda d: d.strftime("%m-%d-%Y"),         # MM-DD-YYYY
    lambda d: d.strftime("%B %d, %Y"),        # Month DD, YYYY
]


def random_date():
    start = datetime.date(2025, 10, 1)
    end = datetime.date(2025, 12, 31)
    delta = (end - start).days
    return start + datetime.timedelta(days=random.randint(0, delta))


def random_amount():
    return round(random.uniform(500, 48000), 2)


def random_budget(amount):
    # Budget is usually close to actual, sometimes higher, sometimes lower
    factor = random.uniform(0.75, 1.30)
    return round(amount * factor, 2)


def build_sample_file():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Q4 Expense Data"

    # --- Headers (Row 1 — intentionally plain, no formatting) ---
    headers = ["Department", "Employee Name", "Transaction Date", "Amount",
               "Budget", "Variance", "Category", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Generate 130 rows of data ---
    data_rows = []
    used_names = set()
    for i in range(130):
        dept = random.choice(DEPARTMENTS)
        # Generate unique-ish names
        while True:
            name = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
            if name not in used_names or random.random() < 0.3:
                used_names.add(name)
                break
        date = random_date()
        amount = random_amount()
        budget = random_budget(amount)
        variance = round(amount - budget, 2)
        category = random.choice(CATEGORIES)
        status = random.choice(STATUSES)
        data_rows.append([dept, name, date, amount, budget, variance, category, status])

    # Sort by department so merging makes sense
    data_rows.sort(key=lambda r: r[0])

    # --- Write data with intentional problems ---
    blank_row_positions = {16, 33, 51, 68, 84, 102}  # 6 blank rows
    text_number_positions = set(random.sample(range(130), 26))  # ~20% text-stored numbers
    text_budget_positions = set(random.sample(range(130), 16))  # ~12% text-stored budgets
    space_name_positions = set(random.sample(range(130), 20))  # ~15% names with spaces

    current_row = 2
    data_idx = 0
    row_map = {}  # data_idx -> actual row

    for data_idx in range(130):
        # Insert blank row at specified positions
        if current_row in blank_row_positions:
            current_row += 1  # skip this row (leave it blank)

        row_map[data_idx] = current_row
        row = data_rows[data_idx]
        dept, name, date, amount, budget, variance, category, status = row

        # Column A: Department (will be merged later)
        ws.cell(row=current_row, column=1, value=dept)

        # Column B: Employee Name (some with leading/trailing spaces)
        if data_idx in space_name_positions:
            # Add invisible spaces
            space_type = random.choice(["leading", "trailing", "both"])
            if space_type == "leading":
                name = "  " + name
            elif space_type == "trailing":
                name = name + "   "
            else:
                name = " " + name + "  "
        ws.cell(row=current_row, column=2, value=name)

        # Column C: Date (mixed formats — ALL stored as text strings)
        fmt_func = random.choice(DATE_FORMATS_TEXT)
        date_str = fmt_func(date)
        cell = ws.cell(row=current_row, column=3, value=date_str)
        cell.number_format = '@'  # Text format

        # Column D: Amount (some stored as text)
        if data_idx in text_number_positions:
            # Store as text string to simulate text-stored numbers
            cell = ws.cell(row=current_row, column=4, value=str(round(amount, 2)))
            cell.number_format = '@'
        else:
            ws.cell(row=current_row, column=4, value=amount)

        # Column E: Budget (some stored as text)
        if data_idx in text_budget_positions:
            cell = ws.cell(row=current_row, column=5, value=str(round(budget, 2)))
            cell.number_format = '@'
        else:
            ws.cell(row=current_row, column=5, value=budget)

        # Column F: Variance (numeric, some negative — NOT red formatted)
        ws.cell(row=current_row, column=6, value=variance)

        # Column G: Category
        ws.cell(row=current_row, column=7, value=category)

        # Column H: Status
        ws.cell(row=current_row, column=8, value=status)

        current_row += 1

    last_data_row = current_row - 1

    # --- Merge cells in Column A (department groups) ---
    # Group consecutive same-department rows and merge
    dept_groups = []
    start_row = 2
    current_dept = ws.cell(row=2, column=1).value
    for r in range(3, last_data_row + 2):
        val = ws.cell(row=r, column=1).value if r <= last_data_row else None
        if val != current_dept or r > last_data_row:
            if r - start_row >= 3:  # Only merge groups of 3+ rows
                dept_groups.append((start_row, r - 1))
            start_row = r
            current_dept = val

    # Only merge the first few large groups (3 merge groups as per spec)
    merge_count = 0
    for start, end in dept_groups:
        if merge_count >= 3:
            break
        if end - start >= 5:  # Only merge groups of 6+ rows
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws.cell(row=start, column=1).alignment = Alignment(vertical='center')
            merge_count += 1

    # --- Total / Summary Row ---
    total_row = last_data_row + 2
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=4).value = f"=SUM(D2:D{last_data_row})"
    ws.cell(row=total_row, column=5).value = f"=SUM(E2:E{last_data_row})"
    ws.cell(row=total_row, column=6).value = f"=SUM(F2:F{last_data_row})"

    # --- Error formulas (below data area) ---
    error_row = total_row + 3
    ws.cell(row=error_row, column=1, value="Error Examples (for Health Check demo)")
    ws.cell(row=error_row + 1, column=1, value="DIV/0 error:")
    ws.cell(row=error_row + 1, column=2).value = "=1/0"
    ws.cell(row=error_row + 2, column=1, value="N/A error:")
    ws.cell(row=error_row + 2, column=2).value = '=VLOOKUP("NONEXISTENT",A1:B5,2,FALSE)'

    # --- External link formula ---
    ext_row = error_row + 4
    ws.cell(row=ext_row, column=1, value="External Link (for Link Finder demo):")
    ws.cell(row=ext_row, column=2).value = "='[OtherWorkbook.xlsx]Sheet1'!A1"

    # --- Intentionally bad column widths ---
    ws.column_dimensions['A'].width = 8   # Too narrow for "Customer Success"
    ws.column_dimensions['B'].width = 35  # Too wide
    ws.column_dimensions['C'].width = 10  # Too narrow for some date formats
    ws.column_dimensions['D'].width = 8   # Too narrow
    ws.column_dimensions['E'].width = 8   # Too narrow
    ws.column_dimensions['F'].width = 7   # Too narrow
    ws.column_dimensions['G'].width = 25  # Too wide
    ws.column_dimensions['H'].width = 6   # Too narrow

    # --- Hidden Sheet 1: "Archive Notes" ---
    ws2 = wb.create_sheet("Archive Notes")
    ws2.cell(row=1, column=1, value="Archive Notes")
    ws2.cell(row=2, column=1, value="This sheet contains historical notes from Q3 review.")
    ws2.cell(row=3, column=1, value="Q3 variance was driven by one-time consulting spend in Marketing.")
    ws2.cell(row=4, column=1, value="CFO approved the overage on 2025-09-15.")
    ws2.cell(row=5, column=1, value="No action needed for Q4 unless the pattern repeats.")
    ws2.sheet_state = 'hidden'

    # --- Hidden Sheet 2: "Legacy Data" ---
    ws3 = wb.create_sheet("Legacy Data")
    ws3.cell(row=1, column=1, value="Department")
    ws3.cell(row=1, column=2, value="Q2 Total")
    ws3.cell(row=1, column=3, value="Q3 Total")
    legacy_data = [
        ("Engineering", 245000, 267500),
        ("Sales", 189000, 201000),
        ("Marketing", 312000, 358000),
        ("Finance", 145000, 148500),
        ("Operations", 210000, 225000),
        ("Customer Success", 167000, 172000),
        ("Product", 198000, 215000),
        ("HR", 125000, 131000),
    ]
    for i, (dept, q2, q3) in enumerate(legacy_data, 2):
        ws3.cell(row=i, column=1, value=dept)
        ws3.cell(row=i, column=2, value=q2)
        ws3.cell(row=i, column=3, value=q3)
    ws3.sheet_state = 'hidden'

    # --- Second visible sheet: "Summary" (gives Branding more to work with) ---
    ws4 = wb.create_sheet("Department Summary")
    sum_headers = ["Department", "Total Spend", "Total Budget", "Variance", "% of Budget", "Status"]
    for col, h in enumerate(sum_headers, 1):
        ws4.cell(row=1, column=col, value=h)

    dept_summary = [
        ("Engineering", 156780.50, 162000, -5219.50, "96.8%", "On Track"),
        ("Sales", 98450.25, 95000, 3450.25, "103.6%", "Over Budget"),
        ("Marketing", 215300.00, 200000, 15300.00, "107.7%", "Over Budget"),
        ("Finance", 67890.75, 72000, -4109.25, "94.3%", "On Track"),
        ("Operations", 134560.00, 140000, -5440.00, "96.1%", "On Track"),
        ("Customer Success", 89200.30, 88000, 1200.30, "101.4%", "Monitor"),
        ("Product", 178900.00, 175000, 3900.00, "102.2%", "Monitor"),
        ("HR", 45600.80, 50000, -4399.20, "91.2%", "On Track"),
    ]
    for i, row_data in enumerate(dept_summary, 2):
        for col, val in enumerate(row_data, 1):
            ws4.cell(row=i, column=col, value=val)

    # Total row
    ws4.cell(row=10, column=1, value="Total")
    ws4.cell(row=10, column=2).value = "=SUM(B2:B9)"
    ws4.cell(row=10, column=3).value = "=SUM(C2:C9)"
    ws4.cell(row=10, column=4).value = "=SUM(D2:D9)"

    # --- Save ---
    output_path = "/home/user/claude-training-lab-code/videodraft/Sample_Quarterly_Report.xlsx"
    wb.save(output_path)
    print(f"Sample file saved to: {output_path}")
    print(f"Main sheet rows: {last_data_row} (including {len(blank_row_positions)} blank rows)")
    print(f"Merge groups: {merge_count}")
    print(f"Text-stored amounts: {len(text_number_positions)}")
    print(f"Text-stored budgets: {len(text_budget_positions)}")
    print(f"Names with spaces: {len(space_name_positions)}")
    print(f"Hidden sheets: 2 (Archive Notes, Legacy Data)")
    print(f"Visible sheets: 2 (Q4 Expense Data, Department Summary)")
    print(f"Error formulas: 2 (#DIV/0!, #N/A)")
    print(f"External link formula: 1")
    print(f"Total row at row: {total_row}")


if __name__ == "__main__":
    build_sample_file()

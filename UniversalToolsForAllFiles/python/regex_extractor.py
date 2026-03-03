"""
KBT Universal Tools — Regex Text Extractor
============================================================
Extracts structured patterns from free-text columns using
regular expressions (regex). No coding knowledge needed —
just pick a preset pattern or describe what you're looking for.

Preset patterns:
  invoice    — Invoice numbers (INV-12345, INV12345)
  email      — Email addresses
  phone      — US phone numbers
  date       — Dates in common formats
  currency   — Dollar amounts ($1,234.56)
  account    — Account codes (GL codes, alphanumeric IDs)
  zipcode    — US zip codes
  ssn        — Social Security Numbers (masked output)

Usage:
    python regex_extractor.py "C:\\path\\data.xlsx" --col "Description" --pattern invoice
    python regex_extractor.py "data.xlsx" --col "Notes" --pattern email
    python regex_extractor.py "data.xlsx" --col "Memo" --custom "[A-Z]{2}-\\d{4}"

Output: Saves "REGEX_EXTRACTED.xlsx" with original data + extracted values
"""

import sys
import os
import re
import argparse

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)


PATTERNS = {
    'invoice':  r'(?i)\b(?:INV|Invoice)[- #]?(\w{4,12})\b',
    'email':    r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b',
    'phone':    r'\b(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}\b',
    'date':     r'\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2})\b',
    'currency': r'\$\s?[\d,]+(?:\.\d{2})?',
    'account':  r'\b[A-Z]{1,4}[-_]?\d{3,8}\b',
    'zipcode':  r'\b\d{5}(?:-\d{4})?\b',
    'ssn':      r'\b\d{3}-\d{2}-\d{4}\b',
}

PATTERN_DESCRIPTIONS = {
    'invoice':  'Invoice numbers (INV-12345)',
    'email':    'Email addresses',
    'phone':    'US phone numbers',
    'date':     'Dates (MM/DD/YYYY, YYYY-MM-DD, etc.)',
    'currency': 'Dollar amounts ($1,234.56)',
    'account':  'Account/GL codes (ABC-12345)',
    'zipcode':  'US ZIP codes',
    'ssn':      'Social Security Numbers (will be masked in output)',
}


def extract_regex(file_path: str, column: str, pattern_key: str,
                  custom_pattern: str = None, sheet_name: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Regex Text Extractor")
    print(f"{'='*55}")
    print(f"  File:    {os.path.basename(file_path)}")
    print(f"  Column:  '{column}'")

    if custom_pattern:
        pattern = custom_pattern
        pattern_label = f"Custom: {custom_pattern}"
    elif pattern_key in PATTERNS:
        pattern = PATTERNS[pattern_key]
        pattern_label = PATTERN_DESCRIPTIONS[pattern_key]
    else:
        print(f"\nERROR: Unknown pattern '{pattern_key}'.")
        print(f"Available presets: {', '.join(PATTERNS.keys())}")
        sys.exit(1)

    print(f"  Pattern: {pattern_label}")
    print(f"{'='*55}\n")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
    df.columns = df.columns.str.strip()

    if column not in df.columns:
        print(f"ERROR: Column '{column}' not found.")
        print(f"Available columns: {', '.join(df.columns.tolist())}")
        sys.exit(1)

    def extract_first(text):
        if pd.isna(text):
            return ""
        matches = re.findall(pattern, str(text))
        if not matches:
            return ""
        result = matches[0] if isinstance(matches[0], str) else matches[0][0]
        # Mask SSNs
        if pattern_key == 'ssn':
            result = "***-**-" + result[-4:]
        return result

    def extract_all(text):
        if pd.isna(text):
            return ""
        matches = re.findall(pattern, str(text))
        if not matches:
            return ""
        all_matches = [m if isinstance(m, str) else m[0] for m in matches]
        if pattern_key == 'ssn':
            all_matches = ["***-**-" + m[-4:] for m in all_matches]
        return " | ".join(all_matches)

    new_col = f"Extracted_{pattern_key or 'Custom'}"
    all_col = f"All_{pattern_key or 'Custom'}_Matches"

    df[new_col] = df[column].apply(extract_first)
    df[all_col] = df[column].apply(extract_all)

    extracted_count = (df[new_col] != "").sum()
    print(f"Records scanned:   {len(df)}")
    print(f"Matches found:     {extracted_count}")
    print(f"No match:          {len(df) - extracted_count}")

    output_path = os.path.join(os.path.dirname(file_path), "REGEX_EXTRACTED.xlsx")

    fill_hit    = PatternFill("solid", fgColor="C6EFCE")
    fill_miss   = PatternFill("solid", fgColor="FFC7CE")
    fill_header = PatternFill("solid", fgColor="1F497D")
    font_header = Font(bold=True, color="FFFFFF")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Extracted', index=False)
        ws = writer.sheets['Extracted']

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header

        extracted_col_idx = df.columns.get_loc(new_col) + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            val = row[extracted_col_idx - 1].value
            fill = fill_hit if val and str(val).strip() != "" else fill_miss
            row[extracted_col_idx - 1].fill = fill

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Matches-only sheet
        matches_df = df[df[new_col] != ""]
        if not matches_df.empty:
            matches_df.to_excel(writer, sheet_name='Matches Only', index=False)

    print(f"\n{'='*55}")
    print(f"  DONE! Saved to: {output_path}")
    print(f"  Green=Match Found | Red=No Match")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Regex Text Extractor')
    parser.add_argument('file', help='Path to the Excel file')
    parser.add_argument('--col', required=True, help='Column name to extract from')
    parser.add_argument('--pattern', default='invoice',
                        choices=list(PATTERNS.keys()),
                        help=f'Preset pattern to use (default: invoice). Options: {", ".join(PATTERNS.keys())}')
    parser.add_argument('--custom', default=None,
                        help='Custom regex pattern (overrides --pattern)')
    parser.add_argument('--sheet', default=None, help='Sheet name to read')
    args = parser.parse_args()
    extract_regex(args.file, args.col, args.pattern, args.custom, args.sheet)


if __name__ == '__main__':
    main()

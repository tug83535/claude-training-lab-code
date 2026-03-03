"""
KBT Universal Tools — Budget vs. Actual Consolidator
============================================================
Merges 50+ department budget files from a folder into one
master file with dollar and percent variance columns.

Assumptions:
  - All files in the folder have the same column structure
  - Files contain an "Actual" column and a "Budget" column
  - Each file represents one department or entity

Usage:
    python consolidate_budget.py "C:\\path\\to\\folder"
    python consolidate_budget.py "C:\\path\\to\\folder" --actual "Actuals" --budget "Plan"
    python consolidate_budget.py "C:\\path\\to\\folder" --sheet "P&L"

Output: Saves "CONSOLIDATED_BUDGET_VS_ACTUAL.xlsx" in the folder
"""

import sys
import os
import glob
import argparse
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("ERROR: Required packages not installed.")
    print("Run: pip install pandas openpyxl")
    sys.exit(1)


def consolidate_budget(folder: str, actual_col: str, budget_col: str, sheet_name: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Budget vs. Actual Consolidator")
    print(f"{'='*55}")
    print(f"  Folder: {folder}")
    print(f"  Actual column: '{actual_col}' | Budget column: '{budget_col}'")
    print(f"  Date:   {datetime.now().strftime('%m/%d/%Y %I:%M %p')}")
    print(f"{'='*55}\n")

    if not os.path.isdir(folder):
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)

    files = glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xls"))
    output_name = "CONSOLIDATED_BUDGET_VS_ACTUAL.xlsx"
    files = [f for f in files if os.path.basename(f) != output_name]

    if not files:
        print("ERROR: No Excel files found in the folder.")
        sys.exit(1)

    print(f"Found {len(files)} file(s) to consolidate:\n")

    all_dfs = []
    skipped = []

    for f in sorted(files):
        fname = os.path.basename(f)
        try:
            if sheet_name:
                df = pd.read_excel(f, sheet_name=sheet_name)
            else:
                df = pd.read_excel(f)

            df.columns = df.columns.str.strip()

            if actual_col not in df.columns or budget_col not in df.columns:
                print(f"  SKIP: {fname} — columns '{actual_col}'/'{budget_col}' not found")
                skipped.append(fname)
                continue

            df['Source_File'] = fname
            all_dfs.append(df)
            print(f"  OK:   {fname} ({len(df)} rows)")
        except Exception as e:
            print(f"  ERROR: {fname} — {e}")
            skipped.append(fname)

    if not all_dfs:
        print("\nERROR: No files could be processed.")
        sys.exit(1)

    print(f"\nConsolidating {len(all_dfs)} file(s)...")

    master = pd.concat(all_dfs, ignore_index=True, sort=False)

    # Move Source_File to front
    cols = ['Source_File'] + [c for c in master.columns if c != 'Source_File']
    master = master[cols]

    # Add variance columns
    master[actual_col] = pd.to_numeric(master[actual_col], errors='coerce')
    master[budget_col] = pd.to_numeric(master[budget_col], errors='coerce')

    master['$ Variance (Act-Bud)'] = master[actual_col] - master[budget_col]
    master['% Variance'] = master.apply(
        lambda r: (r[actual_col] - r[budget_col]) / abs(r[budget_col])
        if pd.notna(r[budget_col]) and r[budget_col] != 0 else None,
        axis=1
    )

    output_path = os.path.join(folder, output_name)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        master.to_excel(writer, sheet_name='Consolidated', index=False)
        ws = writer.sheets['Consolidated']

        # Style header row
        header_fill = PatternFill("solid", fgColor="1F497D")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Find and format variance columns
        header_map = {cell.value: cell.column_letter for cell in ws[1]}

        if '$ Variance (Act-Bud)' in header_map:
            var_col = header_map['$ Variance (Act-Bud)']
            for row in ws.iter_rows(min_row=2, min_col=ws[f'{var_col}1'].column,
                                    max_col=ws[f'{var_col}1'].column):
                for cell in row:
                    cell.number_format = '#,##0.00;[Red](#,##0.00)'

        if '% Variance' in header_map:
            pct_col = header_map['% Variance']
            for row in ws.iter_rows(min_row=2, min_col=ws[f'{pct_col}1'].column,
                                    max_col=ws[f'{pct_col}1'].column):
                for cell in row:
                    cell.number_format = '0.0%'

        # Auto-fit columns
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Summary sheet
        summary_df = master.groupby('Source_File').agg(
            Rows=('Source_File', 'count'),
            Total_Actual=(actual_col, 'sum'),
            Total_Budget=(budget_col, 'sum')
        ).reset_index()
        summary_df['$ Variance'] = summary_df['Total_Actual'] - summary_df['Total_Budget']
        summary_df['% Variance'] = summary_df.apply(
            lambda r: r['$ Variance'] / abs(r['Total_Budget']) if r['Total_Budget'] != 0 else None,
            axis=1
        )
        summary_df.to_excel(writer, sheet_name='Summary by File', index=False)

    print(f"\n{'='*55}")
    print(f"  DONE!")
    print(f"  Files processed: {len(all_dfs)}")
    print(f"  Files skipped:   {len(skipped)}")
    print(f"  Total rows:      {len(master)}")
    print(f"  Saved to: {output_path}")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Budget vs. Actual Consolidator')
    parser.add_argument('folder', help='Path to the folder containing budget files')
    parser.add_argument('--actual', default='Actual',
                        help='Name of the Actual column (default: "Actual")')
    parser.add_argument('--budget', default='Budget',
                        help='Name of the Budget column (default: "Budget")')
    parser.add_argument('--sheet', default=None,
                        help='Specific sheet name to read from each file (default: first sheet)')
    args = parser.parse_args()
    consolidate_budget(args.folder, args.actual, args.budget, args.sheet)


if __name__ == '__main__':
    main()

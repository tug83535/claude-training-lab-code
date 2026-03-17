"""
KBT Universal Tools — Batch Process Folder of Files
============================================================
Runs the Universal Data Cleaner on every Excel file in a folder.
Processes dozens of files automatically and saves cleaned copies.

Can also run a custom Python transformation function if provided.

Usage:
    python batch_process.py "C:\\path\\to\\folder"
    python batch_process.py "C:\\path\\to\\folder" --pattern "*Q1*"
    python batch_process.py "C:\\path\\to\\folder" --output "C:\\output_folder"

Output: Saves "_CLEANED" copies of each file (in same folder or --output folder)
"""

import sys
import os
import glob
import argparse
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)


def clean_dataframe(df: pd.DataFrame) -> tuple:
    """Apply standard cleaning operations to a DataFrame. Returns (cleaned_df, stats_dict)."""
    stats = {'empty_rows': 0, 'spaces': 0, 'text_numbers': 0, 'dupes': 0}

    original_len = len(df)
    df.dropna(how='all', inplace=True)
    df.dropna(axis=1, how='all', inplace=True)
    stats['empty_rows'] = original_len - len(df)

    for col in df.select_dtypes(include='object').columns:
        before = df[col].copy()
        df[col] = df[col].str.strip()
        stats['spaces'] += (before != df[col]).sum()

    for col in df.select_dtypes(include='object').columns:
        converted = pd.to_numeric(df[col], errors='coerce')
        mask = converted.notna() & df[col].notna()
        df.loc[mask, col] = converted[mask]
        stats['text_numbers'] += mask.sum()

    before_len = len(df)
    df.drop_duplicates(inplace=True)
    stats['dupes'] = before_len - len(df)

    return df, stats


def batch_process(folder: str, output_folder: str = None, pattern: str = "*.xlsx") -> None:
    print(f"\n{'='*55}")
    print("  KBT Batch File Processor")
    print(f"{'='*55}")
    print(f"  Folder:  {folder}")
    print(f"  Pattern: {pattern}")
    print(f"  Date:    {datetime.now().strftime('%m/%d/%Y %I:%M %p')}")
    print(f"{'='*55}\n")

    if not os.path.isdir(folder):
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)

    out_dir = output_folder if output_folder else folder
    if output_folder and not os.path.exists(output_folder):
        os.makedirs(output_folder)

    files = [f for f in glob.glob(os.path.join(folder, pattern))
             if '_CLEANED' not in os.path.basename(f)
             and 'BATCH_LOG' not in os.path.basename(f)]

    if not files:
        print(f"ERROR: No files matching '{pattern}' found.")
        sys.exit(1)

    print(f"Found {len(files)} file(s) to process.\n")

    log_rows = []
    processed = 0
    errors = 0

    for f in sorted(files):
        fname = os.path.basename(f)
        base, ext = os.path.splitext(fname)
        out_path = os.path.join(out_dir, base + '_CLEANED' + ext)

        try:
            xl = pd.ExcelFile(f)
            all_stats = {}
            total_rows_in = 0
            total_rows_out = 0

            with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
                for sht in xl.sheet_names:
                    df = pd.read_excel(f, sheet_name=sht)
                    rows_in = len(df)
                    df, stats = clean_dataframe(df)
                    rows_out = len(df)
                    total_rows_in += rows_in
                    total_rows_out += rows_out
                    all_stats[sht] = stats

                    df.to_excel(writer, sheet_name=sht, index=False)

                    ws = writer.sheets[sht]
                    fill = PatternFill("solid", fgColor="1F497D")
                    font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = fill
                        cell.font = font
                    for col in ws.columns:
                        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

            processed += 1
            print(f"  OK:    {fname}")
            print(f"         Rows: {total_rows_in} → {total_rows_out} | "
                  f"Spaces: {sum(s['spaces'] for s in all_stats.values())} | "
                  f"Dupes removed: {sum(s['dupes'] for s in all_stats.values())}")

            log_rows.append({
                'File': fname,
                'Status': 'SUCCESS',
                'Rows In': total_rows_in,
                'Rows Out': total_rows_out,
                'Empty Rows Removed': sum(s['empty_rows'] for s in all_stats.values()),
                'Spaces Fixed': sum(s['spaces'] for s in all_stats.values()),
                'Text-Numbers Fixed': sum(s['text_numbers'] for s in all_stats.values()),
                'Dupes Removed': sum(s['dupes'] for s in all_stats.values()),
                'Output File': os.path.basename(out_path)
            })

        except Exception as e:
            errors += 1
            print(f"  ERROR: {fname} — {e}")
            log_rows.append({'File': fname, 'Status': f'ERROR: {e}',
                             'Rows In': 0, 'Rows Out': 0})

    # Write processing log
    log_path = os.path.join(out_dir, "BATCH_PROCESSING_LOG.xlsx")
    log_df = pd.DataFrame(log_rows)
    with pd.ExcelWriter(log_path, engine='openpyxl') as writer:
        log_df.to_excel(writer, sheet_name='Log', index=False)
        ws = writer.sheets['Log']
        fill = PatternFill("solid", fgColor="1F497D")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    print(f"\n{'='*55}")
    print(f"  DONE!")
    print(f"  Processed: {processed} | Errors: {errors}")
    print(f"  Log saved: {log_path}")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Batch Process Folder of Files')
    parser.add_argument('folder', help='Folder containing Excel files to process')
    parser.add_argument('--output', default=None, help='Output folder (default: same as input)')
    parser.add_argument('--pattern', default='*.xlsx', help='File name pattern (default: *.xlsx)')
    args = parser.parse_args()
    batch_process(args.folder, args.output, args.pattern)


if __name__ == '__main__':
    main()

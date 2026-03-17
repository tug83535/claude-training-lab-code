"""
KBT Universal Tools — Fuzzy-Match Bank Reconciler
============================================================
Probabilistically matches ledger descriptions against bank
statement text using fuzzy string matching. Catches near-
matches that exact lookups miss (e.g., "AMAZON MKTP" vs "Amazon").

Usage:
    python bank_reconciler.py "C:\\path\\ledger.xlsx" "C:\\path\\bank.xlsx"
    python bank_reconciler.py "ledger.xlsx" "bank.xlsx" --desc "Description" --amount "Debit" --bank-desc "Memo"

Output: Saves "BANK_RECONCILIATION.xlsx" with matched/unmatched items
"""

import sys
import os
import argparse
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)

try:
    from thefuzz import fuzz, process
except ImportError:
    print("ERROR: thefuzz not installed. Run: pip install thefuzz python-Levenshtein")
    sys.exit(1)


def bank_reconcile(ledger_file: str, bank_file: str,
                   ledger_desc: str, ledger_amount: str,
                   bank_desc: str, bank_amount: str,
                   threshold: int, date_col: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Fuzzy-Match Bank Reconciler")
    print(f"{'='*55}")
    print(f"  Ledger:     {os.path.basename(ledger_file)}")
    print(f"  Bank stmt:  {os.path.basename(bank_file)}")
    print(f"  Threshold:  {threshold}% match confidence")
    print(f"  As of:      {datetime.now().strftime('%m/%d/%Y')}")
    print(f"{'='*55}\n")

    for f in [ledger_file, bank_file]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)

    ledger_df = pd.read_excel(ledger_file)
    bank_df   = pd.read_excel(bank_file)
    ledger_df.columns = ledger_df.columns.str.strip()
    bank_df.columns   = bank_df.columns.str.strip()

    for col, df, name in [(ledger_desc, ledger_df, 'Ledger'), (ledger_amount, ledger_df, 'Ledger'),
                          (bank_desc, bank_df, 'Bank'), (bank_amount, bank_df, 'Bank')]:
        if col not in df.columns:
            print(f"ERROR: Column '{col}' not found in {name} file.")
            print(f"Available: {', '.join(df.columns.tolist())}")
            sys.exit(1)

    ledger_df[ledger_amount] = pd.to_numeric(ledger_df[ledger_amount], errors='coerce')
    bank_df[bank_amount]     = pd.to_numeric(bank_df[bank_amount], errors='coerce')

    print(f"Ledger records:  {len(ledger_df)}")
    print(f"Bank records:    {len(bank_df)}")
    print(f"\nMatching... this may take a moment for large files.")

    bank_descriptions = bank_df[bank_desc].fillna("").astype(str).tolist()

    results = []
    ledger_df['_matched'] = False
    bank_df['_matched']   = False

    for i, ledger_row in ledger_df.iterrows():
        l_desc   = str(ledger_row[ledger_desc]) if pd.notna(ledger_row[ledger_desc]) else ""
        l_amount = ledger_row[ledger_amount]

        best_match = None
        best_score = 0
        best_bank_idx = None

        for j, bank_row in bank_df[~bank_df['_matched']].iterrows():
            b_desc   = str(bank_row[bank_desc]) if pd.notna(bank_row[bank_desc]) else ""
            b_amount = bank_row[bank_amount]

            desc_score = fuzz.token_sort_ratio(l_desc.lower(), b_desc.lower())
            amount_match = pd.notna(l_amount) and pd.notna(b_amount) and abs(l_amount - b_amount) < 0.01

            combined_score = desc_score * (1.2 if amount_match else 0.8)

            if combined_score > best_score and desc_score >= threshold:
                best_score = combined_score
                best_match = bank_row
                best_bank_idx = j

        if best_match is not None:
            ledger_df.at[i, '_matched'] = True
            bank_df.at[best_bank_idx, '_matched'] = True
            status = "MATCHED" if best_score >= threshold * 1.2 else "FUZZY MATCH"
            results.append({
                'Ledger_Desc': l_desc,
                'Ledger_Amount': l_amount,
                'Bank_Desc': best_match[bank_desc],
                'Bank_Amount': best_match[bank_amount],
                'Match_Score': round(best_score, 1),
                'Status': status
            })
        else:
            results.append({
                'Ledger_Desc': l_desc,
                'Ledger_Amount': l_amount,
                'Bank_Desc': '',
                'Bank_Amount': None,
                'Match_Score': 0,
                'Status': 'UNMATCHED IN LEDGER'
            })

    # Unmatched bank items
    unmatched_bank = bank_df[~bank_df['_matched']]

    matched_count = sum(1 for r in results if r['Status'] in ('MATCHED', 'FUZZY MATCH'))
    unmatched_ledger = sum(1 for r in results if r['Status'] == 'UNMATCHED IN LEDGER')

    print(f"\nReconciliation Results:")
    print(f"  Matched:              {matched_count}")
    print(f"  Unmatched in Ledger:  {unmatched_ledger}")
    print(f"  Unmatched in Bank:    {len(unmatched_bank)}")

    ledger_total  = ledger_df[ledger_amount].sum()
    bank_total    = bank_df[bank_amount].sum()
    difference    = round(ledger_total - bank_total, 2)

    print(f"\n  Ledger Total:  ${ledger_total:>12,.2f}")
    print(f"  Bank Total:    ${bank_total:>12,.2f}")
    print(f"  Difference:    ${difference:>12,.2f}")

    out_dir = os.path.dirname(ledger_file)
    output_path = os.path.join(out_dir, "BANK_RECONCILIATION.xlsx")

    fill_matched  = PatternFill("solid", fgColor="C6EFCE")
    fill_fuzzy    = PatternFill("solid", fgColor="FFEB9C")
    fill_unmatched= PatternFill("solid", fgColor="FFC7CE")
    fill_header   = PatternFill("solid", fgColor="1F497D")
    font_header   = Font(bold=True, color="FFFFFF")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Summary
        summary_data = {
            'Item': ['Ledger Records', 'Bank Records', 'Matched', 'Unmatched (Ledger)',
                     'Unmatched (Bank)', 'Ledger Total', 'Bank Total', 'Difference'],
            'Value': [len(ledger_df), len(bank_df), matched_count, unmatched_ledger,
                      len(unmatched_bank), ledger_total, bank_total, difference]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

        # Match results
        results_df = pd.DataFrame(results)
        results_df.to_excel(writer, sheet_name='Match Results', index=False)
        ws = writer.sheets['Match Results']
        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header
        status_col = results_df.columns.get_loc('Status') + 1
        for row in ws.iter_rows(min_row=2):
            status_val = row[status_col - 1].value
            if status_val == 'MATCHED':
                fill = fill_matched
            elif status_val == 'FUZZY MATCH':
                fill = fill_fuzzy
            else:
                fill = fill_unmatched
            for cell in row:
                cell.fill = fill

        # Unmatched bank items
        if not unmatched_bank.empty:
            unmatched_bank.drop(columns=['_matched']).to_excel(
                writer, sheet_name='Unmatched Bank Items', index=False)

        for sht in writer.sheets.values():
            for col in sht.columns:
                max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                sht.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

    print(f"\n{'='*55}")
    print(f"  DONE! Report saved to: {output_path}")
    print(f"  Green=Matched | Yellow=Fuzzy | Red=Unmatched")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Fuzzy-Match Bank Reconciler')
    parser.add_argument('ledger', help='Path to the ledger/GL Excel file')
    parser.add_argument('bank', help='Path to the bank statement Excel file')
    parser.add_argument('--desc', default='Description',
                        help='Ledger description column (default: Description)')
    parser.add_argument('--amount', default='Amount',
                        help='Ledger amount column (default: Amount)')
    parser.add_argument('--bank-desc', default='Description',
                        help='Bank description column (default: Description)')
    parser.add_argument('--bank-amount', default='Amount',
                        help='Bank amount column (default: Amount)')
    parser.add_argument('--threshold', type=int, default=70,
                        help='Minimum match score 0-100 (default: 70)')
    args = parser.parse_args()
    bank_reconcile(args.ledger, args.bank, args.desc, args.amount,
                   args.bank_desc, args.bank_amount, args.threshold)


if __name__ == '__main__':
    main()

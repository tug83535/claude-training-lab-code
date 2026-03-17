"""
KBT Universal Tools — AR/AP Aging Report
============================================================
Generates an aging report from any Excel file containing
invoice/transaction data with a date column and amount column.

Buckets: Current | 0-30 Days | 31-60 Days | 61-90 Days | 90+ Days

Usage:
    python aging_report.py "C:\\path\\to\\invoices.xlsx"
    python aging_report.py "invoices.xlsx" --date "Due Date" --amount "Balance" --name "Vendor"
    python aging_report.py "invoices.xlsx" --type AR --sheet "Invoices"

Output: Saves "AGING_REPORT.xlsx" in the same folder as the input file
"""

import sys
import os
import argparse
from datetime import datetime, date

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Required packages not installed.")
    print("Run: pip install pandas openpyxl")
    sys.exit(1)


BUCKETS = [
    ("Current (Not Yet Due)", lambda d: d < 0),
    ("0 - 30 Days",           lambda d: 0 <= d <= 30),
    ("31 - 60 Days",          lambda d: 31 <= d <= 60),
    ("61 - 90 Days",          lambda d: 61 <= d <= 90),
    ("90+ Days",              lambda d: d > 90),
]

BUCKET_COLORS = {
    "Current (Not Yet Due)": "C6EFCE",
    "0 - 30 Days":           "FFFFFF",
    "31 - 60 Days":          "FFEB9C",
    "61 - 90 Days":          "FFC7CE",
    "90+ Days":              "FF0000",
}


def generate_aging_report(file_path: str, date_col: str, amount_col: str,
                           name_col: str, report_type: str, sheet_name: str = None) -> None:
    print(f"\n{'='*55}")
    print(f"  KBT {report_type} Aging Report Generator")
    print(f"{'='*55}")
    print(f"  File:        {os.path.basename(file_path)}")
    print(f"  Date column: '{date_col}'")
    print(f"  Amount col:  '{amount_col}'")
    print(f"  Name col:    '{name_col}'")
    print(f"  As of date:  {datetime.now().strftime('%m/%d/%Y')}")
    print(f"{'='*55}\n")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
        df.columns = df.columns.str.strip()
    except Exception as e:
        print(f"ERROR reading file: {e}")
        sys.exit(1)

    for col in [date_col, amount_col]:
        if col not in df.columns:
            print(f"ERROR: Column '{col}' not found.")
            print(f"Available columns: {', '.join(df.columns.tolist())}")
            sys.exit(1)

    if name_col not in df.columns:
        print(f"WARNING: Name column '{name_col}' not found — proceeding without grouping by name.")
        name_col = None

    today = pd.Timestamp(date.today())
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df[amount_col] = pd.to_numeric(df[amount_col], errors='coerce')

    invalid_dates = df[date_col].isna().sum()
    if invalid_dates > 0:
        print(f"WARNING: {invalid_dates} row(s) had unreadable dates — excluded from report.")

    df = df.dropna(subset=[date_col, amount_col])
    df['Days_Outstanding'] = (today - df[date_col]).dt.days
    df['Aging_Bucket'] = df['Days_Outstanding'].apply(assign_bucket)

    print(f"Total records: {len(df)}")
    print(f"Total amount:  ${df[amount_col].sum():,.2f}\n")

    output_path = os.path.join(os.path.dirname(file_path), f"{report_type}_AGING_REPORT.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

        # Sheet 1: Detail
        df_out = df.copy()
        df_out.to_excel(writer, sheet_name='Detail', index=False)
        ws_detail = writer.sheets['Detail']
        _style_header(ws_detail)
        _autofit(ws_detail)

        # Sheet 2: Bucket Summary
        summary_rows = []
        for bucket_name, _ in BUCKETS:
            bucket_df = df[df['Aging_Bucket'] == bucket_name]
            summary_rows.append({
                'Aging Bucket': bucket_name,
                'Invoice Count': len(bucket_df),
                'Total Amount': bucket_df[amount_col].sum(),
                '% of Total': bucket_df[amount_col].sum() / df[amount_col].sum()
                               if df[amount_col].sum() != 0 else 0
            })

        summary_df = pd.DataFrame(summary_rows)
        totals = pd.DataFrame([{
            'Aging Bucket': 'TOTAL',
            'Invoice Count': summary_df['Invoice Count'].sum(),
            'Total Amount': summary_df['Total Amount'].sum(),
            '% of Total': 1.0
        }])
        summary_df = pd.concat([summary_df, totals], ignore_index=True)
        summary_df.to_excel(writer, sheet_name='Bucket Summary', index=False)

        ws_sum = writer.sheets['Bucket Summary']
        _style_header(ws_sum)

        for row_idx, (_, row_data) in enumerate(summary_df.iterrows(), start=2):
            bucket = row_data['Aging Bucket']
            color = BUCKET_COLORS.get(bucket, "FFFFFF")
            for col_idx in range(1, 5):
                ws_sum.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor=color)

        # Format amount and percent columns
        for row in ws_sum.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '$#,##0.00'
        for row in ws_sum.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = '0.0%'

        _autofit(ws_sum)

        # Sheet 3: By Name (if name column exists)
        if name_col:
            name_pivot = df.groupby([name_col, 'Aging_Bucket'])[amount_col].sum().unstack(fill_value=0)
            name_pivot['Total'] = name_pivot.sum(axis=1)
            name_pivot = name_pivot.sort_values('Total', ascending=False)
            name_pivot.to_excel(writer, sheet_name=f'By {name_col}')
            ws_name = writer.sheets[f'By {name_col}']
            _style_header(ws_name)
            _autofit(ws_name)

    print(f"\nAging Summary:")
    for bucket_name, _ in BUCKETS:
        bucket_total = df[df['Aging_Bucket'] == bucket_name][amount_col].sum()
        pct = bucket_total / df[amount_col].sum() * 100 if df[amount_col].sum() != 0 else 0
        print(f"  {bucket_name:<25} ${bucket_total:>12,.2f}  ({pct:.1f}%)")

    print(f"\n{'='*55}")
    print(f"  DONE! Report saved to:")
    print(f"  {output_path}")
    print(f"{'='*55}\n")


def assign_bucket(days: int) -> str:
    for bucket_name, condition in BUCKETS:
        if condition(days):
            return bucket_name
    return "90+ Days"


def _style_header(ws):
    fill = PatternFill("solid", fgColor="1F497D")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)


def main():
    parser = argparse.ArgumentParser(description='KBT AR/AP Aging Report Generator')
    parser.add_argument('file', help='Path to the Excel file with invoice data')
    parser.add_argument('--date', default='Date', help='Name of the date column (default: "Date")')
    parser.add_argument('--amount', default='Amount', help='Name of the amount column (default: "Amount")')
    parser.add_argument('--name', default='Vendor', help='Name of vendor/customer column (default: "Vendor")')
    parser.add_argument('--type', default='AP', choices=['AP', 'AR'],
                        help='Report type — AP or AR (default: AP)')
    parser.add_argument('--sheet', default=None, help='Sheet name to read (default: first sheet)')
    args = parser.parse_args()
    generate_aging_report(args.file, args.date, args.amount, args.name, args.type, args.sheet)


if __name__ == '__main__':
    main()

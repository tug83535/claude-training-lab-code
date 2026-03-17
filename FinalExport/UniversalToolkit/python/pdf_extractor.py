"""
KBT Universal Tools — PDF Tabular Extractor
============================================================
Scans PDF documents and extracts tables directly into Excel.
Solves the "my report only comes as a PDF" problem.

Requires: pip install pdfplumber (in addition to pandas/openpyxl)

Usage:
    python pdf_extractor.py "C:\\path\\report.pdf"
    python pdf_extractor.py "report.pdf" --pages "1-5"
    python pdf_extractor.py "report.pdf" --all-pages

Output: Saves "PDF_EXTRACTED_TABLES.xlsx" with one sheet per table found
"""

import sys
import os
import argparse
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber not installed.")
    print("Run: pip install pdfplumber")
    sys.exit(1)


def extract_pdf_tables(file_path: str, page_range: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT PDF Table Extractor")
    print(f"{'='*55}")
    print(f"  File: {os.path.basename(file_path)}")
    print(f"  Date: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}")
    print(f"{'='*55}\n")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    if not file_path.lower().endswith('.pdf'):
        print("ERROR: File must be a PDF (.pdf extension).")
        sys.exit(1)

    try:
        pdf = pdfplumber.open(file_path)
    except Exception as e:
        print(f"ERROR opening PDF: {e}")
        sys.exit(1)

    total_pages = len(pdf.pages)
    print(f"PDF pages: {total_pages}")

    # Determine which pages to scan
    if page_range:
        try:
            if '-' in page_range:
                start, end = page_range.split('-')
                pages_to_scan = list(range(int(start) - 1, int(end)))
            else:
                pages_to_scan = [int(page_range) - 1]
        except ValueError:
            print(f"ERROR: Invalid page range format. Use '1-5' or '3'.")
            sys.exit(1)
    else:
        pages_to_scan = list(range(total_pages))

    pages_to_scan = [p for p in pages_to_scan if 0 <= p < total_pages]
    print(f"Scanning pages: {', '.join(str(p+1) for p in pages_to_scan)}\n")

    all_tables = []
    table_count = 0

    for page_num in pages_to_scan:
        page = pdf.pages[page_num]
        tables = page.extract_tables()
        if tables:
            for t_idx, table in enumerate(tables):
                if table and len(table) > 1:
                    df = pd.DataFrame(table[1:], columns=table[0])
                    df = df.dropna(how='all').dropna(axis=1, how='all')
                    if len(df) > 0:
                        all_tables.append({
                            'page': page_num + 1,
                            'table_idx': t_idx + 1,
                            'df': df
                        })
                        table_count += 1
                        print(f"  Page {page_num + 1}, Table {t_idx + 1}: {len(df)} rows x {len(df.columns)} columns")

    pdf.close()

    if not all_tables:
        print("\nNo tables found in the specified pages.")
        print("Note: This tool works best on PDFs with structured grid tables.")
        print("Scanned/image PDFs may not work — the text must be selectable.")
        sys.exit(0)

    output_path = os.path.join(os.path.dirname(file_path), "PDF_EXTRACTED_TABLES.xlsx")

    fill_header = PatternFill("solid", fgColor="1F497D")
    font_header = Font(bold=True, color="FFFFFF")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for t in all_tables:
            sheet_name = f"Page{t['page']}_Table{t['table_idx']}"
            t['df'].to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    print(f"\n{'='*55}")
    print(f"  DONE! {table_count} table(s) extracted.")
    print(f"  Saved to: {output_path}")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT PDF Table Extractor')
    parser.add_argument('file', help='Path to the PDF file')
    parser.add_argument('--pages', default=None,
                        help='Page range to scan (e.g. "1-5" or "3"). Default: all pages.')
    args = parser.parse_args()
    extract_pdf_tables(args.file, args.pages)


if __name__ == '__main__':
    main()

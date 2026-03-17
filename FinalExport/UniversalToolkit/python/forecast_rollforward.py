"""
KBT Universal Tools — Forecast Roll-Forward
============================================================
Takes last month's actuals plus growth assumptions and builds
the next 12-month rolling forecast.

Methods available:
  - moving_avg: 3-month moving average
  - growth:     Apply a fixed % growth rate per period
  - flat:       Repeat last period's values

Usage:
    python forecast_rollforward.py "C:\\path\\data.xlsx" --method moving_avg
    python forecast_rollforward.py "data.xlsx" --value-col "Revenue" --method growth --rate 0.05

Output: Saves "FORECAST_ROLLFORWARD.xlsx" with actuals + forecast
"""

import sys
import os
import argparse
from datetime import datetime, date
from dateutil.relativedelta import relativedelta

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.chart import LineChart, Reference
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl python-dateutil")
    sys.exit(1)


def forecast_rollforward(file_path: str, date_col: str, value_col: str,
                          method: str, growth_rate: float, periods: int,
                          sheet_name: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Forecast Roll-Forward")
    print(f"{'='*55}")
    print(f"  File:    {os.path.basename(file_path)}")
    print(f"  Values:  '{value_col}' | Method: {method}")
    print(f"  Periods: {periods} months forward")
    print(f"{'='*55}\n")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
    df.columns = df.columns.str.strip()

    for col in [date_col, value_col]:
        if col not in df.columns:
            print(f"ERROR: Column '{col}' not found. Available: {', '.join(df.columns.tolist())}")
            sys.exit(1)

    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df[value_col] = pd.to_numeric(df[value_col], errors='coerce')
    df = df.dropna(subset=[date_col, value_col]).sort_values(date_col)

    print(f"Actuals loaded: {len(df)} periods")
    print(f"Last actual period: {df[date_col].max().strftime('%B %Y')}\n")

    last_date = df[date_col].max()
    actual_values = df[value_col].tolist()

    forecast_dates = []
    forecast_values = []

    for i in range(1, periods + 1):
        next_date = last_date + relativedelta(months=i)
        forecast_dates.append(next_date)

        if method == 'moving_avg':
            lookback = min(3, len(actual_values) + len(forecast_values))
            all_vals = actual_values + forecast_values
            val = sum(all_vals[-lookback:]) / lookback
        elif method == 'growth':
            all_vals = actual_values + forecast_values
            val = all_vals[-1] * (1 + growth_rate)
        else:  # flat
            all_vals = actual_values + forecast_values
            val = all_vals[-1]

        forecast_values.append(round(val, 2))

    forecast_df = pd.DataFrame({
        date_col: forecast_dates,
        value_col: forecast_values,
        'Type': 'Forecast'
    })

    df['Type'] = 'Actual'
    combined = pd.concat([df[[date_col, value_col, 'Type']], forecast_df], ignore_index=True)

    output_path = os.path.join(os.path.dirname(file_path), "FORECAST_ROLLFORWARD.xlsx")

    fill_actual   = PatternFill("solid", fgColor="DEEAF1")
    fill_forecast = PatternFill("solid", fgColor="E2EFDA")
    fill_header   = PatternFill("solid", fgColor="1F497D")
    font_header   = Font(bold=True, color="FFFFFF")
    font_bold     = Font(bold=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        combined.to_excel(writer, sheet_name='Forecast', index=False)
        ws = writer.sheets['Forecast']

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header

        for row in ws.iter_rows(min_row=2):
            type_val = row[2].value
            fill = fill_actual if type_val == 'Actual' else fill_forecast
            for cell in row:
                cell.fill = fill

        # Format value column
        val_col_letter = 'B'
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = '#,##0.00'

        # Add line chart
        chart = LineChart()
        chart.title = f"{value_col} — Actuals vs Forecast"
        chart.y_axis.title = value_col
        chart.x_axis.title = "Period"
        chart.style = 10

        data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, "E2")

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 20)

    print(f"Forecast Summary ({method} method):")
    for fd, fv in zip(forecast_dates, forecast_values):
        print(f"  {fd.strftime('%B %Y'):15}  ${fv:>12,.2f}")

    print(f"\n{'='*55}")
    print(f"  DONE! Saved to: {output_path}")
    print(f"  Blue=Actual | Green=Forecast")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Forecast Roll-Forward')
    parser.add_argument('file', help='Path to Excel file with historical data')
    parser.add_argument('--date', default='Date', help='Date column name (default: Date)')
    parser.add_argument('--value-col', default='Amount', help='Value column name (default: Amount)')
    parser.add_argument('--method', default='moving_avg',
                        choices=['moving_avg', 'growth', 'flat'],
                        help='Forecast method (default: moving_avg)')
    parser.add_argument('--rate', type=float, default=0.03,
                        help='Growth rate for "growth" method, e.g. 0.05 for 5% (default: 0.03)')
    parser.add_argument('--periods', type=int, default=12,
                        help='Number of months to forecast (default: 12)')
    parser.add_argument('--sheet', default=None, help='Sheet name to read')
    args = parser.parse_args()
    forecast_rollforward(args.file, args.date, args.value_col, args.method,
                         args.rate, args.periods, args.sheet)


if __name__ == '__main__':
    main()

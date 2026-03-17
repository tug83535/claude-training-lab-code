#!/usr/bin/env python3
"""
pnl_monte_carlo.py — Monte Carlo P&L Risk Simulation
=====================================================
iPipeline Finance & Accounting | Keystone BenefitTech P&L Toolkit

PURPOSE:
    Runs N iterations (default 10,000) of the P&L model with randomized inputs
    to produce full probability distributions for key financial outcomes.
    Answers the question: "Given real uncertainty in our expenses and allocations,
    what is the realistic range of outcomes for our P&L?"

WHAT IS RANDOMIZED EACH ITERATION:
    1. Revenue allocation shares — drawn from a Dirichlet distribution centered
       on the configured shares in pnl_config.py. They always sum to exactly 100%.
    2. Expense amounts — drawn from a Normal distribution around actual GL values.
       Standard deviation is derived from the observed month-to-month coefficient
       of variation in the GL data (how volatile each expense category actually is).
    3. Optional: expense shock events — low-probability, high-impact cost spikes
       to stress-test the P&L against unexpected events.

OUTPUTS:
    - Console: percentile table (P5/P25/P50/P75/P95), best/base/worst summary,
               Value at Risk, top variance drivers
    - Excel:   formatted multi-sheet workbook with all results, charts, and inputs
    - Chart:   histogram of total spend distribution across all simulations

USAGE:
    python pnl_monte_carlo.py                          # 10,000 iterations, default file
    python pnl_monte_carlo.py --iterations 50000       # higher precision
    python pnl_monte_carlo.py --export results.xlsx    # save to Excel
    python pnl_monte_carlo.py --shock-prob 0.05        # 5% chance of expense shock
    python pnl_monte_carlo.py --seed 42                # reproducible results

DEPENDENCIES:
    numpy, pandas, matplotlib, openpyxl (all standard in the toolkit)
    scipy (optional — used for advanced distribution fitting; falls back gracefully)
"""

import os
import sys
import argparse
import warnings
import time
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")   # Non-interactive backend — works without a display
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

warnings.filterwarnings("ignore", category=UserWarning)

try:
    from pnl_config import (
        PnLBase, PRODUCTS, REVENUE_SHARES, AWS_COMPUTE_SHARES,
        DEPARTMENTS, SOURCE_FILE, OUTPUT_DIR, CHART_DIR,
        COLORS, PRODUCT_COLORS, FY_LABEL, APP_NAME, APP_VERSION,
        VARIANCE_PCT, format_currency, format_pct
    )
except ImportError:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from pnl_config import (
        PnLBase, PRODUCTS, REVENUE_SHARES, AWS_COMPUTE_SHARES,
        DEPARTMENTS, SOURCE_FILE, OUTPUT_DIR, CHART_DIR,
        COLORS, PRODUCT_COLORS, FY_LABEL, APP_NAME, APP_VERSION,
        VARIANCE_PCT, format_currency, format_pct
    )

# Optional scipy for advanced distribution fitting
try:
    from scipy import stats as scipy_stats
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False


# =============================================================================
# SIMULATION CONFIGURATION DEFAULTS
# =============================================================================

DEFAULT_ITERATIONS    = 10_000   # Number of Monte Carlo iterations
DEFAULT_SEED          = None     # Random seed (None = different every run)
DEFAULT_SHARE_CONC    = 10.0     # Dirichlet concentration (higher = tighter spread)
DEFAULT_EXPENSE_CV    = 0.12     # Default coefficient of variation for expenses (12%)
DEFAULT_SHOCK_PROB    = 0.0      # Probability of an expense shock event per iteration
DEFAULT_SHOCK_SIZE    = 0.25     # Size of shock as fraction of total spend (25%)

PERCENTILES           = [5, 25, 50, 75, 95]
PERCENTILE_LABELS     = ["P5 (Best)", "P25", "P50 (Base)", "P75", "P95 (Worst)"]


# =============================================================================
# MONTE CARLO SIMULATION ENGINE
# =============================================================================

class MonteCarloSimulator(PnLBase):
    """
    Monte Carlo P&L risk simulator.

    Randomizes revenue allocation shares and expense amounts across N iterations
    to produce full probability distributions for total spend, per-product costs,
    and implied margin. Outputs percentile tables, risk metrics, and charts.
    """

    def __init__(
        self,
        file_path:   str   = None,
        iterations:  int   = DEFAULT_ITERATIONS,
        seed:        int   = DEFAULT_SEED,
        share_conc:  float = DEFAULT_SHARE_CONC,
        expense_cv:  float = DEFAULT_EXPENSE_CV,
        shock_prob:  float = DEFAULT_SHOCK_PROB,
        shock_size:  float = DEFAULT_SHOCK_SIZE,
        verbose:     bool  = True,
    ):
        super().__init__(verbose=verbose)
        self.file_path   = file_path or SOURCE_FILE
        self.iterations  = iterations
        self.seed        = seed
        self.share_conc  = share_conc
        self.expense_cv  = expense_cv
        self.shock_prob  = shock_prob
        self.shock_size  = shock_size

        self.rng         = np.random.default_rng(seed)
        self.gl          = None         # Raw GL DataFrame
        self.baseline    = {}           # Baseline metrics from actual GL
        self.results     = None         # Full simulation results DataFrame
        self.summary     = None         # Percentile summary table

    # -------------------------------------------------------------------------
    # STEP 1: Load and characterize the GL data
    # -------------------------------------------------------------------------

    def load(self) -> "MonteCarloSimulator":
        """Load the GL data and compute baseline statistics for simulation."""
        self._section("LOADING & CHARACTERIZING GL DATA")

        self.gl = self._load_gl(self.file_path)

        # --- Baseline: total annual spend ---
        total_spend = self.gl["Abs_Amount"].sum()
        self._print(f"Total GL spend (baseline):  {format_currency(total_spend)}")

        # --- Baseline: spend by product ---
        product_spend = (
            self.gl[self.gl["Product"].isin(PRODUCTS)]
            .groupby("Product")["Abs_Amount"]
            .sum()
            .reindex(PRODUCTS, fill_value=0)
        )

        # --- Baseline: spend by expense category ---
        category_spend = self.gl.groupby("Expense Category")["Abs_Amount"].sum()

        # --- Monthly volatility: compute CV per product from month-to-month data ---
        monthly_product = (
            self.gl[self.gl["Product"].isin(PRODUCTS)]
            .groupby(["Product", "Month"])["Abs_Amount"]
            .sum()
            .reset_index()
        )

        product_cv = {}
        for product in PRODUCTS:
            monthly = monthly_product[monthly_product["Product"] == product]["Abs_Amount"]
            if len(monthly) >= 2 and monthly.mean() > 0:
                cv = monthly.std() / monthly.mean()
                product_cv[product] = max(0.03, min(cv, 0.40))  # Clamp: 3%-40%
            else:
                product_cv[product] = self.expense_cv

        # --- Store baseline ---
        self.baseline = {
            "total_spend":    total_spend,
            "product_spend":  product_spend.to_dict(),
            "category_spend": category_spend.to_dict(),
            "product_cv":     product_cv,
            "months_present": self.gl["Month"].nunique(),
        }

        self._print(f"Products found:             {', '.join(PRODUCTS)}")
        self._print(f"Months with data:           {self.baseline['months_present']}")
        self._print(f"Expense categories:         {len(category_spend)}")
        self._print("")
        self._print("Spend by product (baseline):", "INFO")
        for p in PRODUCTS:
            amt   = product_spend.get(p, 0)
            share = REVENUE_SHARES.get(p, 0)
            cv    = product_cv.get(p, self.expense_cv)
            self._print(f"  {p:<15s}  {format_currency(amt):>12s}  "
                        f"({format_pct(share)} share, CV={cv:.1%})")

        return self

    # -------------------------------------------------------------------------
    # STEP 2: Run the simulation
    # -------------------------------------------------------------------------

    def run(self) -> "MonteCarloSimulator":
        """
        Execute the Monte Carlo simulation.
        Returns self so calls can be chained: sim.load().run().export()
        """
        if self.gl is None:
            raise RuntimeError("Call .load() before .run()")

        self._section(f"RUNNING {self.iterations:,} MONTE CARLO ITERATIONS")
        t_start = time.time()

        # --- Dirichlet concentration vector ---
        # Multiplying configured shares by share_conc gives the Dirichlet alpha
        # parameters. Higher share_conc = tighter distribution around configured
        # shares. Default of 10 allows ~±8% swing at one standard deviation.
        base_shares = np.array([REVENUE_SHARES[p] for p in PRODUCTS])
        alpha       = base_shares * self.share_conc  # Dirichlet alphas

        # --- Storage arrays (pre-allocated for performance) ---
        n = self.iterations
        total_spends          = np.zeros(n)
        product_spends        = np.zeros((n, len(PRODUCTS)))
        allocated_spends      = np.zeros((n, len(PRODUCTS)))
        share_draws           = np.zeros((n, len(PRODUCTS)))
        had_shock             = np.zeros(n, dtype=bool)

        baseline_total = self.baseline["total_spend"]
        product_cvs    = np.array([self.baseline["product_cv"][p] for p in PRODUCTS])
        product_base   = np.array([self.baseline["product_spend"].get(p, 0)
                                   for p in PRODUCTS])

        self._print(f"Share concentration (α):    {self.share_conc}")
        self._print(f"Expense volatility (CV):    {self.expense_cv:.1%} default")
        self._print(f"Shock probability:          {self.shock_prob:.1%}")
        if self.shock_prob > 0:
            self._print(f"Shock size (if triggered):  ±{self.shock_size:.1%} of total spend")
        self._print("")

        # --- Main simulation loop ---
        for i in range(n):

            # 1. Draw revenue allocation shares from Dirichlet distribution
            #    These always sum to exactly 1.0
            shares = self.rng.dirichlet(alpha)
            share_draws[i] = shares

            # 2. Draw total expense amount (normal distribution around baseline)
            #    CV drives the standard deviation: higher CV = more volatile
            sigma         = baseline_total * self.expense_cv
            total_sim     = self.rng.normal(loc=baseline_total, scale=sigma)
            total_sim     = max(total_sim, baseline_total * 0.4)  # floor at 40%

            # 3. Apply expense shock (rare but impactful events)
            if self.shock_prob > 0 and self.rng.random() < self.shock_prob:
                shock_factor  = 1.0 + self.rng.uniform(
                    self.shock_size * 0.5, self.shock_size * 1.5
                )
                total_sim    *= shock_factor
                had_shock[i]  = True

            total_spends[i] = total_sim

            # 4. Draw per-product spend (normal, centered on product baseline)
            #    Each product gets its own volatility (derived from its own CV)
            for j, product in enumerate(PRODUCTS):
                p_sigma            = product_base[j] * product_cvs[j]
                p_draw             = self.rng.normal(loc=product_base[j], scale=p_sigma)
                product_spends[i, j] = max(p_draw, 0)  # Non-negative spend

                # 5. Allocated spend = product spend weighted by the drawn revenue share
                allocated_spends[i, j] = total_sim * shares[j]

        # --- Build results DataFrame ---
        cols_raw  = [f"{p}_RawSpend"   for p in PRODUCTS]
        cols_allc = [f"{p}_AllocSpend" for p in PRODUCTS]
        cols_shr  = [f"{p}_Share"      for p in PRODUCTS]

        self.results = pd.DataFrame({
            "TotalSpend": total_spends,
            "HadShock":   had_shock,
            **{cols_raw[j]:  product_spends[:, j]   for j in range(len(PRODUCTS))},
            **{cols_allc[j]: allocated_spends[:, j] for j in range(len(PRODUCTS))},
            **{cols_shr[j]:  share_draws[:, j]       for j in range(len(PRODUCTS))},
        })

        elapsed = time.time() - t_start
        self._print(f"Simulation complete:  {n:,} iterations in {elapsed:.2f}s", "OK")
        shock_count = had_shock.sum()
        if self.shock_prob > 0:
            self._print(f"Shock events:         {shock_count:,} "
                        f"({shock_count/n:.1%} of iterations)", "INFO")

        self._build_summary()
        return self

    # -------------------------------------------------------------------------
    # STEP 3: Build the percentile summary table
    # -------------------------------------------------------------------------

    def _build_summary(self):
        """Compute percentile tables from simulation results."""
        rows = []

        # Total spend percentiles
        ts = self.results["TotalSpend"]
        row = {"Metric": "Total Annual Spend", "Baseline": self.baseline["total_spend"]}
        for pct, label in zip(PERCENTILES, PERCENTILE_LABELS):
            row[label] = np.percentile(ts, pct)
        row["Mean"]   = ts.mean()
        row["StdDev"] = ts.std()
        row["CV"]     = ts.std() / ts.mean() if ts.mean() > 0 else 0
        rows.append(row)

        # Per-product allocated spend percentiles
        for product in PRODUCTS:
            col  = f"{product}_AllocSpend"
            vals = self.results[col]
            row  = {
                "Metric":   f"  {product} — Allocated Spend",
                "Baseline": self.baseline["product_spend"].get(product, 0),
            }
            for pct, label in zip(PERCENTILES, PERCENTILE_LABELS):
                row[label] = np.percentile(vals, pct)
            row["Mean"]   = vals.mean()
            row["StdDev"] = vals.std()
            row["CV"]     = vals.std() / vals.mean() if vals.mean() > 0 else 0
            rows.append(row)

        # Per-product revenue share percentiles
        for product in PRODUCTS:
            col    = f"{product}_Share"
            vals   = self.results[col]
            config = REVENUE_SHARES.get(product, 0)
            row    = {
                "Metric":   f"  {product} — Revenue Share",
                "Baseline": config,
            }
            for pct, label in zip(PERCENTILES, PERCENTILE_LABELS):
                row[label] = np.percentile(vals, pct)
            row["Mean"]   = vals.mean()
            row["StdDev"] = vals.std()
            row["CV"]     = vals.std() / vals.mean() if vals.mean() > 0 else 0
            rows.append(row)

        self.summary = pd.DataFrame(rows).set_index("Metric")

    # -------------------------------------------------------------------------
    # STEP 4: Print results to console
    # -------------------------------------------------------------------------

    def print_results(self) -> "MonteCarloSimulator":
        """Print a clean, readable summary of simulation results."""
        if self.results is None:
            raise RuntimeError("Call .run() before .print_results()")

        ts = self.results["TotalSpend"]

        self._section("MONTE CARLO RESULTS SUMMARY")

        # --- Key statistics ---
        baseline = self.baseline["total_spend"]
        p5  = np.percentile(ts, 5)
        p50 = np.percentile(ts, 50)
        p95 = np.percentile(ts, 95)
        mean_spend = ts.mean()
        std_spend  = ts.std()

        print(f"\n  Iterations:          {self.iterations:>12,}")
        print(f"  Baseline (actual):   {format_currency(baseline):>12s}")
        print(f"  Mean (simulated):    {format_currency(mean_spend):>12s}  "
              f"({(mean_spend - baseline) / baseline:+.1%} vs baseline)")
        print(f"  Std Deviation:       {format_currency(std_spend):>12s}")
        print(f"  Coeff of Variation:  {std_spend / mean_spend:.1%}")
        print()

        # --- Best / Base / Worst case ---
        print(f"  {'SCENARIO':<22s}  {'TOTAL SPEND':>14s}  {'vs BASELINE':>12s}")
        print(f"  {'-'*22}  {'-'*14}  {'-'*12}")
        for label, val in [("Best Case (P5)",  p5),
                            ("Base Case (P50)", p50),
                            ("Worst Case (P95)", p95)]:
            delta_pct = (val - baseline) / baseline if baseline else 0
            delta_str = f"{delta_pct:+.1%}"
            print(f"  {label:<22s}  {format_currency(val):>14s}  {delta_str:>12s}")

        # --- Value at Risk (VaR) ---
        var_95 = p95 - p50
        print(f"\n  Value at Risk (95%):    {format_currency(var_95)}")
        print(f"  (Unexpected spend above base case at 95th percentile)")

        # --- Per-product summary ---
        print(f"\n  {'PRODUCT':<16s}  {'BASELINE':>12s}  {'P5':>12s}  "
              f"{'P50':>12s}  {'P95':>12s}  {'CV':>8s}")
        print(f"  {'-'*16}  {'-'*12}  {'-'*12}  {'-'*12}  {'-'*12}  {'-'*8}")

        for product in PRODUCTS:
            col  = f"{product}_AllocSpend"
            vals = self.results[col]
            base = self.baseline["product_spend"].get(product, 0)
            cv   = vals.std() / vals.mean() if vals.mean() > 0 else 0
            print(f"  {product:<16s}  {format_currency(base):>12s}  "
                  f"{format_currency(np.percentile(vals, 5)):>12s}  "
                  f"{format_currency(np.percentile(vals, 50)):>12s}  "
                  f"{format_currency(np.percentile(vals, 95)):>12s}  "
                  f"{cv:>8.1%}")

        # --- Share distribution ---
        print(f"\n  {'PRODUCT':<16s}  {'CONFIGURED':>12s}  "
              f"{'SIM MEAN':>10s}  {'SIM STD':>10s}  {'SIM RANGE':>20s}")
        print(f"  {'-'*16}  {'-'*12}  {'-'*10}  {'-'*10}  {'-'*20}")

        for product in PRODUCTS:
            col    = f"{product}_Share"
            vals   = self.results[col]
            config = REVENUE_SHARES.get(product, 0)
            lo, hi = np.percentile(vals, [5, 95])
            print(f"  {product:<16s}  {config:>12.1%}  "
                  f"{vals.mean():>10.1%}  {vals.std():>10.1%}  "
                  f"{lo:.1%} – {hi:.1%}")

        # --- Top variance drivers ---
        print(f"\n  TOP VARIANCE DRIVERS (by CV of allocated spend):")
        cvs = {}
        for product in PRODUCTS:
            col  = f"{product}_AllocSpend"
            vals = self.results[col]
            cvs[product] = vals.std() / vals.mean() if vals.mean() > 0 else 0
        for i, (product, cv) in enumerate(
            sorted(cvs.items(), key=lambda x: x[1], reverse=True), 1
        ):
            bar = "█" * int(cv * 100)
            print(f"  {i}. {product:<15s}  CV={cv:.1%}  {bar}")

        print()
        return self

    # -------------------------------------------------------------------------
    # STEP 5: Generate distribution chart
    # -------------------------------------------------------------------------

    def plot(self, save_path: str = None) -> str:
        """
        Generate a 4-panel distribution chart:
          - Top left:  Total spend histogram with percentile bands
          - Top right: Per-product allocated spend box plot
          - Bottom left: Revenue share distribution violin plot
          - Bottom right: Cumulative distribution (CDF) of total spend
        """
        if self.results is None:
            raise RuntimeError("Call .run() before .plot()")

        ts       = self.results["TotalSpend"]
        baseline = self.baseline["total_spend"]
        p5, p50, p95 = np.percentile(ts, [5, 50, 95])

        # --- Color palette ---
        navy   = COLORS.get("navy",  "#1F4E79")
        green  = COLORS.get("green", "#70AD47")
        red    = COLORS.get("red",   "#C00000")
        amber  = COLORS.get("amber", "#ED7D31")
        grey   = COLORS.get("grey",  "#808080")
        p_colors = [PRODUCT_COLORS.get(p, grey) for p in PRODUCTS]

        fig, axes = plt.subplots(2, 2, figsize=(16, 11))
        fig.suptitle(
            f"{APP_NAME} — Monte Carlo P&L Risk Simulation\n"
            f"{self.iterations:,} Iterations  |  {FY_LABEL}  |  "
            f"Seed: {self.seed if self.seed is not None else 'random'}",
            fontsize=14, fontweight="bold", color=navy, y=0.98
        )
        fig.patch.set_facecolor("#F8F9FA")

        # ── Panel 1: Total Spend Histogram ──────────────────────────────────
        ax1 = axes[0, 0]
        ax1.set_facecolor("#FFFFFF")

        n_bins = min(80, self.iterations // 100)
        ax1.hist(ts / 1e6, bins=n_bins, color=navy, alpha=0.75, edgecolor="white",
                 linewidth=0.3, label="Simulated spend")

        # Percentile shading
        ax1.axvspan(ts.min() / 1e6, p5 / 1e6,  alpha=0.15, color=green, label="P5 zone")
        ax1.axvspan(p95 / 1e6, ts.max() / 1e6, alpha=0.15, color=red,   label="P95 zone")

        # Vertical reference lines
        ax1.axvline(baseline / 1e6, color=amber, linewidth=2.0,
                    linestyle="--", label=f"Baseline: {format_currency(baseline)}")
        ax1.axvline(p50 / 1e6, color=navy, linewidth=2.0,
                    linestyle="-",  label=f"P50: {format_currency(p50)}")
        ax1.axvline(p5  / 1e6, color=green, linewidth=1.5,
                    linestyle=":",  label=f"P5:  {format_currency(p5)}")
        ax1.axvline(p95 / 1e6, color=red,   linewidth=1.5,
                    linestyle=":",  label=f"P95: {format_currency(p95)}")

        ax1.set_title("Total Annual Spend Distribution", fontsize=12, fontweight="bold",
                      color=navy, pad=8)
        ax1.set_xlabel("Total Spend ($M)", fontsize=10)
        ax1.set_ylabel("Frequency", fontsize=10)
        ax1.legend(fontsize=8, loc="upper right")
        ax1.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.1f}M"))
        ax1.grid(axis="y", alpha=0.3, linestyle="--")
        ax1.spines["top"].set_visible(False)
        ax1.spines["right"].set_visible(False)

        # ── Panel 2: Per-Product Box Plot ────────────────────────────────────
        ax2 = axes[0, 1]
        ax2.set_facecolor("#FFFFFF")

        box_data = [self.results[f"{p}_AllocSpend"].values / 1e6 for p in PRODUCTS]
        bp = ax2.boxplot(
            box_data,
            patch_artist=True,
            medianprops=dict(color="white", linewidth=2),
            whiskerprops=dict(linewidth=1.2),
            capprops=dict(linewidth=1.2),
            flierprops=dict(marker=".", markersize=2, alpha=0.3),
            widths=0.5,
        )
        for patch, color in zip(bp["boxes"], p_colors):
            patch.set_facecolor(color)
            patch.set_alpha(0.80)

        # Overlay baseline dots
        for j, product in enumerate(PRODUCTS, 1):
            base_m = self.baseline["product_spend"].get(product, 0) / 1e6
            ax2.scatter(j, base_m, color="white", s=60, zorder=5,
                        edgecolors=p_colors[j - 1], linewidth=2)

        ax2.set_title("Per-Product Allocated Spend Distribution", fontsize=12,
                      fontweight="bold", color=navy, pad=8)
        ax2.set_xlabel("Product", fontsize=10)
        ax2.set_ylabel("Allocated Spend ($M)", fontsize=10)
        ax2.set_xticklabels(PRODUCTS, fontsize=9)
        ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.1f}M"))
        ax2.grid(axis="y", alpha=0.3, linestyle="--")
        ax2.spines["top"].set_visible(False)
        ax2.spines["right"].set_visible(False)
        ax2.text(0.98, 0.02, "○ = baseline", transform=ax2.transAxes,
                 ha="right", va="bottom", fontsize=8, color=grey)

        # ── Panel 3: Revenue Share Violin Plot ───────────────────────────────
        ax3 = axes[1, 0]
        ax3.set_facecolor("#FFFFFF")

        share_data = [self.results[f"{p}_Share"].values * 100 for p in PRODUCTS]
        vp = ax3.violinplot(share_data, showmedians=True, showextrema=True)

        for j, (body, color) in enumerate(zip(vp["bodies"], p_colors)):
            body.set_facecolor(color)
            body.set_alpha(0.70)
        vp["cmedians"].set_color("white")
        vp["cmedians"].set_linewidth(2)
        vp["cmaxes"].set_color(grey)
        vp["cmins"].set_color(grey)
        vp["cbars"].set_color(grey)

        # Configured share markers
        for j, product in enumerate(PRODUCTS, 1):
            config_pct = REVENUE_SHARES.get(product, 0) * 100
            ax3.scatter(j, config_pct, color="white", s=70, zorder=5,
                        edgecolors=p_colors[j - 1], linewidth=2, marker="D")

        ax3.set_title("Revenue Share Distribution Across Simulations", fontsize=12,
                      fontweight="bold", color=navy, pad=8)
        ax3.set_xlabel("Product", fontsize=10)
        ax3.set_ylabel("Revenue Share (%)", fontsize=10)
        ax3.set_xticks(range(1, len(PRODUCTS) + 1))
        ax3.set_xticklabels(PRODUCTS, fontsize=9)
        ax3.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}%"))
        ax3.grid(axis="y", alpha=0.3, linestyle="--")
        ax3.spines["top"].set_visible(False)
        ax3.spines["right"].set_visible(False)
        ax3.text(0.98, 0.02, "◇ = configured share", transform=ax3.transAxes,
                 ha="right", va="bottom", fontsize=8, color=grey)

        # ── Panel 4: Cumulative Distribution (CDF) ───────────────────────────
        ax4 = axes[1, 1]
        ax4.set_facecolor("#FFFFFF")

        sorted_ts = np.sort(ts) / 1e6
        cdf_y     = np.arange(1, len(sorted_ts) + 1) / len(sorted_ts) * 100

        ax4.plot(sorted_ts, cdf_y, color=navy, linewidth=2, label="CDF")
        ax4.fill_between(sorted_ts, cdf_y, alpha=0.08, color=navy)

        # Percentile markers
        for pct, label_str, color in [
            (p5,  "P5",       green),
            (p50, "P50",      navy),
            (p95, "P95",      red),
        ]:
            ax4.axhline(PERCENTILES[PERCENTILE_LABELS.index(
                next(l for l in PERCENTILE_LABELS if str(PERCENTILES[
                    PERCENTILE_LABELS.index(l)]) in l and (
                    "5" in l or "50" in l or "95" in l
                ) and (
                    (pct == p5  and "Best" in l) or
                    (pct == p50 and "Base" in l) or
                    (pct == p95 and "Worst" in l)
                ))
            )] if False else pct / 1e6,  # simplified: use pct value directly
            color=color, linewidth=1, linestyle="--", alpha=0.6)

        # Cleaner approach for CDF percentile lines
        for pct_val, pct_num, color, lbl in [
            (p5,  5,  green, "P5"),
            (p50, 50, navy,  "P50"),
            (p95, 95, red,   "P95"),
        ]:
            ax4.axvline(pct_val / 1e6, color=color, linewidth=1.5,
                        linestyle="--", alpha=0.8)
            ax4.axhline(pct_num, color=color, linewidth=0.8,
                        linestyle=":", alpha=0.5)
            ax4.annotate(
                f"{lbl}\n{format_currency(pct_val)}",
                xy=(pct_val / 1e6, pct_num),
                xytext=(6, 0), textcoords="offset points",
                fontsize=7.5, color=color, va="center"
            )

        ax4.axvline(baseline / 1e6, color=amber, linewidth=2,
                    linestyle="-.", label=f"Baseline {format_currency(baseline)}")

        ax4.set_title("Cumulative Probability of Total Spend", fontsize=12,
                      fontweight="bold", color=navy, pad=8)
        ax4.set_xlabel("Total Spend ($M)", fontsize=10)
        ax4.set_ylabel("Probability (%)", fontsize=10)
        ax4.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.1f}M"))
        ax4.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}%"))
        ax4.set_ylim(0, 100)
        ax4.grid(alpha=0.3, linestyle="--")
        ax4.spines["top"].set_visible(False)
        ax4.spines["right"].set_visible(False)

        plt.tight_layout(rect=[0, 0, 1, 0.95])

        # Save
        if save_path is None:
            PnLBase.ensure_dir(CHART_DIR)
            ts_str    = datetime.now().strftime("%Y%m%d_%H%M%S")
            save_path = os.path.join(CHART_DIR, f"monte_carlo_{ts_str}.png")

        plt.savefig(save_path, dpi=150, bbox_inches="tight",
                    facecolor=fig.get_facecolor())
        plt.close(fig)

        self._print(f"Chart saved: {save_path}", "OK")
        return save_path

    # -------------------------------------------------------------------------
    # STEP 6: Export to Excel
    # -------------------------------------------------------------------------

    def export(self, output_path: str = None) -> str:
        """
        Export simulation results to a formatted Excel workbook with 4 sheets:
          1. Summary      — Percentile table, best/base/worst, key stats
          2. Raw Results  — Full N-row simulation output (sampled if N > 5,000)
          3. Share Dist.  — Revenue share distribution stats per product
          4. Inputs       — All simulation parameters for reproducibility
        """
        if self.results is None:
            raise RuntimeError("Call .run() before .export()")

        PnLBase.ensure_dir(OUTPUT_DIR)
        if output_path is None:
            ts_str      = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(OUTPUT_DIR, f"monte_carlo_{ts_str}.xlsx")

        self._section(f"EXPORTING TO EXCEL")
        self._print(f"Output: {output_path}")

        navy_hex   = "1F4E79"
        green_hex  = "375623"
        amber_hex  = "7F3F00"
        red_hex    = "9C0006"
        white_hex  = "FFFFFF"
        lt_blue    = "D9E2F3"
        lt_green   = "E2EFDA"
        lt_red     = "FFE0E0"
        lt_amber   = "FCE4D6"
        lt_grey    = "F2F2F2"

        ts = self.results["TotalSpend"]
        p5, p25, p50, p75, p95 = [np.percentile(ts, p) for p in PERCENTILES]
        baseline = self.baseline["total_spend"]

        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, numbers)
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            self._print("openpyxl not installed — skipping Excel export", "WARN")
            return None

        wb = openpyxl.Workbook()

        # ── Helper: style functions ──────────────────────────────────────────
        def hdr_fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def hdr_font(bold=True, size=10, color=white_hex):
            return Font(bold=bold, size=size, color=color, name="Calibri")

        def data_font(bold=False, size=10, color="000000"):
            return Font(bold=bold, size=size, color=color, name="Calibri")

        def centered():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

        def right_align():
            return Alignment(horizontal="right", vertical="center")

        def thin_border():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        def write_cell(ws, row, col, value, bold=False, bg=None, fg="000000",
                       align="left", size=10, num_format=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font      = Font(bold=bold, size=size, color=fg, name="Calibri")
            cell.alignment = Alignment(horizontal=align, vertical="center",
                                       wrap_text=False)
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            if num_format:
                cell.number_format = num_format
            cell.border = thin_border()
            return cell

        # ── Sheet 1: Summary ─────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions["A"].width = 32
        for col in "BCDEFGHI":
            ws1.column_dimensions[col].width = 16

        r = 1
        # Title block
        ws1.merge_cells(f"A{r}:I{r}")
        write_cell(ws1, r, 1,
                   f"Monte Carlo P&L Risk Simulation — {FY_LABEL}",
                   bold=True, bg=navy_hex, fg=white_hex, align="center", size=14)
        ws1.row_dimensions[r].height = 26
        r += 1
        ws1.merge_cells(f"A{r}:I{r}")
        write_cell(ws1, r, 1,
                   f"{self.iterations:,} iterations  |  "
                   f"Seed: {self.seed if self.seed is not None else 'random'}  |  "
                   f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                   bg=navy_hex, fg="BDD7EE", align="center", size=9)
        ws1.row_dimensions[r].height = 16
        r += 2

        # Key stats block
        write_cell(ws1, r, 1, "KEY STATISTICS", bold=True, bg=lt_blue, size=10)
        ws1.merge_cells(f"B{r}:I{r}")
        ws1.row_dimensions[r].height = 18
        r += 1

        stats = [
            ("Baseline (Actual GL)",       baseline,           "$#,##0",  None),
            ("Mean (Simulated)",            ts.mean(),          "$#,##0",  None),
            ("Standard Deviation",          ts.std(),           "$#,##0",  None),
            ("Coefficient of Variation",    ts.std()/ts.mean(), "0.0%",    None),
            ("Value at Risk (P95–P50)",     p95 - p50,          "$#,##0",  None),
            ("Shock Events (if enabled)",
             f"{self.results['HadShock'].sum():,} of {self.iterations:,}",
             None, None),
        ]
        for label, val, fmt, _ in stats:
            write_cell(ws1, r, 1, label, bold=False, bg=lt_grey, size=9)
            c = write_cell(ws1, r, 2, val, bold=True, align="right", size=9)
            if fmt:
                c.number_format = fmt
            ws1.merge_cells(f"C{r}:I{r}")
            r += 1

        r += 1

        # Scenario summary
        write_cell(ws1, r, 1, "SCENARIO SUMMARY", bold=True, bg=lt_blue, size=10)
        ws1.merge_cells(f"B{r}:I{r}")
        ws1.row_dimensions[r].height = 18
        r += 1

        hdr_scenarios = ["Scenario", "Total Spend", "vs Baseline", "Probability Level"]
        for ci, h in enumerate(hdr_scenarios, 1):
            write_cell(ws1, r, ci, h, bold=True, bg=navy_hex, fg=white_hex,
                       align="center", size=9)
        ws1.row_dimensions[r].height = 18
        r += 1

        scenarios = [
            ("Best Case",  p5,  green_hex, lt_green),
            ("Base Case",  p50, navy_hex,  lt_blue),
            ("Worst Case", p95, red_hex,   lt_red),
        ]
        for label, val, hx, bg_h in scenarios:
            delta = (val - baseline) / baseline if baseline else 0
            pct_level = "P5 (5th percentile)" if "Best" in label else (
                "P50 (50th percentile)" if "Base" in label else "P95 (95th percentile)"
            )
            write_cell(ws1, r, 1, label, bold=True, bg=bg_h, fg=hx, size=10)
            write_cell(ws1, r, 2, val,   bold=True, bg=bg_h, align="right",
                       size=10, num_format="$#,##0")
            c_delta = write_cell(ws1, r, 3, delta, bold=True, bg=bg_h,
                                 align="right", size=10, num_format="+0.0%;-0.0%;0.0%")
            c_delta.font = Font(bold=True, size=10, color=hx, name="Calibri")
            write_cell(ws1, r, 4, pct_level, bg=bg_h, size=9)
            ws1.row_dimensions[r].height = 20
            r += 1

        r += 1

        # Full percentile table
        write_cell(ws1, r, 1, "FULL PERCENTILE TABLE", bold=True, bg=lt_blue, size=10)
        ws1.merge_cells(f"B{r}:I{r}")
        ws1.row_dimensions[r].height = 18
        r += 1

        hdr_cols = ["Metric", "Baseline"] + PERCENTILE_LABELS + ["Mean", "Std Dev", "CV"]
        for ci, h in enumerate(hdr_cols, 1):
            bg = navy_hex if ci == 1 else (
                green_hex if "Best" in h else (
                red_hex   if "Worst" in h else navy_hex
            ))
            write_cell(ws1, r, ci, h, bold=True, bg=bg, fg=white_hex,
                       align="center", size=9)
        ws1.row_dimensions[r].height = 32
        r += 1

        currency_fmt = "$#,##0"
        pct_fmt      = "0.0%"

        for idx, (metric, row_data) in enumerate(self.summary.iterrows()):
            bg = lt_grey if idx % 2 == 0 else "FFFFFF"
            is_share = "Share" in metric
            fmt      = pct_fmt if is_share else currency_fmt

            write_cell(ws1, r, 1, metric, bg=bg, size=9)
            c = write_cell(ws1, r, 2, row_data["Baseline"], bg=bg,
                           align="right", size=9, num_format=fmt)
            for ci, col in enumerate(PERCENTILE_LABELS, 3):
                if col in row_data:
                    bg2 = lt_green if "Best" in col else (lt_red if "Worst" in col else bg)
                    write_cell(ws1, r, ci, row_data[col], bg=bg2,
                               align="right", size=9, num_format=fmt)
            write_cell(ws1, r, 8, row_data.get("Mean",   0), bg=bg,
                       align="right", size=9, num_format=fmt)
            write_cell(ws1, r, 9, row_data.get("StdDev", 0), bg=bg,
                       align="right", size=9, num_format=fmt)
            # CV column
            cv_c = write_cell(ws1, r, 10, row_data.get("CV", 0), bg=bg,
                              align="right", size=9, num_format=pct_fmt)
            ws1.row_dimensions[r].height = 16
            r += 1

        # ── Sheet 2: Raw Results (sampled) ───────────────────────────────────
        ws2 = wb.create_sheet("Raw Results")
        ws2.sheet_view.showGridLines = False

        # Sample for large simulations
        max_rows = 5000
        if len(self.results) > max_rows:
            sample = self.results.sample(n=max_rows, random_state=42)
            ws2["A1"] = (f"Note: Showing {max_rows:,} random samples "
                         f"of {len(self.results):,} total iterations")
            ws2["A1"].font = Font(italic=True, size=9, color="808080")
            start_row = 2
        else:
            sample    = self.results
            start_row = 1

        # Write headers
        headers = (["TotalSpend", "HadShock"] +
                   [f"{p}_RawSpend" for p in PRODUCTS] +
                   [f"{p}_AllocSpend" for p in PRODUCTS] +
                   [f"{p}_Share"      for p in PRODUCTS])
        for ci, h in enumerate(headers, 1):
            c = ws2.cell(row=start_row, column=ci, value=h)
            c.font      = Font(bold=True, size=9, color=white_hex, name="Calibri")
            c.fill      = PatternFill("solid", fgColor=navy_hex)
            c.alignment = centered()

        # Write data rows
        for ri, (_, row_data) in enumerate(sample.iterrows(), start_row + 1):
            for ci, col in enumerate(headers, 1):
                val = row_data[col]
                c   = ws2.cell(row=ri, column=ci, value=bool(val) if col == "HadShock" else val)
                c.font   = Font(size=8, name="Calibri")
                c.number_format = "0.0%" if "Share" in col else "$#,##0"
            if ri % 1000 == 0:
                self._print(f"  Writing rows: {ri:,}/{len(sample):,}...")

        for ci in range(1, len(headers) + 1):
            ws2.column_dimensions[get_column_letter(ci)].width = 16

        # ── Sheet 3: Inputs ──────────────────────────────────────────────────
        ws3 = wb.create_sheet("Inputs & Parameters")
        ws3.sheet_view.showGridLines = False
        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 20

        r3 = 1
        write_cell(ws3, r3, 1, "SIMULATION INPUTS & PARAMETERS",
                   bold=True, bg=navy_hex, fg=white_hex, size=12)
        ws3.merge_cells(f"A{r3}:B{r3}")
        ws3.row_dimensions[r3].height = 22
        r3 += 2

        params = [
            ("Source File",           self.file_path),
            ("Iterations",            self.iterations),
            ("Random Seed",           str(self.seed) if self.seed is not None else "random"),
            ("Share Concentration α", self.share_conc),
            ("Expense CV (default)",  f"{self.expense_cv:.1%}"),
            ("Shock Probability",     f"{self.shock_prob:.1%}"),
            ("Shock Size",            f"{self.shock_size:.1%}"),
            ("Baseline Total Spend",  baseline),
            ("Months with Data",      self.baseline["months_present"]),
            ("Run Timestamp",         datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for label, val in params:
            write_cell(ws3, r3, 1, label, bg=lt_grey, size=9)
            c = write_cell(ws3, r3, 2, val, align="right", size=9)
            if isinstance(val, float) and val > 100:
                c.number_format = "$#,##0"
            ws3.row_dimensions[r3].height = 16
            r3 += 1

        r3 += 1
        write_cell(ws3, r3, 1, "CONFIGURED REVENUE SHARES",
                   bold=True, bg=lt_blue, size=10)
        ws3.merge_cells(f"A{r3}:B{r3}")
        r3 += 1
        for product in PRODUCTS:
            write_cell(ws3, r3, 1, product,                       bg=lt_grey, size=9)
            write_cell(ws3, r3, 2, REVENUE_SHARES.get(product,0),
                       align="right", size=9, num_format="0.0%")
            r3 += 1

        r3 += 1
        write_cell(ws3, r3, 1, "OBSERVED EXPENSE VOLATILITY (CV BY PRODUCT)",
                   bold=True, bg=lt_blue, size=10)
        ws3.merge_cells(f"A{r3}:B{r3}")
        r3 += 1
        for product in PRODUCTS:
            cv = self.baseline["product_cv"].get(product, self.expense_cv)
            write_cell(ws3, r3, 1, product,          bg=lt_grey, size=9)
            write_cell(ws3, r3, 2, cv, align="right", size=9, num_format="0.0%")
            r3 += 1

        # Save
        wb.save(output_path)
        self._print(f"Excel workbook saved: {output_path}", "OK")
        return output_path


# =============================================================================
# COMMAND-LINE ENTRY POINT
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        prog="pnl_monte_carlo",
        description=(
            f"{APP_NAME} v{APP_VERSION} — Monte Carlo P&L Risk Simulation\n\n"
            "Runs N iterations of the P&L model with randomized allocation shares\n"
            "and expense amounts to produce full probability distributions."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python pnl_monte_carlo.py                              # 10,000 iter, default file
  python pnl_monte_carlo.py --iterations 50000           # higher precision
  python pnl_monte_carlo.py --seed 42                    # reproducible results
  python pnl_monte_carlo.py --shock-prob 0.05            # 5% shock probability
  python pnl_monte_carlo.py --export results.xlsx        # save to Excel
  python pnl_monte_carlo.py --no-chart                   # skip chart generation
  python pnl_monte_carlo.py --concentration 5            # wider share distribution
""",
    )

    parser.add_argument(
        "--file", "-f",
        default=SOURCE_FILE,
        help=f"Source Excel file (default: {SOURCE_FILE})"
    )
    parser.add_argument(
        "--iterations", "-n",
        type=int,
        default=DEFAULT_ITERATIONS,
        help=f"Number of Monte Carlo iterations (default: {DEFAULT_ITERATIONS:,})"
    )
    parser.add_argument(
        "--seed", "-s",
        type=int,
        default=DEFAULT_SEED,
        help="Random seed for reproducible results (default: random)"
    )
    parser.add_argument(
        "--concentration", "-c",
        type=float,
        default=DEFAULT_SHARE_CONC,
        help=(
            f"Dirichlet concentration for share distribution (default: {DEFAULT_SHARE_CONC}). "
            "Higher = tighter distribution around configured shares. "
            "Lower = wider swing in allocations."
        )
    )
    parser.add_argument(
        "--expense-cv",
        type=float,
        default=DEFAULT_EXPENSE_CV,
        help=(
            f"Default expense coefficient of variation (default: {DEFAULT_EXPENSE_CV:.0%}). "
            "Overridden by observed monthly volatility in GL where data allows."
        )
    )
    parser.add_argument(
        "--shock-prob",
        type=float,
        default=DEFAULT_SHOCK_PROB,
        help=(
            f"Probability of an expense shock event per iteration "
            f"(default: {DEFAULT_SHOCK_PROB:.0%}). "
            "Models low-probability, high-impact cost spikes."
        )
    )
    parser.add_argument(
        "--shock-size",
        type=float,
        default=DEFAULT_SHOCK_SIZE,
        help=(
            f"Size of shock as fraction of total spend (default: {DEFAULT_SHOCK_SIZE:.0%}). "
            "Only relevant when --shock-prob > 0."
        )
    )
    parser.add_argument(
        "--export", "-e",
        default=None,
        metavar="PATH",
        help="Export results to Excel workbook at this path"
    )
    parser.add_argument(
        "--chart-path",
        default=None,
        metavar="PATH",
        help="Save distribution chart to this path (PNG)"
    )
    parser.add_argument(
        "--no-chart",
        action="store_true",
        help="Skip chart generation"
    )
    parser.add_argument(
        "--quiet", "-q",
        action="store_true",
        help="Suppress console output (still exports if --export is set)"
    )

    args = parser.parse_args()

    # --- Validate inputs ---
    if args.iterations < 100:
        parser.error("--iterations must be at least 100")
    if not 0.0 <= args.shock_prob <= 1.0:
        parser.error("--shock-prob must be between 0.0 and 1.0")
    if args.concentration <= 0:
        parser.error("--concentration must be positive")

    # --- Run simulation ---
    sim = MonteCarloSimulator(
        file_path   = args.file,
        iterations  = args.iterations,
        seed        = args.seed,
        share_conc  = args.concentration,
        expense_cv  = args.expense_cv,
        shock_prob  = args.shock_prob,
        shock_size  = args.shock_size,
        verbose     = not args.quiet,
    )

    sim.load().run()

    if not args.quiet:
        sim.print_results()

    if not args.no_chart:
        try:
            sim.plot(save_path=args.chart_path)
        except Exception as e:
            print(f"  ⚠ Chart generation failed: {e}")

    if args.export:
        sim.export(output_path=args.export)
    elif not args.quiet:
        print("  Tip: run with --export results.xlsx to save full results to Excel")
        print()


if __name__ == "__main__":
    main()

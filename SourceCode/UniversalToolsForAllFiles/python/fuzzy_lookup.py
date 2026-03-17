"""
KBT Universal Tools — Fuzzy Match / Fuzzy Lookup
============================================================
Matches records between two datasets using fuzzy string
matching. Catches vendor/customer name typos and variations.

Useful for: vendor deduplication, customer master cleanup,
GL account mapping, and any lookup that fails due to spelling.

Usage:
    python fuzzy_lookup.py "C:\\data.xlsx" "C:\\master.xlsx" --lookup-col "Vendor" --match-col "Vendor Name"
    python fuzzy_lookup.py "data.xlsx" "master.xlsx" --threshold 80

Output: Saves "FUZZY_MATCH_RESULTS.xlsx" with match scores
"""

import sys
import os
import argparse
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)

try:
    from thefuzz import fuzz, process
except ImportError:
    print("ERROR: thefuzz not installed. Run: pip install thefuzz python-Levenshtein")
    sys.exit(1)


def fuzzy_match(source_file: str, lookup_file: str, source_col: str,
                lookup_col: str, threshold: int, return_cols: list = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Fuzzy Match / Fuzzy Lookup")
    print(f"{'='*55}")
    print(f"  Source:    {os.path.basename(source_file)} — column: '{source_col}'")
    print(f"  Lookup:    {os.path.basename(lookup_file)} — column: '{lookup_col}'")
    print(f"  Threshold: {threshold}% match confidence required")
    print(f"{'='*55}\n")

    for f in [source_file, lookup_file]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)

    source_df = pd.read_excel(source_file)
    lookup_df = pd.read_excel(lookup_file)
    source_df.columns = source_df.columns.str.strip()
    lookup_df.columns = lookup_df.columns.str.strip()

    for col, df, fname in [(source_col, source_df, source_file), (lookup_col, lookup_df, lookup_file)]:
        if col not in df.columns:
            print(f"ERROR: Column '{col}' not found in {os.path.basename(fname)}.")
            print(f"Available: {', '.join(df.columns.tolist())}")
            sys.exit(1)

    lookup_values = lookup_df[lookup_col].dropna().astype(str).tolist()
    print(f"Lookup table size: {len(lookup_values)} entries")
    print(f"Records to match:  {len(source_df)}\n")
    print("Matching... (this may take a moment for large files)")

    matched_names = []
    scores = []
    match_status = []

    for val in source_df[source_col].astype(str):
        if val.strip() == "" or val == "nan":
            matched_names.append("")
            scores.append(0)
            match_status.append("BLANK")
            continue

        result = process.extractOne(val, lookup_values, scorer=fuzz.token_sort_ratio)
        if result and result[1] >= threshold:
            matched_names.append(result[0])
            scores.append(result[1])
            match_status.append("MATCHED" if result[1] == 100 else "FUZZY MATCH")
        else:
            matched_names.append("")
            scores.append(result[1] if result else 0)
            match_status.append("NO MATCH")

    source_df['Matched_Value'] = matched_names
    source_df['Match_Score_%'] = scores
    source_df['Match_Status'] = match_status

    # Add lookup columns if requested
    if return_cols:
        for rc in return_cols:
            if rc in lookup_df.columns:
                rc_map = dict(zip(lookup_df[lookup_col].astype(str), lookup_df[rc].astype(str)))
                source_df[f'Lookup_{rc}'] = source_df['Matched_Value'].map(rc_map)

    matched_count = (source_df['Match_Status'] != 'NO MATCH').sum()
    no_match_count = (source_df['Match_Status'] == 'NO MATCH').sum()

    print(f"\nResults:")
    print(f"  Exact matches:    {(source_df['Match_Status'] == 'MATCHED').sum()}")
    print(f"  Fuzzy matches:    {(source_df['Match_Status'] == 'FUZZY MATCH').sum()}")
    print(f"  No match found:   {no_match_count}")
    print(f"  Match rate:       {matched_count / len(source_df) * 100:.1f}%")

    out_dir = os.path.dirname(source_file)
    output_path = os.path.join(out_dir, "FUZZY_MATCH_RESULTS.xlsx")

    fill_exact  = PatternFill("solid", fgColor="C6EFCE")
    fill_fuzzy  = PatternFill("solid", fgColor="FFEB9C")
    fill_none   = PatternFill("solid", fgColor="FFC7CE")
    fill_header = PatternFill("solid", fgColor="1F497D")
    font_header = Font(bold=True, color="FFFFFF")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        source_df.to_excel(writer, sheet_name='Results', index=False)
        ws = writer.sheets['Results']

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header

        status_col_idx = source_df.columns.get_loc('Match_Status') + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status_val = row[status_col_idx - 1].value
            if status_val == 'MATCHED':
                fill = fill_exact
            elif status_val == 'FUZZY MATCH':
                fill = fill_fuzzy
            elif status_val == 'NO MATCH':
                fill = fill_none
            else:
                fill = None
            if fill:
                for cell in row:
                    cell.fill = fill

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

        # No-match sheet for easy review
        no_match_df = source_df[source_df['Match_Status'] == 'NO MATCH']
        if not no_match_df.empty:
            no_match_df.to_excel(writer, sheet_name='No Match — Review', index=False)

    print(f"\n{'='*55}")
    print(f"  DONE! Report saved to: {output_path}")
    print(f"  Color: Green=Exact | Yellow=Fuzzy | Red=No Match")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Fuzzy Match / Fuzzy Lookup')
    parser.add_argument('source', help='Path to the source file (records to look up)')
    parser.add_argument('lookup', help='Path to the lookup/master file')
    parser.add_argument('--lookup-col', default='Name', help='Column in source file to match from')
    parser.add_argument('--match-col', default='Name', help='Column in lookup file to match against')
    parser.add_argument('--threshold', type=int, default=80,
                        help='Minimum match score 0-100 (default: 80)')
    parser.add_argument('--return-cols', nargs='+', default=None,
                        help='Extra columns from lookup file to return')
    args = parser.parse_args()
    fuzzy_match(args.source, args.lookup, args.lookup_col, args.match_col,
                args.threshold, args.return_cols)


if __name__ == '__main__':
    main()

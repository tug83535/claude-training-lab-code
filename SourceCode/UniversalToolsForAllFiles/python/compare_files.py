"""
KBT Universal Tools — Compare Two Excel Files
============================================================
Performs a cell-by-cell comparison of two Excel workbooks.
Produces a diff report highlighting every difference found.

Usage:
    python compare_files.py "C:\\path\\file1.xlsx" "C:\\path\\file2.xlsx"
    python compare_files.py "file1.xlsx" "file2.xlsx" --sheet "Sheet1"

Output: Saves "COMPARISON_REPORT.xlsx" in the same folder as file1
"""

import sys
import os
import argparse
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Required packages not installed.")
    print("Run: pip install pandas openpyxl")
    sys.exit(1)


FILL_ADDED   = PatternFill("solid", fgColor="C6EFCE")  # Green — in file2 only
FILL_REMOVED = PatternFill("solid", fgColor="FFC7CE")  # Red   — in file1 only
FILL_CHANGED = PatternFill("solid", fgColor="FFEB9C")  # Yellow — value changed
FILL_HEADER  = PatternFill("solid", fgColor="1F497D")  # Dark blue header
FONT_HEADER  = Font(bold=True, color="FFFFFF")


def compare_sheets(df1: pd.DataFrame, df2: pd.DataFrame, sheet_name: str, output_wb: openpyxl.Workbook) -> dict:
    summary = {'added': 0, 'removed': 0, 'changed': 0, 'sheet': sheet_name}

    ws = output_wb.create_sheet(title=sheet_name[:28] + " DIFF")

    # Headers
    headers = ["Location", "Column", "File 1 Value", "File 2 Value", "Change Type"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    row_out = 2

    # Align columns
    all_cols = list(dict.fromkeys(list(df1.columns) + list(df2.columns)))

    # Check row counts
    max_rows = max(len(df1), len(df2))
    df1 = df1.reindex(range(max_rows))
    df2 = df2.reindex(range(max_rows))

    for i in range(max_rows):
        for col in all_cols:
            val1 = df1[col].iloc[i] if col in df1.columns else None
            val2 = df2[col].iloc[i] if col in df2.columns else None

            val1_str = str(val1) if pd.notna(val1) else ""
            val2_str = str(val2) if pd.notna(val2) else ""

            if val1_str == val2_str:
                continue

            change_type = ""
            fill = None

            if val1_str == "" and val2_str != "":
                change_type = "ADDED"
                fill = FILL_ADDED
                summary['added'] += 1
            elif val1_str != "" and val2_str == "":
                change_type = "REMOVED"
                fill = FILL_REMOVED
                summary['removed'] += 1
            else:
                change_type = "CHANGED"
                fill = FILL_CHANGED
                summary['changed'] += 1

            row_data = [f"Row {i+2}", col, val1_str, val2_str, change_type]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_out, column=col_idx, value=val)
                if fill:
                    cell.fill = fill
            row_out += 1

    # Auto-fit
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    if row_out == 2:
        ws.cell(row=2, column=1, value="No differences found — sheets are identical.")

    return summary


def compare_files(file1: str, file2: str, sheet_name: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Excel File Comparison")
    print(f"{'='*55}")
    print(f"  File 1: {os.path.basename(file1)}")
    print(f"  File 2: {os.path.basename(file2)}")
    print(f"  Date:   {datetime.now().strftime('%m/%d/%Y %I:%M %p')}")
    print(f"{'='*55}\n")

    for f in [file1, file2]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)

    xl1 = pd.ExcelFile(file1)
    xl2 = pd.ExcelFile(file2)

    if sheet_name:
        sheets = [sheet_name]
    else:
        sheets = xl1.sheet_names
        print(f"Comparing {len(sheets)} sheet(s): {', '.join(sheets)}\n")

    output_wb = openpyxl.Workbook()
    output_wb.remove(output_wb.active)

    # Summary sheet
    summary_ws = output_wb.create_sheet(title="SUMMARY", index=0)
    sum_headers = ["Sheet", "Added Cells", "Removed Cells", "Changed Cells", "Total Differences"]
    for col_idx, h in enumerate(sum_headers, 1):
        cell = summary_ws.cell(row=1, column=col_idx, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    sum_row = 2
    total_diffs = 0

    for sht in sheets:
        if sht not in xl1.sheet_names:
            print(f"  Sheet '{sht}' not in File 1 — skipping.")
            continue
        if sht not in xl2.sheet_names:
            print(f"  Sheet '{sht}' not in File 2 — skipping.")
            continue

        print(f"Comparing sheet: {sht}")
        df1 = pd.read_excel(file1, sheet_name=sht, dtype=str).fillna("")
        df2 = pd.read_excel(file2, sheet_name=sht, dtype=str).fillna("")

        summary = compare_sheets(df1, df2, sht, output_wb)
        sheet_total = summary['added'] + summary['removed'] + summary['changed']
        total_diffs += sheet_total

        summary_ws.cell(row=sum_row, column=1, value=sht)
        summary_ws.cell(row=sum_row, column=2, value=summary['added'])
        summary_ws.cell(row=sum_row, column=3, value=summary['removed'])
        summary_ws.cell(row=sum_row, column=4, value=summary['changed'])
        summary_ws.cell(row=sum_row, column=5, value=sheet_total)
        sum_row += 1

        print(f"  Added: {summary['added']} | Removed: {summary['removed']} | Changed: {summary['changed']}")

    for col in summary_ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        summary_ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    out_dir = os.path.dirname(file1)
    out_path = os.path.join(out_dir, "COMPARISON_REPORT.xlsx")
    output_wb.save(out_path)

    print(f"\n{'='*55}")
    print(f"  DONE! Total differences found: {total_diffs}")
    print(f"  Report saved to:")
    print(f"  {out_path}")
    print(f"  Color key: Green=Added | Red=Removed | Yellow=Changed")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Compare Two Excel Files')
    parser.add_argument('file1', help='Path to the first (original) Excel file')
    parser.add_argument('file2', help='Path to the second (new) Excel file')
    parser.add_argument('--sheet', help='Specific sheet to compare (default: all matching sheets)')
    args = parser.parse_args()
    compare_files(args.file1, args.file2, sheet_name=args.sheet)


if __name__ == '__main__':
    main()

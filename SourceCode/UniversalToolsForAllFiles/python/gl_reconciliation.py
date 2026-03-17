"""
KBT Universal Tools — GL Reconciliation Engine
============================================================
Matches two large transaction lists (e.g., GL vs Sub-ledger)
and flags unmatched items for investigation.

Matching logic: Amount + Date (within 3 days) + optional reference

Usage:
    python gl_reconciliation.py "C:\\path\\gl.xlsx" "C:\\path\\subledger.xlsx"
    python gl_reconciliation.py "gl.xlsx" "sub.xlsx" --amount "Amt" --date "Post Date" --ref "Invoice"

Output: Saves "GL_RECONCILIATION_REPORT.xlsx" with matched/unmatched details
"""

import sys
import os
import argparse
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)


def reconcile(file1: str, file2: str, amount_col: str, date_col: str,
              ref_col: str = None, date_tolerance: int = 3) -> None:
    print(f"\n{'='*55}")
    print("  KBT GL Reconciliation Engine")
    print(f"{'='*55}")
    print(f"  File 1 (GL):       {os.path.basename(file1)}")
    print(f"  File 2 (Sub-ledger): {os.path.basename(file2)}")
    print(f"  Amount column: '{amount_col}' | Date column: '{date_col}'")
    print(f"  Date tolerance: ±{date_tolerance} days")
    print(f"{'='*55}\n")

    for f in [file1, file2]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)

    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)
    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    for df, name in [(df1, 'File 1'), (df2, 'File 2')]:
        for col in [amount_col, date_col]:
            if col not in df.columns:
                print(f"ERROR: Column '{col}' not found in {name}.")
                print(f"Available: {', '.join(df.columns.tolist())}")
                sys.exit(1)

    df1[amount_col] = pd.to_numeric(df1[amount_col], errors='coerce')
    df2[amount_col] = pd.to_numeric(df2[amount_col], errors='coerce')
    df1[date_col] = pd.to_datetime(df1[date_col], errors='coerce')
    df2[date_col] = pd.to_datetime(df2[date_col], errors='coerce')

    df1['_matched'] = False
    df2['_matched'] = False
    matched_pairs = []

    for i, row1 in df1.iterrows():
        if pd.isna(row1[amount_col]) or pd.isna(row1[date_col]):
            continue
        for j, row2 in df2[~df2['_matched']].iterrows():
            if pd.isna(row2[amount_col]) or pd.isna(row2[date_col]):
                continue
            amount_match = abs(row1[amount_col] - row2[amount_col]) < 0.01
            date_match = abs((row1[date_col] - row2[date_col]).days) <= date_tolerance
            ref_match = True
            if ref_col and ref_col in df1.columns and ref_col in df2.columns:
                ref_match = str(row1[ref_col]).strip().lower() == str(row2[ref_col]).strip().lower()

            if amount_match and date_match and ref_match:
                df1.at[i, '_matched'] = True
                df2.at[j, '_matched'] = True
                matched_pairs.append({
                    'GL_Row': i + 2,
                    'Sub_Row': j + 2,
                    'Amount': row1[amount_col],
                    'GL_Date': row1[date_col],
                    'Sub_Date': row2[date_col],
                })
                break

    unmatched1 = df1[~df1['_matched']].drop(columns=['_matched'])
    unmatched2 = df2[~df2['_matched']].drop(columns=['_matched'])
    matched_df = df1[df1['_matched']].drop(columns=['_matched'])

    print(f"Results:")
    print(f"  GL records:          {len(df1)}")
    print(f"  Sub-ledger records:  {len(df2)}")
    print(f"  Matched:             {len(matched_pairs)}")
    print(f"  Unmatched in GL:     {len(unmatched1)}")
    print(f"  Unmatched in Sub:    {len(unmatched2)}")

    out_dir = os.path.dirname(file1)
    output_path = os.path.join(out_dir, "GL_RECONCILIATION_REPORT.xlsx")

    fill_matched   = PatternFill("solid", fgColor="C6EFCE")
    fill_unmatched = PatternFill("solid", fgColor="FFC7CE")
    fill_header    = PatternFill("solid", fgColor="1F497D")
    font_header    = Font(bold=True, color="FFFFFF")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Summary
        summary_data = {
            'Metric': ['GL Records', 'Sub-ledger Records', 'Matched Pairs',
                       'Unmatched in GL', 'Unmatched in Sub-ledger',
                       'GL Total Amount', 'Sub Total Amount', 'Difference'],
            'Value': [len(df1), len(df2), len(matched_pairs),
                      len(unmatched1), len(unmatched2),
                      df1[amount_col].sum(), df2[amount_col].sum(),
                      round(df1[amount_col].sum() - df2[amount_col].sum(), 2)]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

        if not unmatched1.empty:
            unmatched1.to_excel(writer, sheet_name='Unmatched GL', index=False)
            _style_sheet(writer.sheets['Unmatched GL'], fill_unmatched, fill_header, font_header)

        if not unmatched2.empty:
            unmatched2.to_excel(writer, sheet_name='Unmatched Sub-ledger', index=False)
            _style_sheet(writer.sheets['Unmatched Sub-ledger'], fill_unmatched, fill_header, font_header)

        if matched_pairs:
            pd.DataFrame(matched_pairs).to_excel(writer, sheet_name='Matched Pairs', index=False)
            _style_sheet(writer.sheets['Matched Pairs'], fill_matched, fill_header, font_header)

    print(f"\n{'='*55}")
    print(f"  DONE! Report saved to: {output_path}")
    print(f"{'='*55}\n")


def _style_sheet(ws, row_fill, header_fill, header_font):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = row_fill
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)


def main():
    parser = argparse.ArgumentParser(description='KBT GL Reconciliation Engine')
    parser.add_argument('file1', help='Path to GL file')
    parser.add_argument('file2', help='Path to Sub-ledger file')
    parser.add_argument('--amount', default='Amount', help='Amount column name')
    parser.add_argument('--date', default='Date', help='Date column name')
    parser.add_argument('--ref', default=None, help='Reference/Invoice column name (optional)')
    parser.add_argument('--tolerance', type=int, default=3, help='Date tolerance in days (default: 3)')
    args = parser.parse_args()
    reconcile(args.file1, args.file2, args.amount, args.date, args.ref, args.tolerance)


if __name__ == '__main__':
    main()

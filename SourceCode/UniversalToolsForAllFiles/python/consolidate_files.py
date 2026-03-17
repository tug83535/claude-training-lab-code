"""
KBT Universal Tools — Multi-File Data Consolidator
============================================================
Combines data from hundreds of Excel files in a folder into
one master sheet with a "Source_File" column added.

All files must have the same column structure (same headers).

Usage:
    python consolidate_files.py "C:\\path\\to\\folder"
    python consolidate_files.py "C:\\path\\to\\folder" --sheet "Data"
    python consolidate_files.py "C:\\path\\to\\folder" --pattern "*Q1*"

Output: Saves "MASTER_CONSOLIDATED.xlsx" in the folder
"""

import sys
import os
import glob
import argparse
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Required packages not installed.")
    print("Run: pip install pandas openpyxl")
    sys.exit(1)


def consolidate_files(folder: str, sheet_name: str = None, pattern: str = "*.xlsx") -> None:
    print(f"\n{'='*55}")
    print("  KBT Multi-File Data Consolidator")
    print(f"{'='*55}")
    print(f"  Folder:  {folder}")
    print(f"  Pattern: {pattern}")
    print(f"  Date:    {datetime.now().strftime('%m/%d/%Y %I:%M %p')}")
    print(f"{'='*55}\n")

    if not os.path.isdir(folder):
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)

    output_name = "MASTER_CONSOLIDATED.xlsx"
    all_patterns = [pattern]
    if not pattern.endswith(".xls"):
        all_patterns.append(pattern.replace(".xlsx", ".xls"))

    files = []
    for p in all_patterns:
        files.extend(glob.glob(os.path.join(folder, p)))

    files = sorted([f for f in files if os.path.basename(f) != output_name])

    if not files:
        print(f"ERROR: No files matching '{pattern}' found in folder.")
        sys.exit(1)

    print(f"Found {len(files)} file(s) to consolidate.\n")

    all_dfs = []
    skipped = []
    processed = 0

    for f in files:
        fname = os.path.basename(f)
        try:
            if sheet_name:
                df = pd.read_excel(f, sheet_name=sheet_name)
            else:
                df = pd.read_excel(f)

            df.insert(0, 'Source_File', fname)
            all_dfs.append(df)
            processed += 1
            print(f"  OK:   {fname:<40} {len(df):>5} rows")
        except Exception as e:
            print(f"  SKIP: {fname} — {e}")
            skipped.append((fname, str(e)))

    if not all_dfs:
        print("\nERROR: No files could be processed.")
        sys.exit(1)

    print(f"\nCombining {len(all_dfs)} file(s)...")
    master = pd.concat(all_dfs, ignore_index=True, sort=False)
    print(f"Total rows in master: {len(master):,}")

    output_path = os.path.join(folder, output_name)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        master.to_excel(writer, sheet_name='Consolidated', index=False)
        ws = writer.sheets['Consolidated']

        # Style header
        fill = PatternFill("solid", fgColor="1F497D")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

        # Auto-fit
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # File summary sheet
        if skipped:
            skip_df = pd.DataFrame(skipped, columns=['File', 'Error'])
            skip_df.to_excel(writer, sheet_name='Skipped Files', index=False)

        # Source summary
        source_summary = master.groupby('Source_File').size().reset_index(name='Row Count')
        source_summary.to_excel(writer, sheet_name='File Summary', index=False)
        ws_sum = writer.sheets['File Summary']
        for cell in ws_sum[1]:
            cell.fill = fill
            cell.font = font
        for col in ws_sum.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws_sum.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    print(f"\n{'='*55}")
    print(f"  DONE!")
    print(f"  Files processed: {processed}")
    print(f"  Files skipped:   {len(skipped)}")
    print(f"  Total rows:      {len(master):,}")
    print(f"  Saved to: {output_path}")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Multi-File Data Consolidator')
    parser.add_argument('folder', help='Path to folder containing Excel files')
    parser.add_argument('--sheet', default=None,
                        help='Sheet name to read from each file (default: first sheet)')
    parser.add_argument('--pattern', default='*.xlsx',
                        help='File name pattern to match (default: *.xlsx)')
    args = parser.parse_args()
    consolidate_files(args.folder, args.sheet, args.pattern)


if __name__ == '__main__':
    main()

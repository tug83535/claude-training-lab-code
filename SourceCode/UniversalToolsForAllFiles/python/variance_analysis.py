"""
KBT Universal Tools — Variance Analysis Generator
============================================================
Compares Actual vs Budget columns across multiple files and
creates a consolidated summary with waterfall-style breakdown.

Usage:
    python variance_analysis.py "C:\\path\\to\\folder"
    python variance_analysis.py "folder" --actual "Actuals" --budget "Plan" --label "Department"

Output: Saves "VARIANCE_ANALYSIS.xlsx" with summary and chart data
"""

import sys
import os
import glob
import argparse
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)


def variance_analysis(folder: str, actual_col: str, budget_col: str,
                       label_col: str, sheet_name: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Variance Analysis Generator")
    print(f"{'='*55}\n")

    files = [f for f in glob.glob(os.path.join(folder, "*.xlsx"))
             if "VARIANCE_ANALYSIS" not in os.path.basename(f)]

    if not files:
        print("ERROR: No Excel files found.")
        sys.exit(1)

    all_dfs = []
    for f in sorted(files):
        try:
            df = pd.read_excel(f, sheet_name=sheet_name or 0)
            df.columns = df.columns.str.strip()
            if actual_col in df.columns and budget_col in df.columns:
                df['Source_File'] = os.path.basename(f)
                all_dfs.append(df)
                print(f"  OK: {os.path.basename(f)}")
            else:
                print(f"  SKIP: {os.path.basename(f)} — required columns not found")
        except Exception as e:
            print(f"  ERROR: {os.path.basename(f)} — {e}")

    if not all_dfs:
        print("ERROR: No files with required columns found.")
        sys.exit(1)

    master = pd.concat(all_dfs, ignore_index=True, sort=False)
    master[actual_col] = pd.to_numeric(master[actual_col], errors='coerce')
    master[budget_col] = pd.to_numeric(master[budget_col], errors='coerce')
    master['$ Variance'] = master[actual_col] - master[budget_col]
    master['% Variance'] = master.apply(
        lambda r: r['$ Variance'] / abs(r[budget_col]) if pd.notna(r[budget_col]) and r[budget_col] != 0 else None,
        axis=1
    )
    master['Favorable?'] = master['$ Variance'].apply(
        lambda v: 'Favorable' if pd.notna(v) and v > 0 else ('Unfavorable' if pd.notna(v) and v < 0 else 'On Budget'))

    output_path = os.path.join(folder, "VARIANCE_ANALYSIS.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        master.to_excel(writer, sheet_name='Detail', index=False)

        # Summary by label column if provided
        if label_col and label_col in master.columns:
            summary = master.groupby(label_col).agg(
                Total_Actual=(actual_col, 'sum'),
                Total_Budget=(budget_col, 'sum')
            ).reset_index()
            summary['$ Variance'] = summary['Total_Actual'] - summary['Total_Budget']
            summary['% Variance'] = summary.apply(
                lambda r: r['$ Variance'] / abs(r['Total_Budget'])
                if r['Total_Budget'] != 0 else None, axis=1)
            summary.to_excel(writer, sheet_name='Summary', index=False)

            ws = writer.sheets['Summary']
            _style_ws(ws)

            # Add a bar chart
            chart = BarChart()
            chart.type = "col"
            chart.title = "Actual vs Budget by " + label_col
            chart.y_axis.title = "Amount"
            chart.x_axis.title = label_col

            data_ref = Reference(ws, min_col=2, max_col=3,
                                 min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            ws.add_chart(chart, "G2")

        _style_ws(writer.sheets['Detail'])

    print(f"\n{'='*55}")
    print(f"  DONE! Saved to: {output_path}")
    print(f"{'='*55}\n")


def _style_ws(ws):
    fill = PatternFill("solid", fgColor="1F497D")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)


def main():
    parser = argparse.ArgumentParser(description='KBT Variance Analysis Generator')
    parser.add_argument('folder', help='Folder containing files to analyze')
    parser.add_argument('--actual', default='Actual', help='Actual column name')
    parser.add_argument('--budget', default='Budget', help='Budget column name')
    parser.add_argument('--label', default='Department', help='Label/category column for grouping')
    parser.add_argument('--sheet', default=None, help='Sheet name to read')
    args = parser.parse_args()
    variance_analysis(args.folder, args.actual, args.budget, args.label, args.sheet)


if __name__ == '__main__':
    main()

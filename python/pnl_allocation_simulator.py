#!/usr/bin/env python3
"""
pnl_allocation_simulator.py — What-If Allocation Simulator
============================================================

PURPOSE: Lets users change revenue shares, AWS compute shares, or headcount
         shares and instantly see the P&L impact per product.

USAGE:
    python pnl_allocation_simulator.py
    python pnl_allocation_simulator.py --scenario "InsureSight=0.20,DocFast=0.10"
    python pnl_allocation_simulator.py --interactive
    python pnl_allocation_simulator.py --export scenario_output.xlsx

    from pnl_allocation_simulator import AllocationSimulator
    sim = AllocationSimulator(SOURCE_FILE)
    result = sim.simulate(revenue_shares={"InsureSight": 0.20, "DocFast": 0.10})
"""

import os
import sys
import argparse
from typing import Dict, List, Optional, Any
from copy import deepcopy

import pandas as pd
import numpy as np

try:
    from pnl_config import *
except ImportError:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from pnl_config import *


class AllocationSimulator(PnLBase):
    """Simulates P&L impact of changing allocation drivers."""

    def __init__(self, file_path: str = None, verbose: bool = True):
        super().__init__(verbose)
        self.file_path = file_path or SOURCE_FILE
        self.gl = None
        self.baseline = None

    def load(self) -> "AllocationSimulator":
        """Load GL data and compute baseline metrics."""
        self.gl = self._load_gl(self.file_path)
        self.baseline = self._compute_metrics(REVENUE_SHARES, "Baseline")
        return self

    def _compute_metrics(self, rev_shares: Dict[str, float], scenario_name: str) -> pd.DataFrame:
        """Compute P&L metrics for a given set of revenue shares."""
        gl = self.gl
        total_spend = gl[gl["Product"].isin(PRODUCTS)].groupby("Product")["Amount"].sum()
        total_abs = gl[gl["Product"].isin(PRODUCTS)].groupby("Product")["Abs_Amount"].sum()
        total_txns = gl[gl["Product"].isin(PRODUCTS)].groupby("Product")["Amount"].count()

        # Total revenue proxy (positive amounts)
        total_revenue = gl[gl["Is_Positive"] == 1]["Amount"].sum()

        rows = []
        for prod in PRODUCTS:
            share = rev_shares.get(prod, REVENUE_SHARES.get(prod, 0))
            est_revenue = total_revenue * share
            cost = total_spend.get(prod, 0)
            abs_cost = total_abs.get(prod, 0)
            cm = est_revenue + cost  # cost is typically negative
            cm_pct = cm / est_revenue if est_revenue != 0 else 0
            cost_to_rev = abs_cost / est_revenue if est_revenue != 0 else 0
            txns = total_txns.get(prod, 0)

            rows.append({
                "Scenario": scenario_name,
                "Product": prod,
                "Revenue_Share": share,
                "Est_Revenue": est_revenue,
                "Total_Cost": cost,
                "Abs_Cost": abs_cost,
                "CM_Dollar": cm,
                "CM_Pct": cm_pct,
                "Cost_to_Rev": cost_to_rev,
                "Transactions": txns,
            })

        return pd.DataFrame(rows)

    def simulate(self, revenue_shares: Dict[str, float] = None,
                 scenario_name: str = "Scenario") -> pd.DataFrame:
        """
        Run a what-if simulation with modified revenue shares.

        Args:
            revenue_shares: Dict of {product: new_share}. Only pass the ones you want to change.
            scenario_name: Label for this scenario.

        Returns:
            DataFrame comparing baseline vs scenario.
        """
        if self.gl is None:
            self.load()

        # Merge overrides into base shares
        new_shares = dict(REVENUE_SHARES)
        if revenue_shares:
            new_shares.update(revenue_shares)

        # Normalize to sum to 1.0
        share_sum = sum(new_shares.values())
        if abs(share_sum - 1.0) > 0.001:
            self._print(f"Shares sum to {share_sum:.3f}, normalizing to 1.0", "WARN")
            factor = 1.0 / share_sum
            new_shares = {k: v * factor for k, v in new_shares.items()}

        scenario = self._compute_metrics(new_shares, scenario_name)
        comparison = self._compare(self.baseline, scenario)
        return comparison

    def _compare(self, baseline: pd.DataFrame, scenario: pd.DataFrame) -> pd.DataFrame:
        """Build a comparison DataFrame showing baseline, scenario, and deltas."""
        base = baseline.set_index("Product")
        scen = scenario.set_index("Product")

        rows = []
        for prod in PRODUCTS:
            if prod in base.index and prod in scen.index:
                b = base.loc[prod]
                s = scen.loc[prod]
                rows.append({
                    "Product": prod,
                    "Base_Share": b["Revenue_Share"],
                    "New_Share": s["Revenue_Share"],
                    "Share_Change": s["Revenue_Share"] - b["Revenue_Share"],
                    "Base_Revenue": b["Est_Revenue"],
                    "New_Revenue": s["Est_Revenue"],
                    "Revenue_Delta": s["Est_Revenue"] - b["Est_Revenue"],
                    "Base_CM": b["CM_Dollar"],
                    "New_CM": s["CM_Dollar"],
                    "CM_Delta": s["CM_Dollar"] - b["CM_Dollar"],
                    "Base_CM_Pct": b["CM_Pct"],
                    "New_CM_Pct": s["CM_Pct"],
                    "CM_Pct_Change": s["CM_Pct"] - b["CM_Pct"],
                    "Base_Cost_Ratio": b["Cost_to_Rev"],
                    "New_Cost_Ratio": s["Cost_to_Rev"],
                })

        return pd.DataFrame(rows)

    def run_scenarios(self, scenarios: Dict[str, Dict[str, float]]) -> pd.DataFrame:
        """Run multiple named scenarios and return combined results."""
        all_results = []
        for name, shares in scenarios.items():
            result = self.simulate(revenue_shares=shares, scenario_name=name)
            result["Scenario_Name"] = name
            all_results.append(result)
        return pd.concat(all_results, ignore_index=True)

    def print_comparison(self, comparison: pd.DataFrame):
        """Print a formatted comparison table."""
        self._section("ALLOCATION SIMULATION RESULTS")

        for _, row in comparison.iterrows():
            prod = row["Product"]
            share_delta = row["Share_Change"]
            rev_delta = row["Revenue_Delta"]
            cm_delta = row["CM_Delta"]

            arrow = "↑" if share_delta > 0 else "↓" if share_delta < 0 else "—"

            self._print(f"\n  {prod}:")
            self._print(f"    Revenue Share:  {row['Base_Share']:.1%} → {row['New_Share']:.1%} ({arrow} {abs(share_delta):.1%})")
            self._print(f"    Est. Revenue:   {format_currency(row['Base_Revenue'])} → {format_currency(row['New_Revenue'])} ({format_currency(rev_delta)})")
            self._print(f"    CM:             {format_currency(row['Base_CM'])} → {format_currency(row['New_CM'])} ({format_currency(cm_delta)})")
            self._print(f"    CM%:            {row['Base_CM_Pct']:.1%} → {row['New_CM_Pct']:.1%}")

        # Net impact
        total_cm_delta = comparison["CM_Delta"].sum()
        self._print(f"\n  NET CM IMPACT: {format_currency(total_cm_delta)}")

    def export(self, comparison: pd.DataFrame, output_path: str = "allocation_scenario.xlsx"):
        """Export comparison to formatted Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, numbers

        wb = Workbook()
        ws = wb.active
        ws.title = "Scenario Comparison"

        # Title
        ws["A1"] = "Allocation What-If Scenario"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws["A2"] = f"Generated: {PnLBase.timestamp()}"

        # Headers
        headers = ["Product", "Base Share", "New Share", "Δ Share",
                    "Base Revenue", "New Revenue", "Δ Revenue",
                    "Base CM", "New CM", "Δ CM",
                    "Base CM%", "New CM%"]
        for col, hdr in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color="1F4E79", fill_type="solid")

        # Data
        for i, (_, row) in enumerate(comparison.iterrows(), 5):
            ws.cell(row=i, column=1, value=row["Product"])
            ws.cell(row=i, column=2, value=row["Base_Share"]).number_format = '0.0%'
            ws.cell(row=i, column=3, value=row["New_Share"]).number_format = '0.0%'
            ws.cell(row=i, column=4, value=row["Share_Change"]).number_format = '+0.0%;-0.0%'
            ws.cell(row=i, column=5, value=row["Base_Revenue"]).number_format = '$#,##0'
            ws.cell(row=i, column=6, value=row["New_Revenue"]).number_format = '$#,##0'
            ws.cell(row=i, column=7, value=row["Revenue_Delta"]).number_format = '+$#,##0;-$#,##0'
            ws.cell(row=i, column=8, value=row["Base_CM"]).number_format = '$#,##0'
            ws.cell(row=i, column=9, value=row["New_CM"]).number_format = '$#,##0'
            ws.cell(row=i, column=10, value=row["CM_Delta"]).number_format = '+$#,##0;-$#,##0'
            ws.cell(row=i, column=11, value=row["Base_CM_Pct"]).number_format = '0.0%'
            ws.cell(row=i, column=12, value=row["New_CM_Pct"]).number_format = '0.0%'

            # Color the delta
            delta_fill = PatternFill(
                start_color="E2EFDA" if row["CM_Delta"] >= 0 else "FFE0E0",
                fill_type="solid"
            )
            ws.cell(row=i, column=10).fill = delta_fill

        wb.save(output_path)
        self._print(f"Scenario exported: {output_path}", "OK")


def main():
    parser = argparse.ArgumentParser(description="Allocation What-If Simulator")
    parser.add_argument("--file", "-f", default=SOURCE_FILE)
    parser.add_argument("--scenario", "-s", default=None,
                        help='Share overrides: "InsureSight=0.20,DocFast=0.10"')
    parser.add_argument("--export", "-e", default=None, help="Export to Excel file")
    parser.add_argument("--presets", action="store_true",
                        help="Run 3 preset scenarios: Growth, Consolidation, Balanced")
    args = parser.parse_args()

    sim = AllocationSimulator(file_path=args.file)
    sim.load()

    if args.presets:
        # Run 3 common what-if scenarios
        scenarios = {
            "InsureSight Growth": {"InsureSight": 0.20, "DocFast": 0.08, "iGO": 0.47, "Affirm": 0.25},
            "iGO Consolidation":  {"iGO": 0.65, "Affirm": 0.22, "InsureSight": 0.08, "DocFast": 0.05},
            "Balanced Portfolio":  {"iGO": 0.35, "Affirm": 0.30, "InsureSight": 0.20, "DocFast": 0.15},
        }
        for name, shares in scenarios.items():
            result = sim.simulate(revenue_shares=shares, scenario_name=name)
            sim.print_comparison(result)
            if args.export:
                sim.export(result, f"scenario_{name.lower().replace(' ', '_')}.xlsx")
    elif args.scenario:
        # Parse "InsureSight=0.20,DocFast=0.10" format
        overrides = {}
        for pair in args.scenario.split(","):
            k, v = pair.strip().split("=")
            overrides[k.strip()] = float(v.strip())

        result = sim.simulate(revenue_shares=overrides)
        sim.print_comparison(result)
        if args.export:
            sim.export(result, args.export)
    else:
        # Default: show baseline
        sim._section("BASELINE P&L BY PRODUCT")
        for _, row in sim.baseline.iterrows():
            sim._print(f"  {row['Product']:15s}  Share: {row['Revenue_Share']:.0%}  "
                        f"Rev: {format_currency(row['Est_Revenue']):>12s}  "
                        f"CM: {format_currency(row['CM_Dollar']):>12s}  "
                        f"CM%: {row['CM_Pct']:.1%}")

        sim._print("\nUse --scenario or --presets to run what-if simulations")


if __name__ == "__main__":
    main()

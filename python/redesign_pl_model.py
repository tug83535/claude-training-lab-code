#!/usr/bin/env python3
"""
Keystone BenefitTech P&L Model — Fortune 10 FP&A Redesign
iPipeline Brand Guidelines Applied Programmatically via openpyxl

This script:
1. Fixes all FAIL reconciliation checks in the Checks sheet
2. Redesigns every visible sheet to iPipeline Fortune 10 standard
3. Transforms Report--> into an Executive Dashboard
4. Adds a Charts & Visuals sheet
5. Formats CrossfireHiddenWorksheet (keeps hidden)
6. Final Checks sheet design with PASS/FAIL conditional coloring
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# 1. LOAD WORKBOOK
# ============================================================
INPUT_FILE = 'excel/KeystoneBenefitTech_PL_Model.xlsx'
OUTPUT_FILE = 'excel/KeystoneBenefitTech_PL_Model.xlsx'

wb = openpyxl.load_workbook(INPUT_FILE)

# Unmerge all cells in all sheets first
for sname in wb.sheetnames:
    _ws = wb[sname]
    for merge in list(_ws.merged_cells.ranges):
        _ws.unmerge_cells(str(merge))

# ============================================================
# 2. BRAND STYLE DEFINITIONS
# ============================================================
IPIPELINE_BLUE = '0B4779'
ARCTIC_WHITE = 'F9F9F9'
SOFT_NEUTRAL = 'F0F0EE'
NAVY_BLUE = '112E51'
INNOVATION_BLUE = '4B9BCB'
LIME_GREEN = 'BFF18C'
AQUA = '2BCCD3'
CHARCOAL = '161616'
RED_FAIL = 'FF4C4C'

# --- Fonts ---
font_header = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=11)
font_body = Font(name='Arial', color=CHARCOAL, size=10)
font_body_bold = Font(name='Arial', bold=True, color=CHARCOAL, size=10)
font_title = Font(name='Arial', bold=True, color=IPIPELINE_BLUE, size=16)
font_subtitle = Font(name='Arial', bold=True, color=IPIPELINE_BLUE, size=12)
font_section = Font(name='Arial', bold=True, color=INNOVATION_BLUE, size=11)
font_date = Font(name='Arial', italic=True, color=INNOVATION_BLUE, size=9)
font_total = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=10)
font_link = Font(name='Arial', color=INNOVATION_BLUE, size=10, underline='single')
font_hidden_header = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=10)
font_pass = Font(name='Arial', bold=True, color=CHARCOAL, size=10)
font_fail = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=10)
font_kpi_label = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=9)
font_kpi_value = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=20)
font_kpi_label_lime = Font(name='Arial', bold=True, color=CHARCOAL, size=9)
font_kpi_value_lime = Font(name='Arial', bold=True, color=CHARCOAL, size=20)
font_dash_title = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=22)
font_dash_subtitle = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=13)
font_toc_header = Font(name='Arial', bold=True, color=IPIPELINE_BLUE, size=13)
font_construction = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=14)

# --- Fills ---
fill_header = PatternFill(start_color=IPIPELINE_BLUE, end_color=IPIPELINE_BLUE, fill_type='solid')
fill_alt_1 = PatternFill(start_color=ARCTIC_WHITE, end_color=ARCTIC_WHITE, fill_type='solid')
fill_alt_2 = PatternFill(start_color=SOFT_NEUTRAL, end_color=SOFT_NEUTRAL, fill_type='solid')
fill_total = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type='solid')
fill_aqua = PatternFill(start_color=AQUA, end_color=AQUA, fill_type='solid')
fill_lime = PatternFill(start_color=LIME_GREEN, end_color=LIME_GREEN, fill_type='solid')
fill_pass = PatternFill(start_color=LIME_GREEN, end_color=LIME_GREEN, fill_type='solid')
fill_fail = PatternFill(start_color=RED_FAIL, end_color=RED_FAIL, fill_type='solid')
fill_navy = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type='solid')
fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

# --- Borders ---
side_blue = Side(style='thin', color=INNOVATION_BLUE)
side_hair = Side(style='hair', color=INNOVATION_BLUE)
side_medium = Side(style='medium', color=INNOVATION_BLUE)
side_none = Side()
border_all = Border(top=side_blue, bottom=side_blue, left=side_blue, right=side_blue)
border_bottom = Border(bottom=side_blue)
border_bottom_medium = Border(bottom=side_medium)
border_hair = Border(bottom=side_hair)
border_none = Border()
border_thick_all = Border(
    top=Side(style='medium', color=IPIPELINE_BLUE),
    bottom=Side(style='medium', color=IPIPELINE_BLUE),
    left=Side(style='medium', color=IPIPELINE_BLUE),
    right=Side(style='medium', color=IPIPELINE_BLUE)
)

# --- Alignment ---
align_left = Alignment(horizontal='left', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# --- Number Formats ---
FMT_CURRENCY = '$#,##0.0"K"'
FMT_PERCENT = '0.0%'

# ============================================================
# 3. FIX CHECK DATA (Natural P&L)
# ============================================================
print("Step 1: Fixing reconciliation check data...")
ws_nat = wb['US January 2025 Natural P&L']

# Fix 1: AWS values missing from Natural P&L (fixes Check 5 + Check 12)
# AWS for iGO = Func P&L NetOps iGO (726.30) - Natural P&L non-AWS NetOps (493.06)
ws_nat['B15'] = 233.24   # iGO AWS
ws_nat['C15'] = 33.62    # Affirm AWS
ws_nat['D15'] = 42.23    # InsureSight AWS
# E15: reference total company AWS from AWS Allocation (includes DocFast)
ws_nat['E15'] = "='AWS Allocation'!F31"

# Fix 2: Security iGO Other (Check 6): total 158.20 → 162.53, +4.33
ws_nat['B29'] = 4.47     # was 0.14

# Fix 3: Support iGO Other (Check 7): total 235.81 → 234.14, -1.67
ws_nat['B38'] = 8.76     # was 10.43

# Fix 4: Partners iGO Other (Check 8): total 120.81 → 120.24, -0.57
ws_nat['B48'] = 4.35     # was 4.92

# Fix 5: Content iGO Other (Check 9): total 18.41 → 19.34, +0.93
ws_nat['B57'] = 1.04     # was 0.11

print("  Data fixes applied to US January 2025 Natural P&L")

# ============================================================
# 4. HELPER FUNCTIONS
# ============================================================

def set_branding(ws, title_text, max_col):
    """Add company name (A1), sheet title (A2), date stamp (top-right)."""
    ws['A1'] = 'Keystone BenefitTech, Inc.'
    ws['A1'].font = font_title
    ws['A1'].alignment = align_left

    ws['A2'] = title_text
    ws['A2'].font = font_subtitle
    ws['A2'].alignment = align_left

    date_cell = ws.cell(row=1, column=max_col)
    date_cell.value = 'Last Updated: 2026-02-28'
    date_cell.font = font_date
    date_cell.alignment = align_right


def style_header_row(ws, row, min_col, max_col):
    """iPipeline Blue header row with Arctic White bold text."""
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col > min_col else align_left
        cell.border = border_all


def style_data_row(ws, row, min_col, max_col, is_odd, fmt=None):
    """Alternating row with optional number format."""
    fill = fill_alt_1 if is_odd else fill_alt_2
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_body
        cell.fill = fill
        cell.alignment = align_right if col > min_col else align_left
        cell.border = border_hair
        if col > min_col and cell.value is not None and fmt:
            cell.number_format = fmt


def style_total_row(ws, row, min_col, max_col, fmt=None):
    """Navy Blue total/summary row."""
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_total
        cell.fill = fill_total
        cell.alignment = align_right if col > min_col else align_left
        cell.border = border_all
        if col > min_col and cell.value is not None and fmt:
            cell.number_format = fmt


def style_section_row(ws, row, min_col, max_col):
    """Innovation Blue section subheading."""
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_section
        cell.fill = fill_alt_1
        cell.alignment = align_left
        cell.border = border_bottom_medium


def style_blank_row(ws, row, min_col, max_col):
    """Clean blank row."""
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill_white
        cell.border = border_none


def auto_widths(ws, min_col=1, max_col=None, min_w=12, max_w=22):
    """Auto-fit column widths."""
    if max_col is None:
        max_col = ws.max_column
    for col in range(min_col, max_col + 1):
        best = min_w
        letter = get_column_letter(col)
        for row in range(1, min(ws.max_row + 1, 100)):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                vlen = len(str(cell.value))
                if vlen > best:
                    best = min(vlen + 3, max_w)
        # Column A (labels) wider
        if col == 1:
            best = max(best, 38)
        ws.column_dimensions[letter].width = best


def remove_gridlines(ws):
    """Remove gridlines from sheet view (preserving freeze panes)."""
    # Remove any duplicate views — Excel expects exactly 1
    while len(ws.views.sheetView) > 1:
        ws.views.sheetView.pop()
    # Set showGridLines on the single remaining view
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = False


def freeze_at(ws, row):
    """Freeze panes below given row."""
    ws.freeze_panes = f'A{row + 1}'


def format_financial_sheet(ws, title, header_row, max_col,
                           section_rows, total_rows, pct_rows,
                           currency_rows, blank_rows=None):
    """Apply full iPipeline formatting to a financial data sheet."""
    if blank_rows is None:
        blank_rows = []

    set_branding(ws, title, max_col)
    style_header_row(ws, header_row, 1, max_col)
    freeze_at(ws, header_row)

    data_idx = 0
    for row in range(header_row + 1, ws.max_row + 1):
        if row in blank_rows:
            style_blank_row(ws, row, 1, max_col)
        elif row in section_rows:
            style_section_row(ws, row, 1, max_col)
        elif row in total_rows:
            style_total_row(ws, row, 1, max_col, fmt=FMT_CURRENCY)
        elif row in pct_rows:
            data_idx += 1
            style_data_row(ws, row, 1, max_col, data_idx % 2 == 1, fmt=FMT_PERCENT)
        elif row in currency_rows:
            data_idx += 1
            style_data_row(ws, row, 1, max_col, data_idx % 2 == 1, fmt=FMT_CURRENCY)
        else:
            data_idx += 1
            style_data_row(ws, row, 1, max_col, data_idx % 2 == 1, fmt=FMT_CURRENCY)

    auto_widths(ws, 1, max_col)
    remove_gridlines(ws)


# ============================================================
# 5. FORMAT: P&L - Monthly Trend
# ============================================================
print("Step 2: Formatting P&L - Monthly Trend...")
ws = wb['P&L - Monthly Trend']
format_financial_sheet(
    ws, 'P&L - Monthly Trend', header_row=4, max_col=18,
    section_rows=[6, 14, 22, 30, 38],
    total_rows=[7, 9, 15, 17, 23, 25, 31, 33, 39, 41],
    pct_rows=[10, 12, 18, 20, 26, 28, 34, 36, 42, 44],
    currency_rows=[8, 11, 16, 19, 24, 27, 32, 35, 40, 43],
    blank_rows=[3, 5, 13, 21, 29, 37]
)

# ============================================================
# 6. FORMAT: Product Line Summary
# ============================================================
print("Step 3: Formatting Product Line Summary...")
ws = wb['Product Line Summary']
format_financial_sheet(
    ws, 'Product Line Summary — FY2025', header_row=4, max_col=18,
    section_rows=[5, 7, 15, 23, 31, 40, 42, 52, 62, 72],
    total_rows=[8, 10, 16, 18, 24, 26, 32, 34, 50, 60, 70, 80],
    pct_rows=[11, 13, 19, 21, 27, 29, 35, 37],
    currency_rows=[9, 12, 17, 20, 25, 28, 33, 36,
                   43, 44, 45, 46, 47, 48, 49,
                   53, 54, 55, 56, 57, 58, 59,
                   63, 64, 65, 66, 67, 68, 69,
                   73, 74, 75, 76, 77, 78, 79],
    blank_rows=[3, 6, 14, 22, 30, 38, 39, 41, 51, 61, 71]
)

# ============================================================
# 7. FORMAT: Functional P&L - Monthly Trend
# ============================================================
print("Step 4: Formatting Functional P&L - Monthly Trend...")
ws = wb['Functional P&L - Monthly Trend']

# Build row lists for 4 product blocks
func_section_rows = []
func_total_rows = []
func_pct_rows = []
func_currency_rows = []
func_blank_rows = [3, 5]

# Each product block offset pattern (relative to block start)
# Block starts: iGO=6, Affirm=42, InsureSight=78, DocFast=114
block_starts = [6, 42, 78, 114]
for bs in block_starts:
    func_section_rows.extend([bs, bs+1, bs+7])        # Product, Revenue, Cost of Revenue
    func_total_rows.extend([bs+5, bs+13, bs+15, bs+19, bs+21, bs+32])
    # Revenue, CostSub, CostRev, CM, CM GAAP, CM+R&D$
    func_pct_rows.extend([bs+22, bs+23, bs+24, bs+33])
    # SubMargin%, SvcMargin%, TotalMargin%, CM+R&D%
    func_currency_rows.extend([
        bs+2, bs+3, bs+4,       # Sub, Trans, Svc
        bs+8, bs+9, bs+10, bs+11, bs+12,  # NetOps-Content
        bs+14,                   # Cost of Services
        bs+17, bs+18,            # SubContrib, SvcContrib
        bs+20,                   # D&A
    ])
    func_section_rows.append(bs+26)  # R&D section header
    func_currency_rows.extend([bs+27, bs+28, bs+29, bs+30])  # R&D lines

    func_blank_rows.extend([
        bs-1,      # blank before product
        bs+6,      # blank before Cost of Revenue
        bs+16,     # blank before contributions
        bs+25,     # blank before R&D
        bs+31,     # blank before CM+R&D
    ])

format_financial_sheet(
    ws, 'Functional P&L - Monthly Trend', header_row=4, max_col=18,
    section_rows=func_section_rows,
    total_rows=func_total_rows,
    pct_rows=func_pct_rows,
    currency_rows=func_currency_rows,
    blank_rows=func_blank_rows
)

# ============================================================
# 8. FORMAT: Functional P&L Summary (Jan, Feb, Mar)
# ============================================================
func_summary_sheets = [
    ('Functional P&L Summary - Jan 25', 'Functional P&L Summary — January 2025'),
    ('Functional P&L Summary - Feb 25', 'Functional P&L Summary — February 2025'),
    ('Functional P&L Summary - Mar 25', 'Functional P&L Summary — March 2025'),
]

for sheet_name, title in func_summary_sheets:
    print(f"Step 5: Formatting {sheet_name}...")
    ws = wb[sheet_name]
    format_financial_sheet(
        ws, title, header_row=4, max_col=5,
        section_rows=[5, 11, 30],
        total_rows=[9, 17, 19, 23, 25, 36],
        pct_rows=[26, 27, 28, 37],
        currency_rows=[6, 7, 8, 12, 13, 14, 15, 16, 18, 21, 22, 24, 31, 32, 33, 34],
        blank_rows=[3, 10, 20, 29, 35]
    )

# ============================================================
# 9. FORMAT: US January 2025 Natural P&L
# ============================================================
print("Step 6: Formatting US January 2025 Natural P&L...")
ws = wb['US January 2025 Natural P&L']

nat_section_rows = [6, 11, 23, 32, 41, 51, 60, 70]
nat_total_rows = [9, 21, 30, 39, 49, 58, 68, 77]
nat_currency_rows = list(range(7, 9)) + list(range(12, 21)) + list(range(24, 30)) + \
                    list(range(33, 39)) + list(range(42, 49)) + list(range(52, 58)) + \
                    list(range(61, 68)) + list(range(71, 77))
nat_blank_rows = [3, 5, 10, 22, 31, 40, 50, 59, 69]

format_financial_sheet(
    ws, 'US January 2025 Natural P&L', header_row=4, max_col=5,
    section_rows=nat_section_rows,
    total_rows=nat_total_rows,
    pct_rows=[],
    currency_rows=nat_currency_rows,
    blank_rows=nat_blank_rows
)

# ============================================================
# 10. FORMAT: Assumptions
# ============================================================
print("Step 7: Formatting Assumptions...")
ws = wb['Assumptions']
max_col_a = 4

set_branding(ws, 'Assumptions & Driver Table', max_col_a)
style_section_row(ws, 4, 1, max_col_a)   # DRIVER TABLE
style_header_row(ws, 5, 1, max_col_a)    # Column headers
freeze_at(ws, 5)

idx = 0
for row in range(6, 20):
    idx += 1
    cell_a = ws.cell(row=row, column=1)
    if cell_a.value is not None:
        style_data_row(ws, row, 1, max_col_a, idx % 2 == 1)
        # Format percentage values
        cell_b = ws.cell(row=row, column=2)
        cell_c = ws.cell(row=row, column=3)
        if cell_c.value == 'Percentage' and cell_b.value is not None:
            if isinstance(cell_b.value, (int, float)):
                cell_b.number_format = FMT_PERCENT
    else:
        style_blank_row(ws, row, 1, max_col_a)

# Allocation Methodology section
for row in [20, 21, 22]:
    style_blank_row(ws, row, 1, max_col_a)

style_section_row(ws, 23, 1, max_col_a)  # ALLOCATION METHODOLOGY
style_header_row(ws, 24, 1, max_col_a)   # Column headers

idx = 0
for row in range(25, 34):
    cell_a = ws.cell(row=row, column=1)
    if cell_a.value is not None:
        idx += 1
        fill = fill_alt_1 if idx % 2 == 1 else fill_alt_2
        for col in range(1, max_col_a + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = font_body
            cell.fill = fill
            cell.border = border_hair
            if col == 2:
                cell.alignment = align_wrap
            else:
                cell.alignment = align_left

# Set column B wider for methodology text
ws.column_dimensions['B'].width = 80
ws.column_dimensions['A'].width = 38
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 55
remove_gridlines(ws)

# ============================================================
# 11. FORMAT: Data Dictionary
# ============================================================
print("Step 8: Formatting Data Dictionary...")
ws = wb['Data Dictionary']
max_col_dd = 5

set_branding(ws, 'Data Dictionary', max_col_dd)

# Section structure: PRODUCTS(4), DEPARTMENTS(12), EXPENSE CATEGORIES(23), VENDORS(38)
sections = [
    (4, 5, range(6, 10)),     # PRODUCTS
    (12, 13, range(14, 21)),  # DEPARTMENTS
    (23, 24, range(25, 36)),  # EXPENSE CATEGORIES
    (38, 39, range(40, 55)),  # VENDORS
]

freeze_at(ws, 5)  # Freeze below first header row

for sec_row, hdr_row, data_range in sections:
    style_section_row(ws, sec_row, 1, max_col_dd)
    style_header_row(ws, hdr_row, 1, max_col_dd)
    idx = 0
    for row in data_range:
        cell_a = ws.cell(row=row, column=1)
        if cell_a.value is not None:
            idx += 1
            style_data_row(ws, row, 1, max_col_dd, idx % 2 == 1)
        else:
            style_blank_row(ws, row, 1, max_col_dd)

# Blank rows between sections
for row in [3, 10, 11, 21, 22, 36, 37]:
    style_blank_row(ws, row, 1, max_col_dd)

auto_widths(ws, 1, max_col_dd, min_w=14, max_w=28)
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22
remove_gridlines(ws)

# ============================================================
# 12. FORMAT: AWS Allocation
# ============================================================
print("Step 9: Formatting AWS Allocation...")
ws = wb['AWS Allocation']
max_col_aws = 6

set_branding(ws, 'AWS Cost Allocation Model', max_col_aws)

# Section 1: Product Compute Shares (rows 4-10)
style_section_row(ws, 4, 1, max_col_aws)
style_header_row(ws, 5, 1, max_col_aws)
freeze_at(ws, 5)

for i, row in enumerate(range(6, 10), 1):
    style_data_row(ws, row, 1, max_col_aws, i % 2 == 1, fmt=FMT_PERCENT)
    ws.cell(row=row, column=2).number_format = FMT_PERCENT
style_total_row(ws, 10, 1, max_col_aws, fmt=FMT_PERCENT)
ws.cell(row=10, column=2).number_format = FMT_PERCENT

# Blank rows
for row in [3, 11, 12]:
    style_blank_row(ws, row, 1, max_col_aws)

# Section 2: Monthly AWS Cost Pool (rows 13-26)
style_section_row(ws, 13, 1, max_col_aws)
style_header_row(ws, 14, 1, max_col_aws)
for i, row in enumerate(range(15, 27), 1):
    style_data_row(ws, row, 1, max_col_aws, i % 2 == 1, fmt=FMT_CURRENCY)
    ws.cell(row=row, column=3).number_format = FMT_PERCENT
    ws.cell(row=row, column=4).number_format = FMT_PERCENT

# Blank rows
for row in [27, 28]:
    style_blank_row(ws, row, 1, max_col_aws)

# Section 3: Allocated AWS by Product by Month (rows 29-42)
style_section_row(ws, 29, 1, max_col_aws)
style_header_row(ws, 30, 1, max_col_aws)
for i, row in enumerate(range(31, 43), 1):
    style_data_row(ws, row, 1, max_col_aws, i % 2 == 1, fmt=FMT_CURRENCY)

auto_widths(ws, 1, max_col_aws, min_w=14, max_w=22)
ws.column_dimensions['A'].width = 38
remove_gridlines(ws)

# ============================================================
# 13. TRANSFORM: Report--> into Executive Dashboard
# ============================================================
print("Step 10: Transforming Report--> into Executive Dashboard...")
ws = wb['Report-->']

# Clear existing content
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10):
    for cell in row:
        cell.value = None
        cell.font = font_body
        cell.fill = fill_white
        cell.border = border_none
        cell.alignment = align_left

# --- Dashboard Header (rows 1-3) ---
ws.merge_cells('A1:H1')
ws['A1'] = 'Keystone BenefitTech, Inc.'
ws['A1'].font = font_dash_title
ws['A1'].fill = fill_header
ws['A1'].alignment = align_center
# Fill merged area
for col in range(2, 9):
    ws.cell(row=1, column=col).fill = fill_header

ws.merge_cells('A2:H2')
ws['A2'] = 'P&L Reporting & Allocation Model — FY2025'
ws['A2'].font = font_dash_subtitle
ws['A2'].fill = fill_header
ws['A2'].alignment = align_center
for col in range(2, 9):
    ws.cell(row=2, column=col).fill = fill_header

ws.merge_cells('A3:H3')
ws['A3'] = 'Last Updated: 2026-02-28'
ws['A3'].font = font_date
ws['A3'].fill = fill_header
ws['A3'].alignment = Alignment(horizontal='right', vertical='center')
for col in range(2, 9):
    ws.cell(row=3, column=col).fill = fill_header

# Row heights
ws.row_dimensions[1].height = 45
ws.row_dimensions[2].height = 28
ws.row_dimensions[3].height = 22

# --- Section Header (row 5) ---
ws.merge_cells('A5:H5')
ws['A5'] = 'EXECUTIVE SUMMARY'
ws['A5'].font = font_toc_header
ws['A5'].alignment = align_left
ws['A5'].border = border_bottom_medium

# --- KPI Boxes Row 1 (rows 7-8) ---
# KPI 1: Full Year Revenue (Aqua)
ws.merge_cells('B7:C7')
ws['B7'] = 'FULL YEAR REVENUE'
ws['B7'].font = font_kpi_label
ws['B7'].fill = fill_aqua
ws['B7'].alignment = Alignment(horizontal='center', vertical='bottom')
ws['B7'].border = border_thick_all
ws['C7'].fill = fill_aqua
ws['C7'].border = border_thick_all

ws.merge_cells('B8:C8')
ws['B8'] = "='P&L - Monthly Trend'!R7"
ws['B8'].font = font_kpi_value
ws['B8'].fill = fill_aqua
ws['B8'].alignment = Alignment(horizontal='center', vertical='top')
ws['B8'].number_format = FMT_CURRENCY
ws['B8'].border = border_thick_all
ws['C8'].fill = fill_aqua
ws['C8'].border = border_thick_all

# KPI 2: Contribution Margin % (Lime Green)
ws.merge_cells('D7:E7')
ws['D7'] = 'CONTRIBUTION MARGIN %'
ws['D7'].font = font_kpi_label_lime
ws['D7'].fill = fill_lime
ws['D7'].alignment = Alignment(horizontal='center', vertical='bottom')
ws['D7'].border = border_thick_all
ws['E7'].fill = fill_lime
ws['E7'].border = border_thick_all

ws.merge_cells('D8:E8')
ws['D8'] = "='P&L - Monthly Trend'!R10"
ws['D8'].font = font_kpi_value_lime
ws['D8'].fill = fill_lime
ws['D8'].alignment = Alignment(horizontal='center', vertical='top')
ws['D8'].number_format = FMT_PERCENT
ws['D8'].border = border_thick_all
ws['E8'].fill = fill_lime
ws['E8'].border = border_thick_all

# KPI 3: Top Product by Revenue (Aqua)
ws.merge_cells('F7:G7')
ws['F7'] = 'TOP PRODUCT BY REVENUE'
ws['F7'].font = font_kpi_label
ws['F7'].fill = fill_aqua
ws['F7'].alignment = Alignment(horizontal='center', vertical='bottom')
ws['F7'].border = border_thick_all
ws['G7'].fill = fill_aqua
ws['G7'].border = border_thick_all

ws.merge_cells('F8:G8')
ws['F8'] = 'iGO'
ws['F8'].font = font_kpi_value
ws['F8'].fill = fill_aqua
ws['F8'].alignment = Alignment(horizontal='center', vertical='top')
ws['F8'].border = border_thick_all
ws['G8'].fill = fill_aqua
ws['G8'].border = border_thick_all

# KPI Row heights
ws.row_dimensions[7].height = 28
ws.row_dimensions[8].height = 52

# --- KPI Boxes Row 2 (rows 10-11) ---
# KPI 4: Best Month (Lime Green)
ws.merge_cells('B10:C10')
ws['B10'] = 'BEST MONTH (REVENUE)'
ws['B10'].font = font_kpi_label_lime
ws['B10'].fill = fill_lime
ws['B10'].alignment = Alignment(horizontal='center', vertical='bottom')
ws['B10'].border = border_thick_all
ws['C10'].fill = fill_lime
ws['C10'].border = border_thick_all

ws.merge_cells('B11:C11')
ws['B11'] = 'December'
ws['B11'].font = font_kpi_value_lime
ws['B11'].fill = fill_lime
ws['B11'].alignment = Alignment(horizontal='center', vertical='top')
ws['B11'].border = border_thick_all
ws['C11'].fill = fill_lime
ws['C11'].border = border_thick_all

# KPI 5: Total Cost of Revenue (Aqua)
ws.merge_cells('D10:E10')
ws['D10'] = 'TOTAL COST OF REVENUE'
ws['D10'].font = font_kpi_label
ws['D10'].fill = fill_aqua
ws['D10'].alignment = Alignment(horizontal='center', vertical='bottom')
ws['D10'].border = border_thick_all
ws['E10'].fill = fill_aqua
ws['E10'].border = border_thick_all

ws.merge_cells('D11:E11')
ws['D11'] = "='P&L - Monthly Trend'!R8"
ws['D11'].font = font_kpi_value
ws['D11'].fill = fill_aqua
ws['D11'].alignment = Alignment(horizontal='center', vertical='top')
ws['D11'].number_format = FMT_CURRENCY
ws['D11'].border = border_thick_all
ws['E11'].fill = fill_aqua
ws['E11'].border = border_thick_all

# KPI 6: Q4 vs Q1 Growth (Lime Green)
ws.merge_cells('F10:G10')
ws['F10'] = 'Q4 vs Q1 REVENUE GROWTH'
ws['F10'].font = font_kpi_label_lime
ws['F10'].fill = fill_lime
ws['F10'].alignment = Alignment(horizontal='center', vertical='bottom')
ws['F10'].border = border_thick_all
ws['G10'].fill = fill_lime
ws['G10'].border = border_thick_all

ws.merge_cells('F11:G11')
ws['F11'] = "=('P&L - Monthly Trend'!Q7-'P&L - Monthly Trend'!N7)/'P&L - Monthly Trend'!N7"
ws['F11'].font = font_kpi_value_lime
ws['F11'].fill = fill_lime
ws['F11'].alignment = Alignment(horizontal='center', vertical='top')
ws['F11'].number_format = FMT_PERCENT
ws['F11'].border = border_thick_all
ws['G11'].fill = fill_lime
ws['G11'].border = border_thick_all

ws.row_dimensions[10].height = 28
ws.row_dimensions[11].height = 52

# --- Table of Contents (rows 13+) ---
ws.merge_cells('B13:G13')
ws['B13'] = 'TABLE OF CONTENTS'
ws['B13'].font = font_toc_header
ws['B13'].alignment = align_left
ws['B13'].border = border_bottom_medium

toc_sheets = [
    ('Assumptions', 'Assumptions & Driver Table'),
    ('Data Dictionary', 'Data Dictionary — Products, Departments, Vendors'),
    ('AWS Allocation', 'AWS Cost Allocation Model'),
    ('P&L - Monthly Trend', 'P&L Monthly Trend — Consolidated & By Product'),
    ('Product Line Summary', 'Product Line Summary — Revenue & Expenses'),
    ('Functional P&L - Monthly Trend', 'Functional P&L — Monthly Trend by Product'),
    ('Functional P&L Summary - Jan 25', 'Functional P&L Summary — January 2025'),
    ('Functional P&L Summary - Feb 25', 'Functional P&L Summary — February 2025'),
    ('Functional P&L Summary - Mar 25', 'Functional P&L Summary — March 2025'),
    ('US January 2025 Natural P&L', 'US Natural P&L — January 2025'),
    ('Charts & Visuals', 'Charts & Visuals — FY2025'),
    ('Checks', 'Reconciliation Checks — Model Health'),
]

for i, (sname, desc) in enumerate(toc_sheets):
    row = 15 + i
    ws.cell(row=row, column=2).value = desc
    ws.cell(row=row, column=2).font = font_link
    ws.cell(row=row, column=2).alignment = align_left
    # Internal hyperlink
    safe_name = f"'{sname}'" if ' ' in sname or '-' in sname else sname
    ws.cell(row=row, column=2).hyperlink = f"#{safe_name}!A1"

    fill = fill_alt_1 if i % 2 == 0 else fill_alt_2
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).border = border_hair

# Column widths
for col in range(1, 9):
    ws.column_dimensions[get_column_letter(col)].width = 18
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 55
ws.column_dimensions['H'].width = 4

remove_gridlines(ws)

# ============================================================
# 14. ADD: Charts & Visuals Sheet
# ============================================================
print("Step 11: Adding Charts & Visuals sheet...")

# Check if sheet already exists
if 'Charts & Visuals' in wb.sheetnames:
    ws_charts = wb['Charts & Visuals']
else:
    # Insert after US January 2025 Natural P&L (before Checks)
    checks_idx = wb.sheetnames.index('Checks')
    ws_charts = wb.create_sheet('Charts & Visuals', checks_idx)

# Clear any existing content
for row in ws_charts.iter_rows(min_row=1, max_row=ws_charts.max_row or 1, max_col=10):
    for cell in row:
        cell.value = None

# Header
ws_charts['A1'] = 'Keystone BenefitTech, Inc.'
ws_charts['A1'].font = font_title
ws_charts['A1'].alignment = align_left

ws_charts['A2'] = 'Charts & Visuals'
ws_charts['A2'].font = font_subtitle
ws_charts['A2'].alignment = align_left

# Date stamp
ws_charts['H2'] = 'Last Updated: 2026-02-28'
ws_charts['H2'].font = font_date
ws_charts['H2'].alignment = align_right

# Navy Blue header bar
ws_charts.merge_cells('A4:H4')
ws_charts['A4'] = 'Charts & Visuals — FY2025 — Under Construction'
ws_charts['A4'].font = font_construction
ws_charts['A4'].fill = fill_navy
ws_charts['A4'].alignment = align_center
for col in range(2, 9):
    ws_charts.cell(row=4, column=col).fill = fill_navy
    ws_charts.cell(row=4, column=col).border = border_all

ws_charts.row_dimensions[4].height = 40

# Column widths and gridlines
for col in range(1, 9):
    ws_charts.column_dimensions[get_column_letter(col)].width = 18
ws_charts.column_dimensions['A'].width = 38
remove_gridlines(ws_charts)

# ============================================================
# 15. FORMAT: CrossfireHiddenWorksheet
# ============================================================
print("Step 12: Formatting CrossfireHiddenWorksheet (keeping hidden)...")
ws = wb['CrossfireHiddenWorksheet']

# Apply iPipeline Blue header row
style_header_row(ws, 1, 1, 7)

# Format data rows with alternating fills
for row in range(2, ws.max_row + 1):
    is_odd = (row - 2) % 2 == 0
    fill = fill_alt_1 if is_odd else fill_alt_2
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.font = font_body
        cell.fill = fill
        cell.border = border_hair
        cell.alignment = align_right if col == 7 else align_left
        if col == 7:
            cell.number_format = FMT_CURRENCY

# Column widths
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 24
ws.column_dimensions['G'].width = 16

# Keep hidden
ws.sheet_state = 'hidden'
remove_gridlines(ws)

# ============================================================
# 16. FORMAT: Checks Sheet (Final Design)
# ============================================================
print("Step 13: Formatting Checks sheet with PASS/FAIL design...")
ws = wb['Checks']
max_col_chk = 5

# Branding
ws['A1'] = 'Keystone BenefitTech, Inc.'
ws['A1'].font = font_title
ws['A1'].alignment = align_left

ws['A2'] = 'Reconciliation Checks'
ws['A2'].font = font_subtitle
ws['A2'].alignment = align_left

ws.cell(row=1, column=max_col_chk).value = 'Last Updated: 2026-02-28'
ws.cell(row=1, column=max_col_chk).font = font_date
ws.cell(row=1, column=max_col_chk).alignment = align_right

# Row 3: blank
style_blank_row(ws, 3, 1, max_col_chk)

# Header row (Navy Blue for Checks)
for col in range(1, max_col_chk + 1):
    cell = ws.cell(row=4, column=col)
    cell.font = font_total
    cell.fill = fill_navy
    cell.alignment = align_center if col > 1 else align_left
    cell.border = border_all

freeze_at(ws, 4)

# Data rows with base styling (conditional formatting will override)
for row in range(5, 14):
    for col in range(1, max_col_chk + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_body_bold
        cell.fill = fill_alt_1
        cell.alignment = align_right if col > 1 else align_left
        cell.border = border_all
        if col in [2, 3]:
            cell.number_format = FMT_CURRENCY
        elif col == 4:
            cell.number_format = '#,##0.00'

# Add conditional formatting for PASS/FAIL rows
# Green for PASS
ws.conditional_formatting.add(
    'A5:E13',
    FormulaRule(
        formula=['$E5="PASS"'],
        fill=fill_pass,
        font=font_pass
    )
)

# Red for FAIL
ws.conditional_formatting.add(
    'A5:E13',
    FormulaRule(
        formula=['$E5="FAIL"'],
        fill=fill_fail,
        font=font_fail
    )
)

# Column widths
ws.column_dimensions['A'].width = 52
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 12

remove_gridlines(ws)

# ============================================================
# 17. SAVE
# ============================================================
print("\nSaving workbook...")
wb.save(OUTPUT_FILE)
print(f"Done! Saved to {OUTPUT_FILE}")

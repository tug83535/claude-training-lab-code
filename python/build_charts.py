#!/usr/bin/env python3
"""
Keystone BenefitTech P&L Model — Charts & Visuals Builder
Builds world-class interactive charts on the Charts & Visuals sheet.

Architecture:
- Data Validation dropdown (cell B6) lets user pick a product
- Lookup table (rows 50-75) uses IF formulas to pull selected product's data
- Section 1: Single-product dashboard (3 charts driven by dropdown)
- Section 2: All-products comparison (3 charts, all 4 products)
- Section 3: Advanced analytics (2 charts, deeper views)
- All charts use iPipeline brand palette

Run: python3 python/build_charts.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont, RichTextProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# 1. LOAD WORKBOOK
# ============================================================
INPUT_FILE = 'excel/KeystoneBenefitTech_PL_Model.xlsx'
OUTPUT_FILE = 'excel/KeystoneBenefitTech_PL_Model.xlsx'

wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb['Charts & Visuals']

# Unmerge any existing merged cells on this sheet first
for merge in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merge))

# Clear existing content below row 3 (keep header rows 1-2)
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=20):
    for cell in row:
        cell.value = None
        cell.font = Font()
        cell.fill = PatternFill()
        cell.alignment = Alignment()
        cell.border = Border()

# Remove any existing charts
ws._charts = []

# ============================================================
# 2. BRAND CONSTANTS (match redesign_pl_model.py exactly)
# ============================================================
IPIPELINE_BLUE = '0B4779'
ARCTIC_WHITE = 'F9F9F9'
SOFT_NEUTRAL = 'F0F0EE'
NAVY_BLUE = '112E51'
INNOVATION_BLUE = '4B9BCB'
LIME_GREEN = 'BFF18C'
AQUA = '2BCCD3'
CHARCOAL = '161616'

# Chart-specific hex colors (with #)
CLR_BLUE = '0B4779'
CLR_NAVY = '112E51'
CLR_AQUA = '2BCCD3'
CLR_LIME = 'BFF18C'
CLR_INNOVATION = '4B9BCB'
CLR_CHARCOAL = '161616'
CLR_ORANGE = 'E8833A'
CLR_TEAL = '1A8A8A'

# Product colors (consistent across all charts)
PRODUCT_COLORS = {
    'iGO': CLR_BLUE,
    'Affirm': CLR_AQUA,
    'InsureSight': CLR_LIME,
    'DocFast': CLR_INNOVATION,
}

# Department colors
DEPT_COLORS = [CLR_BLUE, CLR_AQUA, CLR_LIME, CLR_INNOVATION, CLR_ORANGE, CLR_NAVY, CLR_TEAL]

# Fonts
font_title = Font(name='Arial', bold=True, color=IPIPELINE_BLUE, size=16)
font_subtitle = Font(name='Arial', bold=True, color=IPIPELINE_BLUE, size=12)
font_section = Font(name='Arial', bold=True, color=INNOVATION_BLUE, size=11)
font_date = Font(name='Arial', italic=True, color=INNOVATION_BLUE, size=9)
font_body = Font(name='Arial', color=CHARCOAL, size=10)
font_body_bold = Font(name='Arial', bold=True, color=CHARCOAL, size=10)
font_header = Font(name='Arial', bold=True, color=ARCTIC_WHITE, size=11)
font_dropdown_label = Font(name='Arial', bold=True, color=IPIPELINE_BLUE, size=11)
font_dropdown_value = Font(name='Arial', bold=True, color=NAVY_BLUE, size=14)
font_data_label = Font(name='Arial', color=CHARCOAL, size=9)

# Fills
fill_header = PatternFill(start_color=IPIPELINE_BLUE, end_color=IPIPELINE_BLUE, fill_type='solid')
fill_alt_1 = PatternFill(start_color=ARCTIC_WHITE, end_color=ARCTIC_WHITE, fill_type='solid')
fill_alt_2 = PatternFill(start_color=SOFT_NEUTRAL, end_color=SOFT_NEUTRAL, fill_type='solid')
fill_total = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type='solid')
fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
fill_dropdown = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')

# Borders
thin_blue = Side(style='thin', color=IPIPELINE_BLUE)
medium_blue = Side(style='medium', color=IPIPELINE_BLUE)
hair_blue = Side(style='hair', color=IPIPELINE_BLUE)
border_thin = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)
border_medium = Border(left=medium_blue, right=medium_blue, top=medium_blue, bottom=medium_blue)
border_hair = Border(left=hair_blue, right=hair_blue, top=hair_blue, bottom=hair_blue)

# Number formats
FMT_CURRENCY = '$#,##0.0"K"'
FMT_PERCENT = '0.0%'

# ============================================================
# 3. SHEET HEADER & BRANDING
# ============================================================
ws['A1'] = 'Keystone BenefitTech, Inc.'
ws['A1'].font = font_title
ws['A1'].fill = fill_white

ws['A2'] = 'Charts & Visuals — FY2025'
ws['A2'].font = font_subtitle
ws['A2'].fill = fill_white

ws['H2'] = 'Last Updated: 2026-02-28'
ws['H2'].font = font_date
ws['H2'].alignment = Alignment(horizontal='right')
ws['H2'].fill = fill_white

# Row 3: blank spacer
for col in range(1, 12):
    ws.cell(row=3, column=col).fill = fill_white

# ============================================================
# 4. DROPDOWN SELECTOR (Row 5-6)
# ============================================================
# Section header for dropdown
ws['A5'] = 'SELECT PRODUCT'
ws['A5'].font = Font(name='Arial', bold=True, color=INNOVATION_BLUE, size=13)
ws['A5'].fill = fill_white
ws['A5'].border = Border(bottom=medium_blue)
for col in range(2, 9):
    ws.cell(row=5, column=col).border = Border(bottom=medium_blue)
    ws.cell(row=5, column=col).fill = fill_white

# Dropdown label and value
ws['A6'] = 'Product:'
ws['A6'].font = font_dropdown_label
ws['A6'].alignment = Alignment(vertical='center')
ws['A6'].fill = fill_white

ws['B6'] = 'iGO'
ws['B6'].font = font_dropdown_value
ws['B6'].fill = fill_dropdown
ws['B6'].border = border_medium
ws['B6'].alignment = Alignment(horizontal='center', vertical='center')

# Data Validation dropdown
dv = DataValidation(
    type='list',
    formula1='"iGO,Affirm,InsureSight,DocFast"',
    allow_blank=False,
    showDropDown=False,
)
dv.error = 'Please select a valid product'
dv.errorTitle = 'Invalid Product'
dv.prompt = 'Select a product to view its charts'
dv.promptTitle = 'Product Selector'
ws.add_data_validation(dv)
dv.add(ws['B6'])

# Instruction text
ws['C6'] = '← Select a product from the dropdown. Charts in Section 1 update automatically.'
ws['C6'].font = Font(name='Arial', italic=True, color=INNOVATION_BLUE, size=9)
ws['C6'].alignment = Alignment(vertical='center')
ws['C6'].fill = fill_white

# Row 7: spacer
for col in range(1, 12):
    ws.cell(row=7, column=col).fill = fill_white

# ============================================================
# 5. LOOKUP DATA TABLES (rows 50-75, off-screen)
# ============================================================
# These formulas use IF() to pull the selected product's data
# from 'Product Line Summary' based on the dropdown in B6.
# Charts reference these cells so they update with the dropdown.

PLS = "'Product Line Summary'"  # sheet ref for formulas
PLT = "'P&L - Monthly Trend'"

# --- Row map for Product Line Summary ---
# Product:  Revenue  COGS   CM$   CM%   CM+R&D$  CM+R&D%
# iGO:       8       9      10    11    12       13
# Affirm:    16      17     18    19    20       21
# InsureSight:24     25     26    27    28       29
# DocFast:   32      33     34    35    36       37
#
# Dept expenses (Section 2 of PLS):
# iGO: NetOps=43, Security=44, Support=45, Partners=46, Content=47, R&D=48, ProdMgmt=49, Total=50
# Affirm: 53,54,55,56,57,58,59,60
# InsureSight: 63,64,65,66,67,68,69,70
# DocFast: 73,74,75,76,77,78,79,80

def build_if_formula(col_letter, igo_row, affirm_row, insure_row, docfast_row, sheet=PLS):
    """Build an IF formula that picks the right row based on B6 dropdown."""
    return (
        f'=IF($B$6="iGO",{sheet}!{col_letter}{igo_row},'
        f'IF($B$6="Affirm",{sheet}!{col_letter}{affirm_row},'
        f'IF($B$6="InsureSight",{sheet}!{col_letter}{insure_row},'
        f'{sheet}!{col_letter}{docfast_row})))'
    )

# Lookup table start row
LT_START = 50

# Row 49: Header labels
ws.cell(row=LT_START - 1, column=1).value = 'LOOKUP TABLE (DO NOT EDIT)'
ws.cell(row=LT_START - 1, column=1).font = Font(name='Arial', bold=True, color='999999', size=8)

# Column headers for lookup table (months)
months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
for i, m in enumerate(months):
    ws.cell(row=LT_START, column=i + 2).value = m
    ws.cell(row=LT_START, column=i + 2).font = font_data_label

ws.cell(row=LT_START, column=1).value = 'Metric'
ws.cell(row=LT_START, column=1).font = font_data_label

# Metric rows for lookup:
# LT_START+1 (51): Revenue
# LT_START+2 (52): Cost of Revenue
# LT_START+3 (53): Contribution Margin $
# LT_START+4 (54): Contribution Margin %
# LT_START+5 (55): CM + R&D $
# LT_START+6 (56): CM + R&D %

metric_rows = {
    'Revenue':                 (8, 16, 24, 32),
    'Cost of Revenue':         (9, 17, 25, 33),
    'Contribution Margin $':   (10, 18, 26, 34),
    'Contribution Margin %':   (11, 19, 27, 35),
    'CM + R&D $':              (12, 20, 28, 36),
    'CM + R&D %':              (13, 21, 29, 37),
}

row_offset = 1
for metric_name, (igo_r, aff_r, ins_r, doc_r) in metric_rows.items():
    r = LT_START + row_offset
    ws.cell(row=r, column=1).value = metric_name
    ws.cell(row=r, column=1).font = font_data_label
    for col_idx in range(2, 14):  # B through M (Jan-Dec)
        col_letter = get_column_letter(col_idx)
        ws.cell(row=r, column=col_idx).value = build_if_formula(
            col_letter, igo_r, aff_r, ins_r, doc_r
        )
        if '%' in metric_name:
            ws.cell(row=r, column=col_idx).number_format = FMT_PERCENT
        else:
            ws.cell(row=r, column=col_idx).number_format = FMT_CURRENCY
    row_offset += 1

# Department expense lookup (rows 57-63)
# LT_START+7 (57): NetOps
# LT_START+8 (58): Security
# ...etc
dept_rows = {
    'NetOps':               (43, 53, 63, 73),
    'Security':             (44, 54, 64, 74),
    'Support':              (45, 55, 65, 75),
    'Partners':             (46, 56, 66, 76),
    'Content':              (47, 57, 67, 77),
    'Research & Development': (48, 58, 68, 78),
    'Product Management':   (49, 59, 69, 79),
}

# Blank row separator
row_offset += 1  # skip a row
dept_label_row = LT_START + row_offset
ws.cell(row=dept_label_row, column=1).value = 'Department'
ws.cell(row=dept_label_row, column=1).font = font_data_label
for i, m in enumerate(months):
    ws.cell(row=dept_label_row, column=i + 2).value = m
    ws.cell(row=dept_label_row, column=i + 2).font = font_data_label

row_offset += 1
DEPT_START_ROW = LT_START + row_offset
for dept_name, (igo_r, aff_r, ins_r, doc_r) in dept_rows.items():
    r = LT_START + row_offset
    ws.cell(row=r, column=1).value = dept_name
    ws.cell(row=r, column=1).font = font_data_label
    for col_idx in range(2, 14):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=r, column=col_idx).value = build_if_formula(
            col_letter, igo_r, aff_r, ins_r, doc_r
        )
        ws.cell(row=r, column=col_idx).number_format = FMT_CURRENCY
    row_offset += 1

DEPT_END_ROW = LT_START + row_offset - 1

# --- Full Year totals for department pie chart (column N = sum of B:M) ---
for r in range(DEPT_START_ROW, DEPT_END_ROW + 1):
    ws.cell(row=r, column=14).value = f'=SUM(B{r}:M{r})'
    ws.cell(row=r, column=14).number_format = FMT_CURRENCY

# --- Full Year Revenue per product for pie chart (row 70-73) ---
PIE_DATA_ROW = LT_START + row_offset + 1
ws.cell(row=PIE_DATA_ROW, column=1).value = 'Product'
ws.cell(row=PIE_DATA_ROW, column=1).font = font_data_label
ws.cell(row=PIE_DATA_ROW, column=2).value = 'FY Revenue'
ws.cell(row=PIE_DATA_ROW, column=2).font = font_data_label
ws.cell(row=PIE_DATA_ROW, column=3).value = 'FY COGS'
ws.cell(row=PIE_DATA_ROW, column=3).font = font_data_label
ws.cell(row=PIE_DATA_ROW, column=4).value = 'FY CM$'
ws.cell(row=PIE_DATA_ROW, column=4).font = font_data_label

products_fy = [
    ('iGO', 8, 9, 10),
    ('Affirm', 16, 17, 18),
    ('InsureSight', 24, 25, 26),
    ('DocFast', 32, 33, 34),
]

for i, (prod, rev_r, cogs_r, cm_r) in enumerate(products_fy):
    r = PIE_DATA_ROW + 1 + i
    ws.cell(row=r, column=1).value = prod
    ws.cell(row=r, column=1).font = font_data_label
    ws.cell(row=r, column=2).value = f'={PLS}!R{rev_r}'   # Column R = 2025 Total
    ws.cell(row=r, column=2).number_format = FMT_CURRENCY
    ws.cell(row=r, column=3).value = f'={PLS}!R{cogs_r}'
    ws.cell(row=r, column=3).number_format = FMT_CURRENCY
    ws.cell(row=r, column=4).value = f'={PLS}!R{cm_r}'
    ws.cell(row=r, column=4).number_format = FMT_CURRENCY

PIE_PROD_START = PIE_DATA_ROW + 1
PIE_PROD_END = PIE_DATA_ROW + 4

# --- Department totals across ALL products for stacked bar (row 78+) ---
DEPT_ALL_ROW = PIE_PROD_END + 2
ws.cell(row=DEPT_ALL_ROW, column=1).value = 'Department'
ws.cell(row=DEPT_ALL_ROW, column=1).font = font_data_label
ws.cell(row=DEPT_ALL_ROW, column=2).value = 'iGO'
ws.cell(row=DEPT_ALL_ROW, column=2).font = font_data_label
ws.cell(row=DEPT_ALL_ROW, column=3).value = 'Affirm'
ws.cell(row=DEPT_ALL_ROW, column=3).font = font_data_label
ws.cell(row=DEPT_ALL_ROW, column=4).value = 'InsureSight'
ws.cell(row=DEPT_ALL_ROW, column=4).font = font_data_label
ws.cell(row=DEPT_ALL_ROW, column=5).value = 'DocFast'
ws.cell(row=DEPT_ALL_ROW, column=5).font = font_data_label

# PLS dept rows: iGO 43-49, Affirm 53-59, InsureSight 63-69, DocFast 73-79
dept_names_list = ['NetOps', 'Security', 'Support', 'Partners', 'Content', 'R&D', 'Product Mgmt']
igo_dept_rows = [43, 44, 45, 46, 47, 48, 49]
aff_dept_rows = [53, 54, 55, 56, 57, 58, 59]
ins_dept_rows = [63, 64, 65, 66, 67, 68, 69]
doc_dept_rows = [73, 74, 75, 76, 77, 78, 79]

DEPT_ALL_START = DEPT_ALL_ROW + 1
for i, dept_name in enumerate(dept_names_list):
    r = DEPT_ALL_START + i
    ws.cell(row=r, column=1).value = dept_name
    ws.cell(row=r, column=1).font = font_data_label
    # FY totals (column R) from Product Line Summary
    ws.cell(row=r, column=2).value = f'={PLS}!R{igo_dept_rows[i]}'
    ws.cell(row=r, column=2).number_format = FMT_CURRENCY
    ws.cell(row=r, column=3).value = f'={PLS}!R{aff_dept_rows[i]}'
    ws.cell(row=r, column=3).number_format = FMT_CURRENCY
    ws.cell(row=r, column=4).value = f'={PLS}!R{ins_dept_rows[i]}'
    ws.cell(row=r, column=4).number_format = FMT_CURRENCY
    ws.cell(row=r, column=5).value = f'={PLS}!R{doc_dept_rows[i]}'
    ws.cell(row=r, column=5).number_format = FMT_CURRENCY

DEPT_ALL_END = DEPT_ALL_START + 6

# --- Revenue by product by month for multi-product bar (row 87+) ---
MULTI_REV_ROW = DEPT_ALL_END + 2
ws.cell(row=MULTI_REV_ROW, column=1).value = 'Month'
ws.cell(row=MULTI_REV_ROW, column=1).font = font_data_label
for i, m in enumerate(months):
    ws.cell(row=MULTI_REV_ROW + 1 + i, column=1).value = m
    ws.cell(row=MULTI_REV_ROW + 1 + i, column=1).font = font_data_label

# Product revenues by month (cols B-E: iGO, Affirm, InsureSight, DocFast)
rev_rows_by_product = {'iGO': 8, 'Affirm': 16, 'InsureSight': 24, 'DocFast': 32}
for p_idx, (prod_name, pls_row) in enumerate(rev_rows_by_product.items()):
    col = p_idx + 2  # B, C, D, E
    ws.cell(row=MULTI_REV_ROW, column=col).value = prod_name
    ws.cell(row=MULTI_REV_ROW, column=col).font = font_data_label
    for m_idx in range(12):
        pls_col = get_column_letter(m_idx + 2)  # B through M in PLS
        r = MULTI_REV_ROW + 1 + m_idx
        ws.cell(row=r, column=col).value = f'={PLS}!{pls_col}{pls_row}'
        ws.cell(row=r, column=col).number_format = FMT_CURRENCY

MULTI_REV_START = MULTI_REV_ROW + 1
MULTI_REV_END = MULTI_REV_ROW + 12

# --- CM% by product by month for multi-product line (row 102+) ---
MULTI_CM_ROW = MULTI_REV_END + 2
ws.cell(row=MULTI_CM_ROW, column=1).value = 'Month'
ws.cell(row=MULTI_CM_ROW, column=1).font = font_data_label
for i, m in enumerate(months):
    ws.cell(row=MULTI_CM_ROW + 1 + i, column=1).value = m
    ws.cell(row=MULTI_CM_ROW + 1 + i, column=1).font = font_data_label

cm_pct_rows = {'iGO': 11, 'Affirm': 19, 'InsureSight': 27, 'DocFast': 35}
for p_idx, (prod_name, pls_row) in enumerate(cm_pct_rows.items()):
    col = p_idx + 2
    ws.cell(row=MULTI_CM_ROW, column=col).value = prod_name
    ws.cell(row=MULTI_CM_ROW, column=col).font = font_data_label
    for m_idx in range(12):
        pls_col = get_column_letter(m_idx + 2)
        r = MULTI_CM_ROW + 1 + m_idx
        ws.cell(row=r, column=col).value = f'={PLS}!{pls_col}{pls_row}'
        ws.cell(row=r, column=col).number_format = FMT_PERCENT

MULTI_CM_START = MULTI_CM_ROW + 1
MULTI_CM_END = MULTI_CM_ROW + 12

# --- Revenue vs Total Expenses per product for waterfall (row 117+) ---
WATERFALL_ROW = MULTI_CM_END + 2
ws.cell(row=WATERFALL_ROW, column=1).value = 'Product'
ws.cell(row=WATERFALL_ROW, column=1).font = font_data_label
ws.cell(row=WATERFALL_ROW, column=2).value = 'Revenue'
ws.cell(row=WATERFALL_ROW, column=2).font = font_data_label
ws.cell(row=WATERFALL_ROW, column=3).value = 'Cost of Revenue'
ws.cell(row=WATERFALL_ROW, column=3).font = font_data_label
ws.cell(row=WATERFALL_ROW, column=4).value = 'Contribution Margin'
ws.cell(row=WATERFALL_ROW, column=4).font = font_data_label

waterfall_data = [
    ('iGO', 8, 9, 10),
    ('Affirm', 16, 17, 18),
    ('InsureSight', 24, 25, 26),
    ('DocFast', 32, 33, 34),
]

WATERFALL_START = WATERFALL_ROW + 1
for i, (prod, rev_r, cogs_r, cm_r) in enumerate(waterfall_data):
    r = WATERFALL_START + i
    ws.cell(row=r, column=1).value = prod
    ws.cell(row=r, column=1).font = font_data_label
    ws.cell(row=r, column=2).value = f'={PLS}!R{rev_r}'
    ws.cell(row=r, column=2).number_format = FMT_CURRENCY
    ws.cell(row=r, column=3).value = f'={PLS}!R{cogs_r}'
    ws.cell(row=r, column=3).number_format = FMT_CURRENCY
    ws.cell(row=r, column=4).value = f'={PLS}!R{cm_r}'
    ws.cell(row=r, column=4).number_format = FMT_CURRENCY

WATERFALL_END = WATERFALL_START + 3


# ============================================================
# 6. CHART HELPER FUNCTIONS
# ============================================================
def make_chart_font(size=10, bold=False, color='161616'):
    """Create a drawing font for chart text elements."""
    return CharacterProperties(
        latin=DrawingFont(typeface='Arial'),
        sz=size * 100,
        b=bold,
        solidFill=color,
    )


def style_chart_base(chart, title_text, width=15, height=10):
    """Apply consistent iPipeline branding to any chart."""
    chart.width = width
    chart.height = height
    chart.style = 2

    # Title
    chart.title = title_text
    chart.title.txPr = RichText(
        p=[Paragraph(
            pPr=ParagraphProperties(
                defRPr=make_chart_font(size=12, bold=True, color=IPIPELINE_BLUE)
            ),
            endParaRPr=make_chart_font(size=12, bold=True, color=IPIPELINE_BLUE),
        )]
    )

    # Legend font
    if chart.legend:
        chart.legend.txPr = RichText(
            p=[Paragraph(
                pPr=ParagraphProperties(
                    defRPr=make_chart_font(size=9, color=CHARCOAL)
                ),
                endParaRPr=make_chart_font(size=9, color=CHARCOAL),
            )]
        )

    # Remove chart border for clean look
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.line = LineProperties(noFill=True)

    return chart


def style_axis(axis, fmt=None, title_text=None):
    """Style a chart axis with iPipeline fonts."""
    if axis is None:
        return
    axis.txPr = RichText(
        p=[Paragraph(
            pPr=ParagraphProperties(
                defRPr=make_chart_font(size=8, color=CHARCOAL)
            ),
            endParaRPr=make_chart_font(size=8, color=CHARCOAL),
        )]
    )
    if fmt:
        axis.numFmt = fmt
    if title_text:
        axis.title = title_text


def color_series(series, hex_color):
    """Set a series fill and line to a specific color."""
    series.graphicalProperties.solidFill = hex_color
    series.graphicalProperties.line.solidFill = hex_color


# ============================================================
# 7. SECTION 1: SINGLE-PRODUCT DASHBOARD (Dropdown-Driven)
# ============================================================
# Section header
ws['A8'] = 'SECTION 1: SELECTED PRODUCT DASHBOARD'
ws['A8'].font = Font(name='Arial', bold=True, color=IPIPELINE_BLUE, size=13)
ws['A8'].fill = fill_white
ws['A8'].border = Border(bottom=medium_blue)
for col in range(2, 9):
    ws.cell(row=8, column=col).border = Border(bottom=medium_blue)
    ws.cell(row=8, column=col).fill = fill_white

# --- Chart 1: Revenue vs Cost of Revenue (Grouped Bar) ---
chart1 = BarChart()
chart1.type = 'col'
chart1.grouping = 'clustered'

# Revenue series (row 51)
rev_ref = Reference(ws, min_col=2, max_col=13, min_row=LT_START + 1)
rev_cats = Reference(ws, min_col=2, max_col=13, min_row=LT_START)
chart1.add_data(rev_ref, from_rows=True, titles_from_data=False)
chart1.set_categories(rev_cats)
chart1.series[0].title = openpyxl.chart.series.SeriesLabel(v='Revenue')
color_series(chart1.series[0], CLR_BLUE)

# COGS series (row 52)
cogs_ref = Reference(ws, min_col=2, max_col=13, min_row=LT_START + 2)
chart1.add_data(cogs_ref, from_rows=True, titles_from_data=False)
chart1.series[1].title = openpyxl.chart.series.SeriesLabel(v='Cost of Revenue')
color_series(chart1.series[1], CLR_AQUA)

style_chart_base(chart1, 'Revenue vs Cost of Revenue — Monthly', width=17, height=11)
style_axis(chart1.x_axis)
style_axis(chart1.y_axis, fmt=FMT_CURRENCY, title_text='$K')
chart1.y_axis.delete = False
chart1.legend.position = 'b'

ws.add_chart(chart1, 'A10')

# --- Chart 2: Contribution Margin % Trend (Line) ---
chart2 = LineChart()

cm_pct_ref = Reference(ws, min_col=2, max_col=13, min_row=LT_START + 4)
cm_cats = Reference(ws, min_col=2, max_col=13, min_row=LT_START)
chart2.add_data(cm_pct_ref, from_rows=True, titles_from_data=False)
chart2.set_categories(cm_cats)
chart2.series[0].title = openpyxl.chart.series.SeriesLabel(v='Contribution Margin %')
color_series(chart2.series[0], CLR_BLUE)
chart2.series[0].graphicalProperties.line.width = 28000  # ~2pt

style_chart_base(chart2, 'Contribution Margin % — Monthly Trend', width=17, height=11)
style_axis(chart2.x_axis)
style_axis(chart2.y_axis, fmt=FMT_PERCENT, title_text='Margin %')
chart2.y_axis.delete = False
chart2.legend.position = 'b'

ws.add_chart(chart2, 'A27')

# --- Chart 3: Expense Breakdown by Department (Pie) ---
chart3 = PieChart()

dept_labels = Reference(ws, min_col=1, min_row=DEPT_START_ROW, max_row=DEPT_END_ROW)
dept_data = Reference(ws, min_col=14, min_row=DEPT_START_ROW, max_row=DEPT_END_ROW)
chart3.add_data(dept_data, titles_from_data=False)
chart3.set_categories(dept_labels)
chart3.series[0].title = openpyxl.chart.series.SeriesLabel(v='Expense by Department')

# Color each slice
for i, clr in enumerate(DEPT_COLORS):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = clr
    chart3.series[0].data_points.append(pt)

# Data labels
chart3.series[0].dLbls = DataLabelList()
chart3.series[0].dLbls.showPercent = True
chart3.series[0].dLbls.showCatName = True
chart3.series[0].dLbls.showVal = False
chart3.series[0].dLbls.numFmt = '0.0%'
chart3.series[0].dLbls.txPr = RichText(
    p=[Paragraph(
        pPr=ParagraphProperties(
            defRPr=make_chart_font(size=8, color=CHARCOAL)
        ),
        endParaRPr=make_chart_font(size=8, color=CHARCOAL),
    )]
)

style_chart_base(chart3, 'Expense Breakdown by Department — Full Year', width=17, height=11)

ws.add_chart(chart3, 'A44')


# ============================================================
# 8. SECTION 2: ALL-PRODUCTS COMPARISON
# ============================================================
SEC2_ROW = 62  # visual start row for section 2 header

ws.cell(row=SEC2_ROW, column=1).value = 'SECTION 2: ALL-PRODUCTS COMPARISON'
ws.cell(row=SEC2_ROW, column=1).font = Font(name='Arial', bold=True, color=IPIPELINE_BLUE, size=13)
ws.cell(row=SEC2_ROW, column=1).fill = fill_white
ws.cell(row=SEC2_ROW, column=1).border = Border(bottom=medium_blue)
for col in range(2, 9):
    ws.cell(row=SEC2_ROW, column=col).border = Border(bottom=medium_blue)
    ws.cell(row=SEC2_ROW, column=col).fill = fill_white

# --- Chart 4: Revenue by Product (Grouped Bar, Monthly) ---
chart4 = BarChart()
chart4.type = 'col'
chart4.grouping = 'clustered'

# Categories (month labels)
multi_cats = Reference(ws, min_col=1, min_row=MULTI_REV_START, max_row=MULTI_REV_END)

# Add each product as a series
product_list = ['iGO', 'Affirm', 'InsureSight', 'DocFast']
product_colors_list = [CLR_BLUE, CLR_AQUA, CLR_LIME, CLR_INNOVATION]

for p_idx in range(4):
    col = p_idx + 2  # B, C, D, E
    data_ref = Reference(ws, min_col=col, min_row=MULTI_REV_ROW, max_row=MULTI_REV_END)
    chart4.add_data(data_ref, titles_from_data=True)

chart4.set_categories(multi_cats)

for i, clr in enumerate(product_colors_list):
    color_series(chart4.series[i], clr)

style_chart_base(chart4, 'Monthly Revenue by Product — FY2025', width=17, height=11)
style_axis(chart4.x_axis)
style_axis(chart4.y_axis, fmt=FMT_CURRENCY, title_text='Revenue ($K)')
chart4.y_axis.delete = False
chart4.legend.position = 'b'

ws.add_chart(chart4, 'A64')

# --- Chart 5: Contribution Margin % by Product (Line, Monthly) ---
chart5 = LineChart()

multi_cm_cats = Reference(ws, min_col=1, min_row=MULTI_CM_START, max_row=MULTI_CM_END)
for p_idx in range(4):
    col = p_idx + 2
    data_ref = Reference(ws, min_col=col, min_row=MULTI_CM_ROW, max_row=MULTI_CM_END)
    chart5.add_data(data_ref, titles_from_data=True)

chart5.set_categories(multi_cm_cats)

for i, clr in enumerate(product_colors_list):
    color_series(chart5.series[i], clr)
    chart5.series[i].graphicalProperties.line.width = 28000

style_chart_base(chart5, 'Contribution Margin % by Product — FY2025', width=17, height=11)
style_axis(chart5.x_axis)
style_axis(chart5.y_axis, fmt=FMT_PERCENT, title_text='Margin %')
chart5.y_axis.delete = False
chart5.legend.position = 'b'

ws.add_chart(chart5, 'A81')

# --- Chart 6: Full Year Revenue Mix (Pie) ---
chart6 = PieChart()

pie_labels = Reference(ws, min_col=1, min_row=PIE_PROD_START, max_row=PIE_PROD_END)
pie_data = Reference(ws, min_col=2, min_row=PIE_PROD_START, max_row=PIE_PROD_END)
chart6.add_data(pie_data, titles_from_data=False)
chart6.set_categories(pie_labels)
chart6.series[0].title = openpyxl.chart.series.SeriesLabel(v='Revenue Share')

for i, clr in enumerate(product_colors_list):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = clr
    chart6.series[0].data_points.append(pt)

chart6.series[0].dLbls = DataLabelList()
chart6.series[0].dLbls.showPercent = True
chart6.series[0].dLbls.showCatName = True
chart6.series[0].dLbls.showVal = False
chart6.series[0].dLbls.numFmt = '0.0%'
chart6.series[0].dLbls.txPr = RichText(
    p=[Paragraph(
        pPr=ParagraphProperties(
            defRPr=make_chart_font(size=9, color=CHARCOAL)
        ),
        endParaRPr=make_chart_font(size=9, color=CHARCOAL),
    )]
)

style_chart_base(chart6, 'Full Year Revenue Mix — FY2025', width=17, height=11)

ws.add_chart(chart6, 'A98')


# ============================================================
# 9. SECTION 3: ADVANCED ANALYTICS
# ============================================================
SEC3_ROW = 116

ws.cell(row=SEC3_ROW, column=1).value = 'SECTION 3: ADVANCED ANALYTICS'
ws.cell(row=SEC3_ROW, column=1).font = Font(name='Arial', bold=True, color=IPIPELINE_BLUE, size=13)
ws.cell(row=SEC3_ROW, column=1).fill = fill_white
ws.cell(row=SEC3_ROW, column=1).border = Border(bottom=medium_blue)
for col in range(2, 9):
    ws.cell(row=SEC3_ROW, column=col).border = Border(bottom=medium_blue)
    ws.cell(row=SEC3_ROW, column=col).fill = fill_white

# --- Chart 7: Revenue vs COGS vs CM by Product (Grouped Bar) ---
chart7 = BarChart()
chart7.type = 'col'
chart7.grouping = 'clustered'

wf_cats = Reference(ws, min_col=1, min_row=WATERFALL_START, max_row=WATERFALL_END)

# Revenue series
wf_rev = Reference(ws, min_col=2, min_row=WATERFALL_ROW, max_row=WATERFALL_END)
chart7.add_data(wf_rev, titles_from_data=True)
color_series(chart7.series[0], CLR_BLUE)

# COGS series
wf_cogs = Reference(ws, min_col=3, min_row=WATERFALL_ROW, max_row=WATERFALL_END)
chart7.add_data(wf_cogs, titles_from_data=True)
color_series(chart7.series[1], CLR_AQUA)

# CM series
wf_cm = Reference(ws, min_col=4, min_row=WATERFALL_ROW, max_row=WATERFALL_END)
chart7.add_data(wf_cm, titles_from_data=True)
color_series(chart7.series[2], CLR_LIME)

chart7.set_categories(wf_cats)

style_chart_base(chart7, 'Revenue vs Cost of Revenue vs Contribution Margin — FY2025', width=17, height=11)
style_axis(chart7.x_axis)
style_axis(chart7.y_axis, fmt=FMT_CURRENCY, title_text='$K')
chart7.y_axis.delete = False
chart7.legend.position = 'b'

ws.add_chart(chart7, 'A118')

# --- Chart 8: Department Cost Distribution Across Products (Stacked Bar) ---
chart8 = BarChart()
chart8.type = 'col'
chart8.grouping = 'stacked'

dept_all_cats = Reference(ws, min_col=1, min_row=DEPT_ALL_START, max_row=DEPT_ALL_END)

# Each product is a series
for p_idx in range(4):
    col = p_idx + 2
    data_ref = Reference(ws, min_col=col, min_row=DEPT_ALL_ROW, max_row=DEPT_ALL_END)
    chart8.add_data(data_ref, titles_from_data=True)

chart8.set_categories(dept_all_cats)

for i, clr in enumerate(product_colors_list):
    color_series(chart8.series[i], clr)

style_chart_base(chart8, 'Department Cost Distribution by Product — FY2025', width=17, height=11)
style_axis(chart8.x_axis)
style_axis(chart8.y_axis, fmt=FMT_CURRENCY, title_text='$K')
chart8.y_axis.delete = False
chart8.legend.position = 'b'

ws.add_chart(chart8, 'A135')


# ============================================================
# 10. COLUMN WIDTHS & SHEET SETTINGS
# ============================================================
ws.column_dimensions['A'].width = 38
for col_idx in range(2, 15):
    ws.column_dimensions[get_column_letter(col_idx)].width = 14
ws.column_dimensions['H'].width = 22  # for date stamp

# Remove gridlines
ws.sheet_view.showGridLines = False

# Freeze at row 4 (keep header visible)
ws.freeze_panes = 'A4'

# Set sheet tab color to iPipeline blue
ws.sheet_properties.tabColor = IPIPELINE_BLUE

# Fill visible area background white
for row_num in range(3, 160):
    for col_idx in range(1, 12):
        c = ws.cell(row=row_num, column=col_idx)
        if c.fill == PatternFill() or c.fill.start_color.rgb == '00000000':
            c.fill = fill_white


# ============================================================
# 11. SAVE
# ============================================================
wb.save(OUTPUT_FILE)
print(f'Charts & Visuals built successfully → {OUTPUT_FILE}')
print(f'  8 charts created across 3 sections')
print(f'  Dropdown selector in B6 (iGO, Affirm, InsureSight, DocFast)')
print(f'  Lookup data tables in rows {LT_START}-{WATERFALL_END}')

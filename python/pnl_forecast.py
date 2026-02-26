#!/usr/bin/env python3
"""
pnl_forecast.py — Rolling Forecast Model
==========================================

PURPOSE: Time-series forecasting for P&L line items using multiple methods:
         - Simple moving average
         - Exponential smoothing (Holt-Winters)
         - Linear trend with seasonal decomposition
         - Scenario-based projections (optimistic/base/pessimistic)

USAGE:
    python pnl_forecast.py
    python pnl_forecast.py --months 6 --method ets
    python pnl_forecast.py --product iGO --export forecast.xlsx

    from pnl_forecast import PnLForecaster
    fc = PnLForecaster("KeystoneBenefitTech_PL_Model.xlsx")
    result = fc.forecast(periods=6)
"""

import os
import sys
import argparse
from typing import Dict, List, Optional, Tuple

import pandas as pd
import numpy as np

try:
    from pnl_config import *
except ImportError:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from pnl_config import *


class PnLForecaster(PnLBase):
    """P&L forecasting engine with multiple methods."""

    METHODS = ["sma", "ets", "trend", "scenario"]

    def __init__(self, file_path: str = None, verbose: bool = True):
        super().__init__(verbose)
        self.file_path = file_path or SOURCE_FILE
        self.gl = None
        self.monthly = None

    def load(self) -> "PnLForecaster":
        """Load data and build monthly time series."""
        self.gl = self._load_gl(self.file_path)
        self._build_monthly_series()
        return self

    def _build_monthly_series(self):
        """Aggregate GL into monthly time series by various dimensions."""
        gl = self.gl[self.gl["Month"].notna()].copy()

        # Total monthly
        total = gl.groupby("Month").agg(
            Spend=("Amount", "sum"),
            Abs_Spend=("Abs_Amount", "sum"),
            Txn_Count=("Amount", "count"),
            Vendors=("Vendor", "nunique"),
        ).reset_index()
        total["Month"] = total["Month"].astype(int)
        total = total.sort_values("Month")

        # By product
        by_product = gl[gl["Product"].isin(PRODUCTS)].groupby(["Month", "Product"]).agg(
            Spend=("Amount", "sum"),
            Abs_Spend=("Abs_Amount", "sum"),
        ).reset_index()
        by_product["Month"] = by_product["Month"].astype(int)

        # By department
        by_dept = gl[gl["Department"].isin(DEPARTMENTS)].groupby(["Month", "Department"]).agg(
            Spend=("Amount", "sum"),
            Abs_Spend=("Abs_Amount", "sum"),
        ).reset_index()
        by_dept["Month"] = by_dept["Month"].astype(int)

        self.monthly = {
            "total": total,
            "by_product": by_product,
            "by_department": by_dept,
        }

    # ─────────────────────────────────────────────────────────
    # METHOD 1: Simple Moving Average
    # ─────────────────────────────────────────────────────────

    def forecast_sma(self, series: pd.Series, periods: int = 3, window: int = 3) -> pd.DataFrame:
        """Simple moving average forecast."""
        values = series.values
        last_n = values[-window:]
        avg = np.mean(last_n)

        forecasts = []
        for i in range(periods):
            forecasts.append({
                "Period": len(values) + i + 1,
                "Forecast": avg,
                "Method": "SMA",
                "Lower": avg * 0.85,
                "Upper": avg * 1.15,
            })
        return pd.DataFrame(forecasts)

    # ─────────────────────────────────────────────────────────
    # METHOD 2: Exponential Smoothing
    # ─────────────────────────────────────────────────────────

    def forecast_ets(self, series: pd.Series, periods: int = 3, alpha: float = 0.3) -> pd.DataFrame:
        """Exponential smoothing (simple Holt method with trend)."""
        values = series.values.astype(float)
        n = len(values)

        if n < 3:
            return self.forecast_sma(series, periods)

        # Initialize
        level = values[0]
        trend = (values[-1] - values[0]) / max(n - 1, 1)
        beta = 0.1  # trend smoothing

        # Fit
        for i in range(1, n):
            prev_level = level
            level = alpha * values[i] + (1 - alpha) * (level + trend)
            trend = beta * (level - prev_level) + (1 - beta) * trend

        # Forecast
        residuals = []
        fit_level = values[0]
        fit_trend = trend
        for i in range(1, n):
            prev = fit_level
            fit_level = alpha * values[i] + (1 - alpha) * (fit_level + fit_trend)
            fit_trend = beta * (fit_level - prev) + (1 - beta) * fit_trend
            residuals.append(values[i] - (prev + fit_trend))

        std_residual = np.std(residuals) if residuals else abs(level) * 0.1

        forecasts = []
        for i in range(1, periods + 1):
            fc = level + trend * i
            forecasts.append({
                "Period": n + i,
                "Forecast": fc,
                "Method": "ETS",
                "Lower": fc - 1.96 * std_residual * np.sqrt(i),
                "Upper": fc + 1.96 * std_residual * np.sqrt(i),
            })
        return pd.DataFrame(forecasts)

    # ─────────────────────────────────────────────────────────
    # METHOD 3: Linear Trend
    # ─────────────────────────────────────────────────────────

    def forecast_trend(self, series: pd.Series, periods: int = 3) -> pd.DataFrame:
        """Linear regression trend forecast."""
        values = series.values.astype(float)
        n = len(values)
        x = np.arange(n)

        if n < 2:
            return self.forecast_sma(series, periods)

        # Fit linear regression
        slope, intercept = np.polyfit(x, values, 1)
        fitted = intercept + slope * x
        residuals = values - fitted
        std_resid = np.std(residuals)
        r_squared = 1 - (np.sum(residuals**2) / np.sum((values - np.mean(values))**2))

        forecasts = []
        for i in range(1, periods + 1):
            fc = intercept + slope * (n + i - 1)
            forecasts.append({
                "Period": n + i,
                "Forecast": fc,
                "Method": "Trend",
                "Lower": fc - 1.96 * std_resid,
                "Upper": fc + 1.96 * std_resid,
                "R_Squared": r_squared,
                "Slope_Monthly": slope,
            })
        return pd.DataFrame(forecasts)

    # ─────────────────────────────────────────────────────────
    # METHOD 4: Scenario-Based
    # ─────────────────────────────────────────────────────────

    def forecast_scenario(self, series: pd.Series, periods: int = 3) -> pd.DataFrame:
        """Three-scenario forecast: optimistic, base, pessimistic."""
        values = series.values.astype(float)
        n = len(values)

        # Base: last 3-month average
        base = np.mean(values[-min(3, n):])

        # Growth rate from trend
        if n >= 3:
            recent_growth = (values[-1] - values[-3]) / abs(values[-3]) if values[-3] != 0 else 0
            monthly_growth = recent_growth / 3
        else:
            monthly_growth = 0

        forecasts = []
        for i in range(1, periods + 1):
            optimistic = base * (1 + (monthly_growth + 0.02) * i)
            base_fc = base * (1 + monthly_growth * i)
            pessimistic = base * (1 + (monthly_growth - 0.02) * i)

            forecasts.append({
                "Period": n + i,
                "Forecast": base_fc,
                "Method": "Scenario",
                "Lower": pessimistic,
                "Upper": optimistic,
                "Optimistic": optimistic,
                "Base": base_fc,
                "Pessimistic": pessimistic,
            })
        return pd.DataFrame(forecasts)

    # ─────────────────────────────────────────────────────────
    # MAIN FORECAST
    # ─────────────────────────────────────────────────────────

    def forecast(self, periods: int = 3, method: str = "ets",
                 product: str = None, department: str = None) -> Dict[str, pd.DataFrame]:
        """
        Run forecast and return results.

        Args:
            periods: Number of months to forecast
            method: sma, ets, trend, or scenario
            product: Filter to specific product (None = total)
            department: Filter to specific department (None = total)
        """
        if self.gl is None:
            self.load()

        method_map = {
            "sma": self.forecast_sma,
            "ets": self.forecast_ets,
            "trend": self.forecast_trend,
            "scenario": self.forecast_scenario,
        }
        fc_func = method_map.get(method, self.forecast_ets)

        results = {}

        self._section(f"FORECAST — {method.upper()} | {periods} months ahead")

        if product:
            # Single product forecast
            data = self.monthly["by_product"]
            prod_data = data[data["Product"] == product].sort_values("Month")
            if len(prod_data) > 0:
                fc = fc_func(prod_data["Abs_Spend"], periods)
                fc["Dimension"] = product
                results[product] = fc
                self._print_forecast(product, prod_data["Abs_Spend"], fc)
        elif department:
            # Single department forecast
            data = self.monthly["by_department"]
            dept_data = data[data["Department"] == department].sort_values("Month")
            if len(dept_data) > 0:
                fc = fc_func(dept_data["Abs_Spend"], periods)
                fc["Dimension"] = department
                results[department] = fc
                self._print_forecast(department, dept_data["Abs_Spend"], fc)
        else:
            # Total + all products
            total = self.monthly["total"].sort_values("Month")
            fc = fc_func(total["Abs_Spend"], periods)
            fc["Dimension"] = "Total"
            results["Total"] = fc
            self._print_forecast("Total", total["Abs_Spend"], fc)

            for prod in PRODUCTS:
                prod_data = self.monthly["by_product"]
                ps = prod_data[prod_data["Product"] == prod].sort_values("Month")
                if len(ps) > 0:
                    fc = fc_func(ps["Abs_Spend"], periods)
                    fc["Dimension"] = prod
                    results[prod] = fc
                    self._print_forecast(prod, ps["Abs_Spend"], fc)

        return results

    def _print_forecast(self, label: str, actuals: pd.Series, forecast: pd.DataFrame):
        """Print formatted forecast results."""
        self._print(f"\n  {label}:")
        last_actual = actuals.iloc[-1] if len(actuals) > 0 else 0
        self._print(f"    Last actual:  {format_currency(last_actual)}")

        for _, row in forecast.iterrows():
            period = int(row["Period"])
            month_name = MONTH_ABBREVS[period - 1] if 1 <= period <= 12 else f"M{period}"
            fc = row["Forecast"]
            lo = row.get("Lower", fc * 0.9)
            hi = row.get("Upper", fc * 1.1)
            self._print(f"    {month_name}:  {format_currency(fc)}  "
                        f"(range: {format_currency(lo)} — {format_currency(hi)})")

    def export(self, results: Dict[str, pd.DataFrame], output_path: str = "forecast_output.xlsx"):
        """Export forecast results to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Forecast Summary"

        ws["A1"] = f"P&L Forecast — {FY_LABEL}"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws["A2"] = f"Generated: {PnLBase.timestamp()}"

        row = 4
        headers = ["Dimension", "Period", "Month", "Forecast", "Lower", "Upper", "Method"]
        for col, hdr in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E79", fill_type="solid")

        row = 5
        for dim_name, fc_df in results.items():
            for _, fc_row in fc_df.iterrows():
                period = int(fc_row["Period"])
                month_name = MONTH_ABBREVS[period - 1] if 1 <= period <= 12 else f"M{period}"
                ws.cell(row=row, column=1, value=dim_name)
                ws.cell(row=row, column=2, value=period)
                ws.cell(row=row, column=3, value=month_name)
                ws.cell(row=row, column=4, value=round(fc_row["Forecast"], 2))
                ws.cell(row=row, column=5, value=round(fc_row.get("Lower", 0), 2))
                ws.cell(row=row, column=6, value=round(fc_row.get("Upper", 0), 2))
                ws.cell(row=row, column=7, value=fc_row.get("Method", ""))
                row += 1

        # Also write actuals
        ws2 = wb.create_sheet("Actuals")
        ws2["A1"] = "Monthly Actuals"
        ws2["A1"].font = Font(bold=True, size=12, color="1F4E79")
        if self.monthly:
            total = self.monthly["total"]
            for col, hdr in enumerate(total.columns, 1):
                ws2.cell(row=3, column=col, value=hdr).font = Font(bold=True)
            for i, (_, r) in enumerate(total.iterrows(), 4):
                for col, val in enumerate(r, 1):
                    ws2.cell(row=i, column=col, value=val)

        wb.save(output_path)
        self._print(f"\nForecast exported: {output_path}", "OK")


def main():
    parser = argparse.ArgumentParser(description="P&L Rolling Forecast")
    parser.add_argument("--file", "-f", default=SOURCE_FILE)
    parser.add_argument("--months", "-m", type=int, default=3, help="Months to forecast")
    parser.add_argument("--method", default="ets", choices=PnLForecaster.METHODS)
    parser.add_argument("--product", "-p", default=None, help="Specific product")
    parser.add_argument("--department", "-d", default=None, help="Specific department")
    parser.add_argument("--export", "-e", default=None, help="Export to Excel")
    parser.add_argument("--all-methods", action="store_true", help="Run all forecast methods")
    args = parser.parse_args()

    fc = PnLForecaster(file_path=args.file)
    fc.load()

    if args.all_methods:
        for method in PnLForecaster.METHODS:
            fc.forecast(periods=args.months, method=method,
                       product=args.product, department=args.department)
    else:
        results = fc.forecast(periods=args.months, method=args.method,
                             product=args.product, department=args.department)
        if args.export:
            fc.export(results, args.export)


if __name__ == "__main__":
    main()

"""
Build all demo input files for Video 4 Python script demos.
Each script needs specific input files to run on camera.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import random
import csv

random.seed(42)
BASE = r"C:\Users\connor.atlee\RecTrial\Video4DemoFiles"

# ============================================================
# 1. COMPARE FILES — Two similar Excel files with differences
# ============================================================
def build_compare_files():
    # File A: Q1_Revenue_v1.xlsx
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Revenue"
    headers = ["Region", "Sales Rep", "Product", "Customer", "Amount", "Status"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)

    data = [
        ["Northeast", "Marcus Chen", "iGO", "MetLife", 125000, "Closed Won"],
        ["Southeast", "Aisha Johnson", "Affirm", "Prudential", 87500, "Closed Won"],
        ["West", "Sofia Rodriguez", "InsureSight", "Pacific Life", 210000, "Pending"],
        ["Midwest", "James Wilson", "DocFast", "Northwestern Mutual", 45000, "Closed Won"],
        ["Northeast", "Rachel Kim", "iGO", "MassMutual", 175000, "Closed Won"],
        ["Southeast", "David Okafor", "Affirm", "Lincoln Financial", 92000, "Pipeline"],
        ["West", "Emily Chen", "InsureSight", "Transamerica", 156000, "Closed Won"],
        ["International", "Michael Thompson", "iGO", "Unum Group", 310000, "Closed Won"],
        ["Midwest", "Jessica Patel", "DocFast", "Aflac", 67000, "Pending"],
        ["Northeast", "Chris Morrison", "Affirm", "Guardian Life", 118000, "Closed Won"],
        ["Southeast", "Lisa Hernandez", "InsureSight", "Nationwide", 85000, "Closed Won"],
        ["West", "Andrew Walsh", "iGO", "Hartford Financial", 225000, "Closed Won"],
        ["International", "Sarah Mitchell", "Affirm", "Voya Financial", 143000, "Pipeline"],
        ["Midwest", "Tom Brown", "DocFast", "Equitable Holdings", 52000, "Closed Won"],
        ["Northeast", "Amanda Foster", "iGO", "Principal Financial", 195000, "Closed Won"],
    ]
    for i, row in enumerate(data, 2):
        for c, v in enumerate(row, 1):
            ws1.cell(row=i, column=c, value=v)

    wb1.save(os.path.join(BASE, "Q1_Revenue_v1.xlsx"))

    # File B: Q1_Revenue_v2.xlsx (8 differences)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Revenue"
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    for i, row in enumerate(data, 2):
        for c, v in enumerate(row, 1):
            ws2.cell(row=i, column=c, value=v)

    # Changes
    ws2.cell(row=2, column=5, value=132000)       # Amount changed
    ws2.cell(row=4, column=6, value="Closed Won")  # Status changed
    ws2.cell(row=6, column=5, value=98500)          # Amount changed
    ws2.cell(row=7, column=6, value="Pipeline")     # Status changed
    ws2.cell(row=9, column=5, value=320000)          # Amount changed
    ws2.cell(row=11, column=4, value="Aflac Inc")    # Customer name changed
    ws2.cell(row=14, column=6, value="Closed Won")   # Status changed
    # Add new row
    ws2.cell(row=17, column=1, value="International")
    ws2.cell(row=17, column=2, value="New Hire TBD")
    ws2.cell(row=17, column=3, value="iGO")
    ws2.cell(row=17, column=4, value="Zurich Insurance")
    ws2.cell(row=17, column=5, value=290000)
    ws2.cell(row=17, column=6, value="Pipeline")

    wb2.save(os.path.join(BASE, "Q1_Revenue_v2.xlsx"))
    print("1. Compare Files: Q1_Revenue_v1.xlsx + Q1_Revenue_v2.xlsx (8 diffs + 1 new row)")

# ============================================================
# 2. PDF EXTRACTOR — Need a PDF with tables
# We can't create a PDF from Python easily without extra libs,
# so we'll create a text file with instructions
# ============================================================
def build_pdf_note():
    note = os.path.join(BASE, "PDF_DEMO_README.txt")
    with open(note, "w") as f:
        f.write("PDF EXTRACTOR DEMO\n")
        f.write("==================\n\n")
        f.write("You need a PDF file with data tables for the demo.\n\n")
        f.write("Best options:\n")
        f.write("1. Any financial statement PDF (quarterly report, 10-K, etc.)\n")
        f.write("2. An invoice summary PDF\n")
        f.write("3. A vendor report PDF\n\n")
        f.write("Requirements:\n")
        f.write("- Must have SELECTABLE TEXT (not a scanned image)\n")
        f.write("- Must have at least one visible table with rows and columns\n")
        f.write("- Save it to this folder as 'sample_report.pdf'\n\n")
        f.write("The script command will be:\n")
        f.write("  python pdf_extractor.py sample_report.pdf\n")
    print("2. PDF Extractor: PDF_DEMO_README.txt (need real PDF for demo)")

# ============================================================
# 3. FUZZY LOOKUP — Two vendor lists with mismatched names
# ============================================================
def build_fuzzy_files():
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Our Vendors"
    ws1.cell(row=1, column=1, value="Vendor Name")
    ws1.cell(row=1, column=2, value="Account Number")

    our_vendors = [
        ["MetLife Insurance Company", "ACCT-001"],
        ["JP Morgan Chase", "ACCT-002"],
        ["Amazon Web Services", "ACCT-003"],
        ["Microsoft Corporation", "ACCT-004"],
        ["Deloitte Consulting LLP", "ACCT-005"],
        ["Delta Air Lines Inc", "ACCT-006"],
        ["Marriott International", "ACCT-007"],
        ["Salesforce Inc", "ACCT-008"],
        ["LinkedIn Corporation", "ACCT-009"],
        ["Workday Inc", "ACCT-010"],
        ["KPMG LLP", "ACCT-011"],
        ["Baker McKenzie Law Firm", "ACCT-012"],
        ["CrowdStrike Holdings", "ACCT-013"],
        ["Prudential Financial Inc", "ACCT-014"],
        ["Northwestern Mutual Life", "ACCT-015"],
    ]
    for i, (name, acct) in enumerate(our_vendors, 2):
        ws1.cell(row=i, column=1, value=name)
        ws1.cell(row=i, column=2, value=acct)

    wb1.save(os.path.join(BASE, "our_vendor_list.xlsx"))

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Bank Vendors"
    ws2.cell(row=1, column=1, value="Vendor Name")
    ws2.cell(row=1, column=2, value="Last Payment")

    bank_vendors = [
        ["METLIFE INS CO", "$125,000"],
        ["JPMorgan Chase & Co", "$87,500"],
        ["AWS Amazon", "$78,000"],
        ["MICROSOFT CORP", "$45,000"],
        ["Deloitte & Touche", "$125,000"],
        ["DELTA AIRLINES", "$8,900"],
        ["Marriott Hotels", "$5,600"],
        ["SALESFORCE.COM", "$45,000"],
        ["LinkedIn Talent Solutions", "$15,000"],
        ["WORKDAY INC", "$18,000"],
        ["KPMG Audit Services", "$85,000"],
        ["Baker & McKenzie", "$45,000"],
        ["CROWDSTRIKE INC", "$22,000"],
        ["Prudential Financial", "$92,000"],
        ["NW Mutual Life Insurance", "$175,000"],
        ["UNKNOWN VENDOR XYZ", "$3,500"],
        ["Staples Office Supply", "$950"],
    ]
    for i, (name, amt) in enumerate(bank_vendors, 2):
        ws2.cell(row=i, column=1, value=name)
        ws2.cell(row=i, column=2, value=amt)

    wb2.save(os.path.join(BASE, "bank_vendor_list.xlsx"))
    print("3. Fuzzy Lookup: our_vendor_list.xlsx + bank_vendor_list.xlsx (15 vs 17 vendors)")

# ============================================================
# 4. BANK RECONCILER — Ledger + Bank Statement
# ============================================================
def build_bank_recon_files():
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "GL Entries"
    headers = ["Date", "Description", "Amount", "Reference"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)

    gl_entries = [
        ["01/05/2026", "AWS Monthly - January", -78000, "INV-2026-001"],
        ["01/10/2026", "MetLife Premium Collection", 125000, "REC-2026-001"],
        ["01/12/2026", "GitHub Enterprise License", -24800, "INV-2026-002"],
        ["01/15/2026", "Prudential Q1 Payment", 87500, "REC-2026-002"],
        ["01/18/2026", "Office Supplies - Staples", -950, "PO-2026-015"],
        ["01/22/2026", "Delta Airlines - Travel", -3200, "EXP-2026-001"],
        ["01/25/2026", "Salesforce Annual License", -45000, "INV-2026-003"],
        ["01/28/2026", "Pacific Life Contract", 210000, "REC-2026-003"],
        ["02/01/2026", "Marriott - Conference", -5600, "EXP-2026-002"],
        ["02/05/2026", "Northwestern Mutual Premium", 175000, "REC-2026-004"],
        ["02/10/2026", "KPMG Audit Services Q1", -85000, "INV-2026-004"],
        ["02/14/2026", "LinkedIn Recruiter Annual", -15000, "INV-2026-005"],
        ["02/18/2026", "MassMutual Collection", 92000, "REC-2026-005"],
        ["02/22/2026", "Workday HR Platform", -18000, "INV-2026-006"],
        ["02/25/2026", "Deloitte Consulting Phase 1", -125000, "INV-2026-007"],
        ["03/01/2026", "Guardian Life Premium", 118000, "REC-2026-006"],
        ["03/05/2026", "CrowdStrike Security Annual", -22000, "INV-2026-008"],
        ["03/10/2026", "Equitable Holdings Payment", 52000, "REC-2026-007"],
        ["03/15/2026", "Baker McKenzie Legal Fees", -45000, "INV-2026-009"],
        ["03/20/2026", "Voya Financial Collection", 143000, "REC-2026-008"],
    ]
    for i, row in enumerate(gl_entries, 2):
        for c, v in enumerate(row, 1):
            ws1.cell(row=i, column=c, value=v)

    wb1.save(os.path.join(BASE, "gl_ledger.xlsx"))

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Bank Statement"
    headers2 = ["Date", "Description", "Amount", "Bank Ref"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)

    bank_entries = [
        ["01/05/2026", "AMAZON WEB SERVICES AWS", -78000, "BNK-90001"],
        ["01/10/2026", "METLIFE INS PREMIUM DEP", 125000, "BNK-90002"],
        ["01/12/2026", "GITHUB INC", -24800, "BNK-90003"],
        ["01/15/2026", "PRUDENTIAL FINL DEP", 87500, "BNK-90004"],
        ["01/19/2026", "STAPLES STORE #4521", -950, "BNK-90005"],
        ["01/23/2026", "DELTA AIR LINES ATL", -3200, "BNK-90006"],
        ["01/25/2026", "SALESFORCE COM INC", -45000, "BNK-90007"],
        ["01/28/2026", "PACIFIC LIFE INS DEP", 210000, "BNK-90008"],
        ["02/01/2026", "MARRIOTT INTL HOTEL", -5600, "BNK-90009"],
        ["02/05/2026", "NW MUTUAL LIFE DEP", 175000, "BNK-90010"],
        ["02/10/2026", "KPMG LLP AUDIT SVC", -85000, "BNK-90011"],
        ["02/14/2026", "LINKEDIN CORP TALENT", -15000, "BNK-90012"],
        ["02/18/2026", "MASSMUTUAL FINANCIAL DEP", 92000, "BNK-90013"],
        ["02/22/2026", "WORKDAY INC SAAS", -18000, "BNK-90014"],
        ["02/25/2026", "DELOITTE TOUCHE CONSULT", -125000, "BNK-90015"],
        ["03/01/2026", "GUARDIAN LIFE INS DEP", 118000, "BNK-90016"],
        ["03/05/2026", "CROWDSTRIKE HLDGS SEC", -22000, "BNK-90017"],
        ["03/10/2026", "EQUITABLE HLDGS DEP", 52000, "BNK-90018"],
        ["03/15/2026", "BAKER MCKENZIE LAW", -45000, "BNK-90019"],
        ["03/20/2026", "VOYA FINANCIAL DEP", 143000, "BNK-90020"],
        ["03/25/2026", "UNKNOWN ACH DEPOSIT", 8750, "BNK-90021"],
        ["03/28/2026", "NSF CHECK RETURN", -2100, "BNK-90022"],
    ]
    for i, row in enumerate(bank_entries, 2):
        for c, v in enumerate(row, 1):
            ws2.cell(row=i, column=c, value=v)

    wb2.save(os.path.join(BASE, "bank_statement.xlsx"))
    print("4. Bank Reconciler: gl_ledger.xlsx (20 entries) + bank_statement.xlsx (22 entries)")

# ============================================================
# 5. AGING REPORT — Invoice data with dates and amounts
# ============================================================
def build_aging_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    headers = ["Invoice #", "Customer", "Invoice Date", "Amount", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    customers = ["MetLife", "Prudential", "New York Life", "Northwestern Mutual",
                 "MassMutual", "Lincoln Financial", "Transamerica", "Pacific Life",
                 "Unum Group", "Aflac", "Guardian Life", "Nationwide"]

    invoices = []
    inv_num = 10001
    # Current (within 30 days of today ~April 2026)
    for _ in range(8):
        invoices.append([f"INV-{inv_num}", random.choice(customers), f"03/{random.randint(5,28):02d}/2026",
                        random.choice([15000, 25000, 45000, 67000, 85000, 125000]), "Open"])
        inv_num += 1
    # 31-60 days
    for _ in range(5):
        invoices.append([f"INV-{inv_num}", random.choice(customers), f"02/{random.randint(1,15):02d}/2026",
                        random.choice([22000, 35000, 48000, 92000, 175000]), "Open"])
        inv_num += 1
    # 61-90 days
    for _ in range(4):
        invoices.append([f"INV-{inv_num}", random.choice(customers), f"01/{random.randint(5,20):02d}/2026",
                        random.choice([18000, 55000, 78000, 210000]), "Overdue"])
        inv_num += 1
    # 90+ days
    for _ in range(3):
        invoices.append([f"INV-{inv_num}", random.choice(customers), f"12/{random.randint(1,28):02d}/2025",
                        random.choice([35000, 95000, 310000]), "Overdue"])
        inv_num += 1

    random.shuffle(invoices)
    for i, row in enumerate(invoices, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=i, column=c, value=v)

    wb.save(os.path.join(BASE, "open_invoices.xlsx"))
    print(f"5. Aging Report: open_invoices.xlsx ({len(invoices)} invoices across all aging buckets)")

# ============================================================
# 6. VARIANCE DECOMPOSITION — Product data with actual/budget units and prices
# ============================================================
def build_variance_decomp_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Data"
    headers = ["Product", "Budget Units", "Budget Price", "Actual Units", "Actual Price"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    products = [
        ["iGO Enterprise", 500, 1250, 520, 1300],
        ["iGO Standard", 1200, 450, 1150, 475],
        ["Affirm Pro", 800, 750, 850, 725],
        ["Affirm Basic", 2000, 200, 2100, 195],
        ["InsureSight Analytics", 350, 1800, 380, 1750],
        ["InsureSight Core", 900, 600, 870, 620],
        ["DocFast Premium", 600, 500, 550, 520],
        ["DocFast Lite", 1500, 150, 1600, 145],
    ]
    for i, row in enumerate(products, 2):
        for c, v in enumerate(row, 1):
            ws.cell(row=i, column=c, value=v)

    wb.save(os.path.join(BASE, "product_budget_vs_actual.xlsx"))
    print(f"6. Variance Decomposition: product_budget_vs_actual.xlsx ({len(products)} products)")

# ============================================================
# 7. FORECAST ROLLFORWARD — Historical monthly data
# ============================================================
def build_forecast_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Revenue"
    ws.cell(row=1, column=1, value="Date")
    ws.cell(row=1, column=2, value="Revenue")

    # 12 months of historical data with upward trend
    base = 8500000
    for m in range(1, 13):
        ws.cell(row=m+1, column=1, value=f"{m:02d}/01/2025")
        rev = base + (m * 250000) + random.randint(-200000, 300000)
        ws.cell(row=m+1, column=2, value=rev)

    # 3 months of 2026 actuals
    for m in range(1, 4):
        ws.cell(row=m+13, column=1, value=f"{m:02d}/01/2026")
        rev = 11500000 + (m * 300000) + random.randint(-150000, 250000)
        ws.cell(row=m+13, column=2, value=rev)

    wb.save(os.path.join(BASE, "monthly_revenue_history.xlsx"))
    print("7. Forecast Rollforward: monthly_revenue_history.xlsx (15 months of data)")

# ============================================================
# 8. VARIANCE ANALYSIS — Multiple budget files in a folder
# ============================================================
def build_variance_analysis_files():
    folder = os.path.join(BASE, "budget_files")
    departments = [
        ("Engineering", 485000, 502000),
        ("Marketing", 165000, 152800),
        ("Sales", 95000, 93400),
        ("Finance", 72000, 68000),
        ("HR", 58000, 61000),
        ("Operations", 125000, 118000),
        ("Legal", 85000, 92000),
    ]

    for dept, budget, actual in departments:
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget vs Actual"
        ws.cell(row=1, column=1, value="Category")
        ws.cell(row=1, column=2, value="Budget")
        ws.cell(row=1, column=3, value="Actual")

        categories = ["Salary & Benefits", "Software & Tools", "Travel", "Consulting", "Other"]
        remaining_bud = budget
        remaining_act = actual

        for i, cat in enumerate(categories):
            if i < len(categories) - 1:
                bud_pct = random.uniform(0.15, 0.35)
                act_pct = bud_pct + random.uniform(-0.05, 0.05)
                cat_bud = round(budget * bud_pct)
                cat_act = round(actual * act_pct)
                remaining_bud -= cat_bud
                remaining_act -= cat_act
            else:
                cat_bud = remaining_bud
                cat_act = remaining_act

            ws.cell(row=i+2, column=1, value=cat)
            ws.cell(row=i+2, column=2, value=cat_bud)
            ws.cell(row=i+2, column=3, value=cat_act)

        ws.cell(row=8, column=1, value="TOTAL")
        ws.cell(row=8, column=2, value=budget)
        ws.cell(row=8, column=3, value=actual)

        fname = f"{dept}_Q1_Budget.xlsx"
        wb.save(os.path.join(folder, fname))

    print(f"8. Variance Analysis: {len(departments)} department budget files in budget_files/")

# ============================================================
# RUN ALL
# ============================================================
print("Building Video 4 demo input files...")
print("=" * 50)
build_compare_files()
build_pdf_note()
build_fuzzy_files()
build_bank_recon_files()
build_aging_file()
build_variance_decomp_file()
build_forecast_file()
build_variance_analysis_files()

print("=" * 50)
print(f"\nAll files saved to: {BASE}")
print("\nFiles created:")
for f in sorted(os.listdir(BASE)):
    fpath = os.path.join(BASE, f)
    if os.path.isfile(fpath):
        size = os.path.getsize(fpath)
        print(f"  {f} ({size:,} bytes)")
    elif os.path.isdir(fpath):
        count = len(os.listdir(fpath))
        print(f"  {f}/ ({count} files)")

print("\nNOTE: You still need a sample PDF for the PDF Extractor demo.")
print("Find any financial PDF with tables and save it as:")
print(f"  {os.path.join(BASE, 'sample_report.pdf')}")

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import random

random.seed(42)
wb = Workbook()

BLUE = "0B4779"
NAVY = "112E51"

# ============================================================
# SHEET 1: Q1 Revenue (~40 rows, intentionally messy)
# ============================================================
ws1 = wb.active
ws1.title = "Q1 Revenue"

hdrs = ["Region", "Sales Rep", "Product", "Customer", "Date", "Amount", "Status", "Commission %", "Notes"]
for c, h in enumerate(hdrs, 1):
    ws1.cell(row=1, column=c, value=h)

reps = ["Sarah Mitchell", "James Rodriguez", "Emily Chen", "Michael Thompson", "Rachel Kim",
        "David Okafor", "Jessica Patel", " Andrew Walsh", "Lisa Hernandez ", "Chris Morrison"]
products = ["iGO", "Affirm", "InsureSight", "DocFast"]
regions = ["Northeast", "Southeast", "Midwest", "West", "International"]
customers = ["MetLife", "Prudential", "New York Life", "Northwestern Mutual", "MassMutual",
             "Lincoln Financial", "Transamerica", "Pacific Life", "Unum Group", "Aflac",
             "Principal Financial", "Guardian Life", "Nationwide", "Hartford Financial",
             "Voya Financial", "Equitable Holdings", "Mutual of Omaha", "Securian Financial"]
statuses = ["Closed Won", "Closed Won", "Closed Won", "Pending", "Pipeline", "Closed Won"]
dates_all = ["01/15/2026", "2026-01-08", "Jan 12, 2026", "01/22/2026", "1/19/26",
             "01/28/2026", "2026-02-06", "Feb 17 2026", "02/03/2026", "2/22/26",
             "02/10/2026", "02/14/2026", "2026-03-03", "Mar 8 2026", "02/19/2026",
             "02/25/2026", "3/14/26", "03/01/2026", "03/05/2026", "2026-03-20",
             "03/11/2026", "Mar 26, 2026", "03/15/2026", "3/30/26", "03/18/2026",
             "03/22/2026", "03/25/2026", "03/28/2026"]
notes_pool = ["Enterprise renewal", "Multi-year contract", "Expansion deal", "New logo",
              "Upsell from DocFast", "Competitive displacement", "POC converted", "RFP winner",
              "Channel partner referral", "Executive sponsor: CFO", "Fast close", "Q1 push",
              "Needs legal review", "Volume discount applied", "Annual subscription", ""]
amounts = [15000, 22000, 35000, 48000, 67000, 85000, 125000, 175000, 225000, 310000, 385000, 450000]

blank_rows = {8, 19, 31}
r = 2
for i in range(38):
    if r - 2 in blank_rows:
        r += 1
    ws1.cell(row=r, column=1, value=random.choice(regions))
    ws1.cell(row=r, column=2, value=random.choice(reps))
    ws1.cell(row=r, column=3, value=random.choice(products))
    ws1.cell(row=r, column=4, value=random.choice(customers))
    ws1.cell(row=r, column=5, value=dates_all[i % len(dates_all)])
    ws1.cell(row=r, column=6, value=random.choice(amounts))
    ws1.cell(row=r, column=7, value=random.choice(statuses))
    ws1.cell(row=r, column=8, value=random.choice([0.08, 0.10, 0.12, 0.14, 0.15]))
    ws1.cell(row=r, column=9, value=random.choice(notes_pool))
    # Duplicate row 12
    if i == 12:
        r += 1
        for c in range(1, 10):
            ws1.cell(row=r, column=c, value=ws1.cell(row=r-1, column=c).value)
    r += 1

# Text-stored numbers
for tr in [5, 16, 28]:
    cell = ws1.cell(row=tr, column=6)
    if cell.value and isinstance(cell.value, (int, float)):
        cell.value = str(int(cell.value))
        cell.number_format = "@"

# Negative adjustment
ws1.cell(row=r, column=1, value="Northeast")
ws1.cell(row=r, column=2, value="System")
ws1.cell(row=r, column=3, value="iGO")
ws1.cell(row=r, column=4, value="MetLife")
ws1.cell(row=r, column=5, value="03/31/2026")
ws1.cell(row=r, column=6, value=-12500)
ws1.cell(row=r, column=7, value="Credit")
ws1.cell(row=r, column=8, value=0)
ws1.cell(row=r, column=9, value="Q1 billing adjustment - approved by Controller")

# Error formula
ws1.cell(row=15, column=9, value='=VLOOKUP("missing_ref",A1:A5,1,FALSE)')

# Comments
ws1.cell(row=3, column=6).comment = Comment("Verify with accounting - amount seems high for single quarter", "Finance Team")
ws1.cell(row=7, column=7).comment = Comment("Follow up by end of month - needs VP approval", "Sarah Mitchell")
ws1.cell(row=11, column=3).comment = Comment("New enterprise tier pricing - check with Product team", "James Rodriguez")
ws1.cell(row=22, column=8).comment = Comment("Override commission approved by VP Sales - see email 2/15", "HR Comp Team")
ws1.cell(row=18, column=4).comment = Comment("Note: Acquired by Lincoln Financial in Q4 2025", "M&A Team")

for c, w in enumerate([14, 18, 14, 22, 16, 14, 13, 14, 30], 1):
    ws1.column_dimensions[get_column_letter(c)].width = w
ws1.sheet_view.showGridLines = False
print(f"Q1 Revenue: {r} rows")

# ============================================================
# SHEET 2: Q1 Expenses
# ============================================================
ws2 = wb.create_sheet("Q1 Expenses")
exp_hdrs = ["Department", "Category", "Vendor", "Invoice Date", "Amount", "Approved By", "PO Number"]
for c, h in enumerate(exp_hdrs, 1):
    ws2.cell(row=c if c == 1 else 1, column=c, value=h)
# Fix: write headers to row 1
for c, h in enumerate(exp_hdrs, 1):
    ws2.cell(row=1, column=c, value=h)

expenses_data = [
    ["Engineering", "Cloud Infrastructure", "AWS", "01/10/2026", 78000, "T. Davis (CTO)", "PO-2026-001"],
    ["Engineering", "Software Licenses", "GitHub Enterprise", "01/15/2026", 22000, "T. Davis (CTO)", "PO-2026-002"],
    ["Engineering", "Hardware", "Dell Technologies", "02/08/2026", 45000, "T. Davis (CTO)", "PO-2026-003"],
    ["Engineering", "Consulting", "Deloitte", "03/01/2026", 125000, "Pending", "PO-2026-004"],
    ["Marketing", "Digital Advertising", "Google Ads", "01/20/2026", 35000, "R. Patel (Controller)", "PO-2026-005"],
    ["Marketing", "Events & Conferences", "Reed Exhibitions", "02/05/2026", 18000, "R. Patel (Controller)", "PO-2026-006"],
    ["Marketing", "Marketing Software", "HubSpot", "02/15/2026", 15000, "R. Patel (Controller)", "PO-2026-007"],
    ["Sales", "Travel", "Delta Airlines", "01/22/2026", 8900, "M. Lee (VP Ops)", "PO-2026-008"],
    ["Sales", "Sales Tools", "Salesforce", "02/01/2026", 45000, "J. Smith (CFO)", "PO-2026-009"],
    ["Sales", "Entertainment", "Client Events", "02/18/2026", 5600, "M. Lee (VP Ops)", "PO-2026-010"],
    ["Sales", "Training", "Sandler Training", "03/10/2026", 12500, "M. Lee (VP Ops)", "PO-2026-011"],
    ["Finance", "Audit & Tax", "KPMG", "01/30/2026", 85000, "J. Smith (CFO)", "PO-2026-012"],
    ["Finance", "Financial Software", "Workday", "02/20/2026", 18000, "J. Smith (CFO)", "PO-2026-013"],
    ["HR", "Recruiting", "LinkedIn Recruiter", "01/25/2026", 15000, "M. Lee (VP Ops)", "PO-2026-014"],
    ["HR", "Training & Development", "Udemy Business", "02/12/2026", 5600, "M. Lee (VP Ops)", "PO-2026-015"],
    ["HR", "Benefits Administration", "ADP", "03/05/2026", 8900, "M. Lee (VP Ops)", "PO-2026-016"],
    ["Operations", "Facilities", "WeWork", "01/05/2026", 95000, "M. Lee (VP Ops)", "PO-2026-017"],
    ["Operations", "Security", "CrowdStrike", "02/10/2026", 22000, "T. Davis (CTO)", "PO-2026-018"],
    ["Operations", "Telecommunications", "Verizon", "03/01/2026", 8500, "M. Lee (VP Ops)", "PO-2026-019"],
    ["Legal", "Outside Counsel", "Baker McKenzie", "01/18/2026", 45000, "J. Smith (CFO)", "PO-2026-020"],
    ["Legal", "Compliance", "OneTrust", "02/25/2026", 12500, "R. Patel (Controller)", "PO-2026-021"],
    ["Engineering", "Cloud Infrastructure", "Azure", "03/15/2026", 35000, "T. Davis (CTO)", "PO-2026-022"],
    ["Marketing", "Content Production", "Contently", "03/20/2026", 8900, "R. Patel (Controller)", "PO-2026-023"],
    ["Engineering", "Software Licenses", "Datadog", "03/25/2026", 18000, "T. Davis (CTO)", "PO-2026-024"],
    ["Engineering", "Software Licenses", "Snowflake", "03/28/2026", 25000, "T. Davis (CTO)", "PO-2026-025"],
]

for i, row in enumerate(expenses_data, 2):
    for c, v in enumerate(row, 1):
        ws2.cell(row=i, column=c, value=v)

for c, w in enumerate([16, 22, 20, 14, 14, 20, 14], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w
ws2.sheet_view.showGridLines = False
print(f"Q1 Expenses: {len(expenses_data)} rows")

# ============================================================
# SHEET 3: Q1 Revenue v2 (copy with 8 differences)
# ============================================================
ws3 = wb.create_sheet("Q1 Revenue v2")
for r in range(1, ws1.max_row + 1):
    for c in range(1, 10):
        ws3.cell(row=r, column=c, value=ws1.cell(row=r, column=c).value)

# 8 differences
if ws3.cell(row=3, column=6).value and isinstance(ws3.cell(row=3, column=6).value, (int, float)):
    ws3.cell(row=3, column=6, value=ws3.cell(row=3, column=6).value + 25000)
if ws3.cell(row=7, column=6).value and isinstance(ws3.cell(row=7, column=6).value, (int, float)):
    ws3.cell(row=7, column=6, value=ws3.cell(row=7, column=6).value - 15000)
if ws3.cell(row=12, column=6).value and isinstance(ws3.cell(row=12, column=6).value, (int, float)):
    ws3.cell(row=12, column=6, value=ws3.cell(row=12, column=6).value + 40000)
ws3.cell(row=10, column=7, value="Pipeline")
ws3.cell(row=20, column=7, value="Closed Won")
ws3.cell(row=6, column=9, value="UPDATED: Contract extended to 3 years")
ws3.cell(row=25, column=9, value="REVISED: Pricing renegotiated per Q1 review")
ws3.cell(row=30, column=5, value="03/15/2026")

nr = ws3.max_row + 1
ws3.cell(row=nr, column=1, value="International")
ws3.cell(row=nr, column=2, value="New Hire TBD")
ws3.cell(row=nr, column=3, value="iGO")
ws3.cell(row=nr, column=4, value="Zurich Insurance")
ws3.cell(row=nr, column=5, value="03/30/2026")
ws3.cell(row=nr, column=6, value=290000)
ws3.cell(row=nr, column=7, value="Pipeline")
ws3.cell(row=nr, column=8, value=0.12)
ws3.cell(row=nr, column=9, value="International expansion - Q2 target")

for c, w in enumerate([14, 18, 14, 22, 16, 14, 13, 14, 30], 1):
    ws3.column_dimensions[get_column_letter(c)].width = w
ws3.sheet_view.showGridLines = False
print("Q1 Revenue v2: copy with 8 differences + 1 new row")

# ============================================================
# SHEET 4: Budget Summary (styled)
# ============================================================
ws4 = wb.create_sheet("Budget Summary")
bhdrs = ["Department", "Q1 Budget", "Q1 Actual", "Variance ($)", "Variance (%)", "Status"]
hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=BLUE)

for c, h in enumerate(bhdrs, 1):
    cell = ws4.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

budgets = [
    ["Engineering", 485000, 502000, "On Track"],
    ["Marketing", 165000, 178000, "Over Budget"],
    ["Sales", 95000, 87000, "Under Budget"],
    ["Finance", 72000, 68000, "Under Budget"],
    ["HR", 58000, 61000, "Over Budget"],
    ["Operations", 125000, 118000, "Under Budget"],
    ["Legal", 85000, 92000, "Over Budget"],
]

for i, (dept, bud, act, status) in enumerate(budgets, 2):
    ws4.cell(row=i, column=1, value=dept)
    ws4.cell(row=i, column=1).font = Font(name="Arial", bold=True, size=10)
    ws4.cell(row=i, column=2, value=bud)
    ws4.cell(row=i, column=2).number_format = "$#,##0"
    ws4.cell(row=i, column=3, value=act)
    ws4.cell(row=i, column=3).number_format = "$#,##0"
    ws4.cell(row=i, column=4, value=f"=C{i}-B{i}")
    ws4.cell(row=i, column=4).number_format = "$#,##0;($#,##0)"
    ws4.cell(row=i, column=5, value=f"=IF(B{i}=0,0,(C{i}-B{i})/B{i})")
    ws4.cell(row=i, column=5).number_format = "0.0%"
    ws4.cell(row=i, column=6, value=status)
    if status == "Over Budget":
        ws4.cell(row=i, column=6).font = Font(name="Arial", color="CC0000", bold=True)
    elif status == "Under Budget":
        ws4.cell(row=i, column=6).font = Font(name="Arial", color="008000", bold=True)
    else:
        ws4.cell(row=i, column=6).font = Font(name="Arial", color="FF8C00", bold=True)
    if i % 2 == 0:
        for c in range(1, 7):
            ws4.cell(row=i, column=c).fill = PatternFill("solid", fgColor="F2F7FB")

tr = len(budgets) + 2
ws4.cell(row=tr, column=1, value="TOTAL")
ws4.cell(row=tr, column=1).font = Font(name="Arial", bold=True, size=11)
for c in range(2, 5):
    ws4.cell(row=tr, column=c, value=f"=SUM({get_column_letter(c)}2:{get_column_letter(c)}{tr-1})")
    ws4.cell(row=tr, column=c).font = Font(name="Arial", bold=True)
    ws4.cell(row=tr, column=c).number_format = "$#,##0;($#,##0)"
ws4.cell(row=tr, column=5, value=f"=IF(B{tr}=0,0,(C{tr}-B{tr})/B{tr})")
ws4.cell(row=tr, column=5).number_format = "0.0%"
ws4.cell(row=tr, column=5).font = Font(name="Arial", bold=True)
for c in range(1, 7):
    ws4.cell(row=tr, column=c).border = Border(top=Side(style="double", color="000000"))

# Dropdown source
ws4.cell(row=tr+3, column=1, value="Status Options (for dropdown):")
ws4.cell(row=tr+3, column=1).font = Font(name="Arial", italic=True, size=9, color="888888")
for i, s in enumerate(["On Track", "Over Budget", "Under Budget", "Under Review", "Frozen"]):
    ws4.cell(row=tr+4+i, column=1, value=s)

for c, w in enumerate([18, 14, 14, 14, 14, 16], 1):
    ws4.column_dimensions[get_column_letter(c)].width = w
ws4.sheet_view.showGridLines = False
print(f"Budget Summary: {len(budgets)} depts + totals")

# ============================================================
# SHEET 5: Contact List
# ============================================================
ws5 = wb.create_sheet("Contact List")
chdrs = ["Full Name", "Title", "Department", "Email", "Phone", "Office Location"]
for c, h in enumerate(chdrs, 1):
    ws5.cell(row=1, column=c, value=h)

contacts = [
    ["Robert Chen", "Chief Financial Officer", "Finance", "robert.chen@ipipeline.com", "555-0101", "Exton, PA"],
    ["Maria Santos", "VP of Engineering", "Engineering", "maria.santos@ipipeline.com", "(555) 010-2", "Exton, PA"],
    ["David Kim", "Director of Sales", "Sales", "david.kim@ipipeline.com", "555.0103", "New York, NY"],
    ["Jennifer Williams", "Senior Financial Analyst", "Finance", "jennifer.williams@ipipeline.com", "555-0104", "Exton, PA"],
    ["Marcus Johnson", "Head of Product", "Engineering", "marcus.johnson@ipipeline.com", "(555) 010-5", "Exton, PA"],
    ["Priya Sharma", "VP of Marketing", "Marketing", "priya.sharma@ipipeline.com", "555.0106", "Boston, MA"],
    ["Thomas Anderson", "General Counsel", "Legal", "thomas.anderson@ipipeline.com", "555-0107", "Exton, PA"],
    ["Amanda Foster", "Controller", "Finance", "amanda.foster@ipipeline.com", "(555) 010-8", "Exton, PA"],
    ["Ryan O'Brien", "Senior DevOps Engineer", "Engineering", "ryan.obrien@ipipeline.com", "555.0109", "Remote"],
    ["Sarah Mitchell", "Regional Sales Manager", "Sales", "sarah.mitchell@ipipeline.com", "555-0110", "Chicago, IL"],
    ["Kevin Park", "HR Business Partner", "HR", "kevin.park@ipipeline.com", "(555) 011-1", "Exton, PA"],
    ["Lisa Tran", "Data Analytics Lead", "Finance", "lisa.tran@ipipeline.com", "555.0112", "Exton, PA"],
    ["Christopher Lee", "Chief Technology Officer", "Engineering", "christopher.lee@ipipeline.com", "555-0113", "Exton, PA"],
    ["Nicole Brown", "VP of Operations", "Operations", "nicole.brown@ipipeline.com", "(555) 011-4", "Exton, PA"],
]

for i, row in enumerate(contacts, 2):
    for c, v in enumerate(row, 1):
        ws5.cell(row=i, column=c, value=v)

for c, w in enumerate([22, 28, 16, 32, 14, 16], 1):
    ws5.column_dimensions[get_column_letter(c)].width = w
ws5.sheet_view.showGridLines = False
print(f"Contact List: {len(contacts)} contacts")

# ============================================================
# SHEET 6: Archive_Q4_2025 (HIDDEN)
# ============================================================
ws6 = wb.create_sheet("Archive_Q4_2025")
ws6.cell(row=1, column=1, value="Q4 2025 ARCHIVE - DO NOT MODIFY")
ws6.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=12, color="CC0000")
ws6.cell(row=2, column=1, value="This sheet contains archived Q4 2025 data.")
ws6.cell(row=3, column=1, value="Contact Finance team before making any changes.")
ws6.cell(row=3, column=1).font = Font(name="Arial", italic=True, size=10, color="888888")
ws6.cell(row=5, column=1, value="Department")
ws6.cell(row=5, column=2, value="Q4 Revenue")
ws6.cell(row=5, column=3, value="Q4 Expenses")
ws6.cell(row=6, column=1, value="Engineering")
ws6.cell(row=6, column=2, value=2150000)
ws6.cell(row=6, column=3, value=485000)
ws6.cell(row=7, column=1, value="Sales")
ws6.cell(row=7, column=2, value=3800000)
ws6.cell(row=7, column=3, value=95000)
ws6.cell(row=8, column=1, value="Marketing")
ws6.cell(row=8, column=2, value=0)
ws6.cell(row=8, column=3, value=165000)
ws6.sheet_state = "hidden"
ws6.sheet_view.showGridLines = False
print("Archive_Q4_2025: hidden")

# ============================================================
# SAVE
# ============================================================
path = r"C:\Users\connor.atlee\RecTrial\SampleFile\Sample_Quarterly_Report.xlsx"
wb.save(path)
print(f"\nSaved: {path}")
print(f"Sheets: {wb.sheetnames}")
print("\nIntentional mess:")
print("  - 3 blank rows in Q1 Revenue")
print("  - 3 text-stored numbers")
print("  - 1 duplicate row")
print("  - 2 names with leading/trailing spaces")
print("  - Mixed date formats (6 different formats)")
print("  - 1 negative amount (-$12,500)")
print("  - 1 #N/A error formula")
print("  - 5 comments")
print("  - 1 hidden sheet")
print("  - Unstyled headers on sheets 1, 2, 3, 5")
print("  - Mixed phone formats on Contact List")
print("  - Enterprise-scale deal amounts ($15K-$450K)")
print("  - Real insurance company customer names")
print("  - Real vendor names")

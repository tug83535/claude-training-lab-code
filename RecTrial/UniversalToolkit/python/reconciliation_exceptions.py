"""
KBT Universal Tools — Reconciliation Exception Generator
============================================================
Compares two Excel datasets, filters out all matched items,
and outputs ONLY the unmatched exceptions for human review.

Matching is done on exact values in one or more key columns.

Usage:
    python reconciliation_exceptions.py "C:\\list1.xlsx" "C:\\list2.xlsx" --key "Invoice No"
    python reconciliation_exceptions.py "list1.xlsx" "list2.xlsx" --key "Invoice No" "Amount"

Output: Saves "RECONCILIATION_EXCEPTIONS.xlsx" — exceptions only
"""

import sys
import os
import argparse

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)


def generate_exceptions(file1: str, file2: str, key_cols: list,
                         sheet1: str = None, sheet2: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Reconciliation Exception Generator")
    print(f"{'='*55}")
    print(f"  File 1:  {os.path.basename(file1)}")
    print(f"  File 2:  {os.path.basename(file2)}")
    print(f"  Match on: {key_cols}")
    print(f"{'='*55}\n")

    for f in [file1, file2]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)

    df1 = pd.read_excel(file1, sheet_name=sheet1 or 0)
    df2 = pd.read_excel(file2, sheet_name=sheet2 or 0)
    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    for col in key_cols:
        for df, name in [(df1, 'File 1'), (df2, 'File 2')]:
            if col not in df.columns:
                print(f"ERROR: Key column '{col}' not found in {name}.")
                print(f"Available: {', '.join(df.columns.tolist())}")
                sys.exit(1)

    print(f"File 1 records: {len(df1):,}")
    print(f"File 2 records: {len(df2):,}\n")

    # Create composite key for matching
    def make_key(df, cols):
        return df[cols].astype(str).apply(lambda r: '|'.join(r.values), axis=1)

    df1['_key'] = make_key(df1, key_cols)
    df2['_key'] = make_key(df2, key_cols)

    # Exceptions in File 1 not in File 2
    exceptions_1 = df1[~df1['_key'].isin(df2['_key'])].drop(columns=['_key'])
    # Exceptions in File 2 not in File 1
    exceptions_2 = df2[~df2['_key'].isin(df1['_key'])].drop(columns=['_key'])
    # Matched records
    matched_keys = df1[df1['_key'].isin(df2['_key'])]['_key']
    matched_count = len(matched_keys)

    df1_clean = df1.drop(columns=['_key'])
    df2_clean = df2.drop(columns=['_key'])

    total_exceptions = len(exceptions_1) + len(exceptions_2)

    print(f"Results:")
    print(f"  Matched (in both files):   {matched_count:,}")
    print(f"  Exceptions in File 1 only: {len(exceptions_1):,}")
    print(f"  Exceptions in File 2 only: {len(exceptions_2):,}")
    print(f"  Total exceptions:          {total_exceptions:,}")

    fill_exc1   = PatternFill("solid", fgColor="FFC7CE")
    fill_exc2   = PatternFill("solid", fgColor="FFEB9C")
    fill_header = PatternFill("solid", fgColor="1F497D")
    font_header = Font(bold=True, color="FFFFFF")

    out_dir = os.path.dirname(file1)
    output_path = os.path.join(out_dir, "RECONCILIATION_EXCEPTIONS.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Summary
        summary_data = {
            'Metric': ['File 1 Records', 'File 2 Records', 'Matched (both files)',
                       'Exceptions in File 1 Only', 'Exceptions in File 2 Only',
                       'Total Exceptions', 'Key Columns Used'],
            'Value': [len(df1_clean), len(df2_clean), matched_count,
                      len(exceptions_1), len(exceptions_2), total_exceptions,
                      ', '.join(key_cols)]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

        if not exceptions_1.empty:
            f1_label = os.path.basename(file1)[:28] + " ONLY"
            exceptions_1.to_excel(writer, sheet_name=f1_label, index=False)
            ws = writer.sheets[f1_label]
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = fill_exc1
            _autofit(ws)

        if not exceptions_2.empty:
            f2_label = os.path.basename(file2)[:28] + " ONLY"
            exceptions_2.to_excel(writer, sheet_name=f2_label, index=False)
            ws = writer.sheets[f2_label]
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = fill_exc2
            _autofit(ws)

        if total_exceptions == 0:
            note_ws = writer.book.create_sheet("Result")
            note_ws['A1'] = "ALL RECORDS MATCH — No exceptions found."
            note_ws['A1'].font = Font(bold=True, size=14)

        _autofit(writer.sheets['Summary'])

    print(f"\n{'='*55}")
    print(f"  DONE! Exceptions report saved to:")
    print(f"  {output_path}")
    if total_exceptions == 0:
        print(f"  ALL RECORDS MATCHED — no exceptions.")
    else:
        print(f"  Red = in File 1 only | Yellow = in File 2 only")
    print(f"{'='*55}\n")


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)


def main():
    parser = argparse.ArgumentParser(description='KBT Reconciliation Exception Generator')
    parser.add_argument('file1', help='Path to the first file')
    parser.add_argument('file2', help='Path to the second file')
    parser.add_argument('--key', nargs='+', required=True,
                        help='Column name(s) to match on (e.g. --key "Invoice No" "Amount")')
    parser.add_argument('--sheet1', default=None, help='Sheet name in file 1')
    parser.add_argument('--sheet2', default=None, help='Sheet name in file 2')
    args = parser.parse_args()
    generate_exceptions(args.file1, args.file2, args.key, args.sheet1, args.sheet2)


if __name__ == '__main__':
    main()

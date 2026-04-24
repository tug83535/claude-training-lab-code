"""
KBT Universal Tools — Variance Decomposition Analyzer
============================================================
Automatically breaks down financial variances into their
price, volume, and mix components — standard FP&A analysis.

Formula:
  Total Variance = Price Effect + Volume Effect + Mix Effect

  Price Effect  = (Actual Price   - Budget Price)   × Actual Volume
  Volume Effect = (Actual Volume  - Budget Volume)  × Budget Price
  Mix Effect    = Remaining (Total - Price - Volume)

Usage:
    python variance_decomposition.py "C:\\data.xlsx"
    python variance_decomposition.py "data.xlsx" --product "Product" --act-vol "Act Units"
        --bud-vol "Bud Units" --act-price "Act Price" --bud-price "Bud Price"

Output: Saves "VARIANCE_DECOMPOSITION.xlsx" with bridge analysis
"""

import sys
import os
import argparse

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)


def decompose_variance(file_path: str, product_col: str, act_vol_col: str, bud_vol_col: str,
                        act_price_col: str, bud_price_col: str, sheet_name: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Variance Decomposition Analyzer")
    print(f"{'='*55}")
    print(f"  File: {os.path.basename(file_path)}")
    print(f"{'='*55}\n")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
    df.columns = df.columns.str.strip()

    required = [product_col, act_vol_col, bud_vol_col, act_price_col, bud_price_col]
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"ERROR: Columns not found: {missing}")
        print(f"Available: {', '.join(df.columns.tolist())}")
        sys.exit(1)

    for col in [act_vol_col, bud_vol_col, act_price_col, bud_price_col]:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    df = df.dropna(subset=[act_vol_col, bud_vol_col, act_price_col, bud_price_col])

    # Core decomposition
    df['Budget_Revenue']  = df[bud_vol_col]   * df[bud_price_col]
    df['Actual_Revenue']  = df[act_vol_col]   * df[act_price_col]
    df['Total_Variance']  = df['Actual_Revenue'] - df['Budget_Revenue']

    df['Price_Effect']    = (df[act_price_col] - df[bud_price_col]) * df[act_vol_col]
    df['Volume_Effect']   = (df[act_vol_col]  - df[bud_vol_col])   * df[bud_price_col]
    df['Mix_Effect']      = df['Total_Variance'] - df['Price_Effect'] - df['Volume_Effect']

    df['Price_Favorable']  = df['Price_Effect'].apply(lambda v: 'Favorable' if v > 0 else 'Unfavorable')
    df['Volume_Favorable'] = df['Volume_Effect'].apply(lambda v: 'Favorable' if v > 0 else 'Unfavorable')

    # Totals row
    totals = pd.DataFrame([{
        product_col:       'TOTAL',
        bud_vol_col:       df[bud_vol_col].sum(),
        act_vol_col:       df[act_vol_col].sum(),
        bud_price_col:     '',
        act_price_col:     '',
        'Budget_Revenue':  df['Budget_Revenue'].sum(),
        'Actual_Revenue':  df['Actual_Revenue'].sum(),
        'Total_Variance':  df['Total_Variance'].sum(),
        'Price_Effect':    df['Price_Effect'].sum(),
        'Volume_Effect':   df['Volume_Effect'].sum(),
        'Mix_Effect':      df['Mix_Effect'].sum(),
        'Price_Favorable': '',
        'Volume_Favorable': ''
    }])
    df_with_totals = pd.concat([df, totals], ignore_index=True)

    print(f"Products analyzed: {len(df)}")
    print(f"\nSummary:")
    print(f"  Total Variance:  ${df['Total_Variance'].sum():>12,.2f}")
    print(f"  Price Effect:    ${df['Price_Effect'].sum():>12,.2f}")
    print(f"  Volume Effect:   ${df['Volume_Effect'].sum():>12,.2f}")
    print(f"  Mix Effect:      ${df['Mix_Effect'].sum():>12,.2f}")

    output_path = os.path.join(os.path.dirname(file_path), "VARIANCE_DECOMPOSITION.xlsx")

    fill_fav   = PatternFill("solid", fgColor="C6EFCE")
    fill_unfav = PatternFill("solid", fgColor="FFC7CE")
    fill_total = PatternFill("solid", fgColor="1F497D")
    fill_header= PatternFill("solid", fgColor="1F497D")
    font_header= Font(bold=True, color="FFFFFF")
    font_total = Font(bold=True, color="FFFFFF")

    currency_cols = ['Budget_Revenue', 'Actual_Revenue', 'Total_Variance',
                     'Price_Effect', 'Volume_Effect', 'Mix_Effect']

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_with_totals.to_excel(writer, sheet_name='Decomposition', index=False)
        ws = writer.sheets['Decomposition']

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header

        total_row = ws.max_row
        for cell in ws[total_row]:
            cell.fill = fill_total
            cell.font = font_total

        # Format currency columns
        col_map = {cell.value: cell.column_letter for cell in ws[1]}
        for cname in currency_cols:
            if cname in col_map:
                for row in ws.iter_rows(min_row=2, min_col=ws[f'{col_map[cname]}1'].column,
                                        max_col=ws[f'{col_map[cname]}1'].column):
                    for cell in row:
                        cell.number_format = '#,##0.00;[Red](#,##0.00)'

        # Color favorable/unfavorable rows by total variance
        total_var_col = col_map.get('Total_Variance')
        if total_var_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
                var_cell = ws[f'{total_var_col}{row[0].row}']
                if var_cell.value and isinstance(var_cell.value, (int, float)):
                    fill = fill_fav if var_cell.value >= 0 else fill_unfav
                    for cell in row:
                        if not cell.fill.fgColor.rgb == 'FF1F497D':
                            cell.fill = fill

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 22)

        # Bridge chart data
        bridge_data = {
            'Component': ['Budget', 'Price Effect', 'Volume Effect', 'Mix Effect', 'Actual'],
            'Amount': [
                df['Budget_Revenue'].sum(),
                df['Price_Effect'].sum(),
                df['Volume_Effect'].sum(),
                df['Mix_Effect'].sum(),
                df['Actual_Revenue'].sum()
            ]
        }
        bridge_df = pd.DataFrame(bridge_data)
        bridge_df.to_excel(writer, sheet_name='Bridge Chart Data', index=False)

    print(f"\n{'='*55}")
    print(f"  DONE! Saved to: {output_path}")
    print(f"  Green=Favorable | Red=Unfavorable")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Variance Decomposition Analyzer')
    parser.add_argument('file', help='Path to the Excel file with product/volume/price data')
    parser.add_argument('--product',   default='Product',       help='Product/item column (default: Product)')
    parser.add_argument('--act-vol',   default='Actual_Volume', help='Actual volume column')
    parser.add_argument('--bud-vol',   default='Budget_Volume', help='Budget volume column')
    parser.add_argument('--act-price', default='Actual_Price',  help='Actual price column')
    parser.add_argument('--bud-price', default='Budget_Price',  help='Budget price column')
    parser.add_argument('--sheet',     default=None,            help='Sheet name to read')
    args = parser.parse_args()
    decompose_variance(args.file, args.product, args.act_vol, args.bud_vol,
                       args.act_price, args.bud_price, args.sheet)


if __name__ == '__main__':
    main()

"""
KBT Universal Tools — Unstructured Data Unpivoter
============================================================
Converts wide-format pivot-style data (one column per month/
product/category) to tall database format (one row per record).

Example: Converts this:
  Name | Jan | Feb | Mar
  ACME | 100 | 200 | 300

To this:
  Name | Month | Amount
  ACME | Jan   | 100
  ACME | Feb   | 200
  ACME | Mar   | 300

Usage:
    python unpivot_data.py "C:\\path\\data.xlsx"
    python unpivot_data.py "data.xlsx" --id-cols "Name" "Department" --value-name "Amount" --var-name "Period"

Output: Saves "UNPIVOTED.xlsx" in the same folder
"""

import sys
import os
import argparse

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)


def unpivot_data(file_path: str, id_cols: list, value_name: str,
                 var_name: str, sheet_name: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Data Unpivoter")
    print(f"{'='*55}")
    print(f"  File:       {os.path.basename(file_path)}")
    print(f"  ID columns: {id_cols if id_cols else 'Auto-detect'}")
    print(f"  Value name: '{value_name}' | Variable name: '{var_name}'")
    print(f"{'='*55}\n")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
    df.columns = df.columns.str.strip()

    print(f"Original shape: {df.shape[0]} rows x {df.shape[1]} columns")

    # Validate ID columns
    if id_cols:
        missing = [c for c in id_cols if c not in df.columns]
        if missing:
            print(f"ERROR: Columns not found: {missing}")
            print(f"Available columns: {', '.join(df.columns.tolist())}")
            sys.exit(1)
    else:
        # Auto-detect: treat non-numeric columns as ID columns
        id_cols = df.select_dtypes(exclude='number').columns.tolist()
        if not id_cols:
            id_cols = [df.columns[0]]
        print(f"Auto-detected ID columns: {id_cols}")

    value_cols = [c for c in df.columns if c not in id_cols]
    print(f"Value columns to unpivot: {len(value_cols)} ({', '.join(value_cols[:5])}{'...' if len(value_cols) > 5 else ''})\n")

    melted = df.melt(id_vars=id_cols, value_vars=value_cols,
                     var_name=var_name, value_name=value_name)

    # Remove rows where the value is null
    original_count = len(melted)
    melted = melted.dropna(subset=[value_name])
    dropped = original_count - len(melted)

    print(f"Unpivoted shape: {melted.shape[0]} rows x {melted.shape[1]} columns")
    if dropped > 0:
        print(f"Null value rows removed: {dropped}")

    output_path = os.path.join(os.path.dirname(file_path), "UNPIVOTED.xlsx")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        melted.to_excel(writer, sheet_name='Unpivoted', index=False)
        ws = writer.sheets['Unpivoted']
        fill = PatternFill("solid", fgColor="1F497D")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    print(f"\n{'='*55}")
    print(f"  DONE! Saved to: {output_path}")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Data Unpivoter')
    parser.add_argument('file', help='Path to the wide-format Excel file')
    parser.add_argument('--id-cols', nargs='+', default=None,
                        help='Columns to keep as identifiers (default: auto-detect non-numeric)')
    parser.add_argument('--value-name', default='Value',
                        help='Name for the new values column (default: Value)')
    parser.add_argument('--var-name', default='Variable',
                        help='Name for the new variable/category column (default: Variable)')
    parser.add_argument('--sheet', default=None, help='Sheet name to read')
    args = parser.parse_args()
    unpivot_data(args.file, args.id_cols, args.value_name, args.var_name, args.sheet)


if __name__ == '__main__':
    main()

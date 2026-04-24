"""
KBT Universal Tools — Dynamic Master Data Mapper
============================================================
Performs SQL-style joins between datasets on common keys.
Replaces complex nested VLOOKUP arrays with a simple command.

Join types: left (default), inner, outer, right

Usage:
    python master_data_mapper.py "C:\\data.xlsx" "C:\\master.xlsx" --key "Vendor ID"
    python master_data_mapper.py "data.xlsx" "master.xlsx" --key "Account" --join inner
    python master_data_mapper.py "data.xlsx" "master.xlsx" --key "ID" --cols "Name" "Category" "Rate"

Output: Saves "MASTER_DATA_MAPPED.xlsx" with joined results
"""

import sys
import os
import argparse

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("ERROR: Run: pip install pandas openpyxl")
    sys.exit(1)


def master_data_map(source_file: str, master_file: str, key_col: str,
                    join_type: str, return_cols: list,
                    source_sheet: str = None, master_sheet: str = None) -> None:
    print(f"\n{'='*55}")
    print("  KBT Master Data Mapper")
    print(f"{'='*55}")
    print(f"  Source:  {os.path.basename(source_file)}")
    print(f"  Master:  {os.path.basename(master_file)}")
    print(f"  Key:     '{key_col}'")
    print(f"  Join:    {join_type}")
    print(f"  Return:  {return_cols if return_cols else 'All master columns'}")
    print(f"{'='*55}\n")

    for f in [source_file, master_file]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)

    source_df = pd.read_excel(source_file, sheet_name=source_sheet or 0)
    master_df = pd.read_excel(master_file, sheet_name=master_sheet or 0)
    source_df.columns = source_df.columns.str.strip()
    master_df.columns = master_df.columns.str.strip()

    for col, df, name in [(key_col, source_df, 'Source'), (key_col, master_df, 'Master')]:
        if col not in df.columns:
            print(f"ERROR: Key column '{col}' not found in {name} file.")
            print(f"Available in {name}: {', '.join(df.columns.tolist())}")
            sys.exit(1)

    if return_cols:
        missing = [c for c in return_cols if c not in master_df.columns]
        if missing:
            print(f"ERROR: Columns not found in master: {missing}")
            sys.exit(1)
        master_subset = master_df[[key_col] + return_cols].drop_duplicates(subset=[key_col])
    else:
        master_subset = master_df.drop_duplicates(subset=[key_col])

    print(f"Source records:  {len(source_df):,}")
    print(f"Master records:  {len(master_subset):,}")
    print(f"\nPerforming {join_type} join on '{key_col}'...")

    merged = source_df.merge(master_subset, on=key_col, how=join_type,
                             suffixes=('', '_master'))

    unmatched = merged[merged[master_subset.columns[1]].isna()] if return_cols else pd.DataFrame()

    print(f"Result rows:     {len(merged):,}")
    if join_type in ('left', 'outer'):
        unmatched_count = merged[merged.isnull().any(axis=1)].shape[0]
        print(f"Unmatched rows:  {unmatched_count:,}")

    output_path = os.path.join(os.path.dirname(source_file), "MASTER_DATA_MAPPED.xlsx")

    fill_matched  = PatternFill("solid", fgColor="C6EFCE")
    fill_unmatched= PatternFill("solid", fgColor="FFC7CE")
    fill_header   = PatternFill("solid", fgColor="1F497D")
    font_header   = Font(bold=True, color="FFFFFF")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        merged.to_excel(writer, sheet_name='Mapped Results', index=False)
        ws = writer.sheets['Mapped Results']

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header

        # Highlight rows with null values (unmatched)
        if return_cols:
            check_col_idx = merged.columns.get_loc(return_cols[0]) + 1
            for row in ws.iter_rows(min_row=2):
                val = row[check_col_idx - 1].value
                if val is None or str(val).strip() == "" or str(val) == "nan":
                    for cell in row:
                        cell.fill = fill_unmatched

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 35)

        # Summary sheet
        summary_data = {
            'Metric': ['Source Records', 'Master Records', 'Result Records',
                       'Join Type', 'Key Column'],
            'Value': [len(source_df), len(master_subset), len(merged),
                      join_type, key_col]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

    print(f"\n{'='*55}")
    print(f"  DONE! Saved to: {output_path}")
    print(f"  Green=Matched | Red=No master record found")
    print(f"{'='*55}\n")


def main():
    parser = argparse.ArgumentParser(description='KBT Dynamic Master Data Mapper')
    parser.add_argument('source', help='Path to the source data file')
    parser.add_argument('master', help='Path to the master/lookup data file')
    parser.add_argument('--key', required=True, help='Column name to join on (must exist in both files)')
    parser.add_argument('--join', default='left',
                        choices=['left', 'inner', 'outer', 'right'],
                        help='Join type (default: left — keeps all source records)')
    parser.add_argument('--cols', nargs='+', default=None,
                        help='Specific columns to pull from master (default: all)')
    parser.add_argument('--source-sheet', default=None, help='Source sheet name')
    parser.add_argument('--master-sheet', default=None, help='Master sheet name')
    args = parser.parse_args()
    master_data_map(args.source, args.master, args.key, args.join, args.cols,
                    args.source_sheet, args.master_sheet)


if __name__ == '__main__':
    main()

"""
Fix Demo File Data — Make all macros produce impressive output.
Run this ONCE on your demo Excel file before recording.

What it does:
1. P&L Monthly Trend: Ensures Budget column exists with values that create
   meaningful variances against Actuals (so Variance Commentary works)
2. P&L Monthly Trend: Ensures row labels match what macros look for
   (Total Revenue, Gross Profit, Operating Expenses, Net Income)
3. General Ledger (CrossfireHiddenWorksheet): Adds realistic GL transactions
   so Reconciliation checks have data to validate
4. Assumptions: Ensures driver values exist for What-If demo
5. Checks: Adds headers if missing

Usage:
  python fix_demo_data.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from copy import copy
import os
import shutil
import sys

# Path to demo file
DEMO_PATH = r"C:\Users\connor.atlee\RecTrial\DemoFile\ExcelDemoFile_adv.xlsm"
BACKUP_PATH = DEMO_PATH.replace(".xlsm", "_BACKUP.xlsm")

if not os.path.exists(DEMO_PATH):
    print(f"ERROR: Demo file not found at {DEMO_PATH}")
    sys.exit(1)

# Backup first
shutil.copy2(DEMO_PATH, BACKUP_PATH)
print(f"Backup saved: {BACKUP_PATH}")

wb = openpyxl.load_workbook(DEMO_PATH, keep_vba=True)
print(f"Opened: {DEMO_PATH}")
print(f"Sheets: {wb.sheetnames}")

# ============================================================
# 1. FIX P&L - Monthly Trend
# ============================================================
PL_SHEET = "P&L - Monthly Trend"
if PL_SHEET in wb.sheetnames:
    ws = wb[PL_SHEET]
    print(f"\n--- Fixing {PL_SHEET} ---")
    print(f"Current range: {ws.min_row}-{ws.max_row} rows, {ws.min_column}-{ws.max_column} cols")

    # Read row 4 headers
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=4, column=c).value
        if val:
            headers[c] = str(val).strip()
            print(f"  Col {c}: '{val}'")

    # Read column A labels (rows 5+)
    labels = {}
    for r in range(5, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val:
            labels[r] = str(val).strip()

    print(f"\nRow labels found:")
    for r, lbl in labels.items():
        print(f"  Row {r}: '{lbl}'")

    # Find key rows
    key_rows = {}
    for r, lbl in labels.items():
        lbl_lower = lbl.lower()
        if "total revenue" in lbl_lower:
            key_rows["Total Revenue"] = r
        elif "gross profit" in lbl_lower:
            key_rows["Gross Profit"] = r
        elif "operating expense" in lbl_lower or "total opex" in lbl_lower or "total operating" in lbl_lower:
            key_rows["Operating Expenses"] = r
        elif "net income" in lbl_lower or "net profit" in lbl_lower:
            key_rows["Net Income"] = r
        elif "cost of revenue" in lbl_lower or "total cost" in lbl_lower or "cost of goods" in lbl_lower or "total cor" in lbl_lower:
            key_rows["Cost of Revenue"] = r

    print(f"\nKey rows identified: {key_rows}")

    # If key rows are missing, add them
    if "Total Revenue" not in key_rows:
        # Look for just "Revenue" as fallback
        for r, lbl in labels.items():
            if lbl.lower() == "revenue" or "total rev" in lbl.lower():
                key_rows["Total Revenue"] = r
                print(f"  Mapped '{lbl}' -> Total Revenue (row {r})")
                break

    # Find month columns (columns with month data, not label/total columns)
    month_cols = []
    budget_col = None
    fy_total_col = None
    for c, h in headers.items():
        h_lower = h.lower()
        if "budget" in h_lower or "plan" in h_lower:
            budget_col = c
        elif "total" in h_lower or "fy" in h_lower:
            fy_total_col = c
        elif any(m in h_lower for m in ["jan", "feb", "mar", "apr", "may", "jun",
                                         "jul", "aug", "sep", "oct", "nov", "dec"]):
            month_cols.append(c)

    print(f"\nMonth columns: {month_cols}")
    print(f"Budget column: {budget_col}")
    print(f"FY Total column: {fy_total_col}")

    # If no budget column exists, create one
    if budget_col is None and month_cols:
        # Add budget column after the last month column or after FY Total
        if fy_total_col:
            budget_col = fy_total_col + 1
        else:
            budget_col = max(month_cols) + 1

        # Shift existing data right if needed (only if there's data in that column)
        existing_val = ws.cell(row=4, column=budget_col).value
        if existing_val and "budget" not in str(existing_val).lower():
            budget_col = ws.max_column + 1

        ws.cell(row=4, column=budget_col, value="FY2025 Budget")
        # Copy header style from neighboring cell
        src_style = ws.cell(row=4, column=month_cols[0])
        tgt = ws.cell(row=4, column=budget_col)
        if src_style.font:
            tgt.font = copy(src_style.font)
        if src_style.fill:
            tgt.fill = copy(src_style.fill)
        if src_style.alignment:
            tgt.alignment = copy(src_style.alignment)

        print(f"Created Budget column at col {budget_col}")

    # Now populate budget values — take the latest month actual and add 5-15% variance
    if budget_col and month_cols:
        import random
        random.seed(42)  # Reproducible
        latest_month = max(month_cols)

        for r in range(5, ws.max_row + 1):
            actual_val = ws.cell(row=r, column=latest_month).value
            if actual_val is not None and isinstance(actual_val, (int, float)) and actual_val != 0:
                # Budget is actual +/- 5-20% to create meaningful variances
                variance_pct = random.uniform(-0.20, 0.15)
                budget_val = actual_val * (1 + variance_pct)
                ws.cell(row=r, column=budget_col, value=round(budget_val, 2))
                # Copy number format
                src_fmt = ws.cell(row=r, column=latest_month).number_format
                ws.cell(row=r, column=budget_col).number_format = src_fmt

        print(f"Populated budget values from col {latest_month} with 5-20% variance")

    # Ensure the key summary rows have values if they're currently 0 or empty
    # Check latest month column for zero values in key rows
    if month_cols:
        latest = max(month_cols)
        for label, r in key_rows.items():
            val = ws.cell(row=r, column=latest).value
            if val is None or val == 0:
                print(f"  WARNING: {label} (row {r}) has value {val} in latest month col {latest}")
else:
    print(f"WARNING: Sheet '{PL_SHEET}' not found!")

# ============================================================
# 2. FIX GENERAL LEDGER (CrossfireHiddenWorksheet)
# ============================================================
GL_SHEET = "CrossfireHiddenWorksheet"
if GL_SHEET in wb.sheetnames:
    ws_gl = wb[GL_SHEET]
    print(f"\n--- Checking {GL_SHEET} ---")
    print(f"Current range: {ws_gl.max_row} rows, {ws_gl.max_column} cols")

    # Check if GL has data
    if ws_gl.max_row < 10:
        print("GL has very little data — adding sample transactions")

        # Set headers if missing
        gl_headers = ["GL_ID", "Date", "Department", "Product", "Category", "Vendor", "Amount"]
        for c, h in enumerate(gl_headers, 1):
            if ws_gl.cell(row=1, column=c).value is None:
                ws_gl.cell(row=1, column=c, value=h)

        # Add sample GL transactions
        import random
        random.seed(42)
        departments = ["Engineering", "Marketing", "Sales", "Finance", "HR", "Operations"]
        products = ["iGO", "Affirm", "InsureSight", "DocFast"]
        categories = ["Revenue", "COGS", "Salary", "Software", "Travel", "Marketing", "AWS", "Consulting"]
        vendors = ["Internal", "AWS", "Google", "Microsoft", "Deloitte", "Delta", "Marriott", "HubSpot"]

        row = ws_gl.max_row + 1 if ws_gl.max_row > 1 else 2
        gl_id = 10001

        for month in range(1, 4):  # Jan, Feb, Mar
            for _ in range(50):  # 50 transactions per month
                dept = random.choice(departments)
                prod = random.choice(products)
                cat = random.choice(categories)
                vendor = random.choice(vendors)

                if cat == "Revenue":
                    amount = round(random.uniform(5000, 50000), 2)
                elif cat in ["Salary", "AWS", "Consulting"]:
                    amount = round(random.uniform(-30000, -5000), 2)
                else:
                    amount = round(random.uniform(-10000, -500), 2)

                day = random.randint(1, 28)
                date_str = f"{month:02d}/{day:02d}/2025"

                ws_gl.cell(row=row, column=1, value=gl_id)
                ws_gl.cell(row=row, column=2, value=date_str)
                ws_gl.cell(row=row, column=3, value=dept)
                ws_gl.cell(row=row, column=4, value=prod)
                ws_gl.cell(row=row, column=5, value=cat)
                ws_gl.cell(row=row, column=6, value=vendor)
                ws_gl.cell(row=row, column=7, value=amount)

                gl_id += 1
                row += 1

        print(f"Added {row - 2} GL transactions (Jan-Mar 2025)")
    else:
        print(f"GL already has {ws_gl.max_row} rows — looks OK")
else:
    print(f"WARNING: Sheet '{GL_SHEET}' not found!")

# ============================================================
# 3. FIX ASSUMPTIONS
# ============================================================
ASSUME_SHEET = "Assumptions"
if ASSUME_SHEET in wb.sheetnames:
    ws_a = wb[ASSUME_SHEET]
    print(f"\n--- Checking {ASSUME_SHEET} ---")
    print(f"Current range: {ws_a.max_row} rows")

    # Check if row 6+ has driver data
    has_drivers = False
    for r in range(6, min(ws_a.max_row + 1, 25)):
        if ws_a.cell(row=r, column=1).value and ws_a.cell(row=r, column=2).value:
            has_drivers = True
            break

    if not has_drivers:
        print("Assumptions has no driver values — adding defaults")
        drivers = [
            ["Revenue Growth Rate", 0.08],
            ["iGO Revenue Share", 0.35],
            ["Affirm Revenue Share", 0.25],
            ["InsureSight Revenue Share", 0.22],
            ["DocFast Revenue Share", 0.18],
            ["AWS Cost per Transaction", 2.50],
            ["AWS Monthly Base", 12000],
            ["Headcount - Engineering", 45],
            ["Headcount - Sales", 30],
            ["Headcount - Marketing", 15],
            ["Headcount - Finance", 10],
            ["Headcount - HR", 8],
            ["Avg Salary - Engineering", 125000],
            ["Avg Salary - Sales", 95000],
            ["Avg Salary - Marketing", 85000],
            ["Monthly Overhead Allocation", 500],
        ]
        for i, (name, val) in enumerate(drivers):
            ws_a.cell(row=6 + i, column=1, value=name)
            ws_a.cell(row=6 + i, column=2, value=val)
        print(f"Added {len(drivers)} driver values starting at row 6")
    else:
        print("Assumptions already has driver data — OK")
        for r in range(6, min(ws_a.max_row + 1, 15)):
            n = ws_a.cell(row=r, column=1).value
            v = ws_a.cell(row=r, column=2).value
            if n:
                print(f"  Row {r}: {n} = {v}")
else:
    print(f"WARNING: Sheet '{ASSUME_SHEET}' not found!")

# ============================================================
# 4. FIX CHECKS SHEET
# ============================================================
CHECKS_SHEET = "Checks"
if CHECKS_SHEET in wb.sheetnames:
    ws_chk = wb[CHECKS_SHEET]
    print(f"\n--- Checking {CHECKS_SHEET} ---")
    if ws_chk.cell(row=1, column=1).value is None:
        ws_chk.cell(row=1, column=1, value="Check Name")
        ws_chk.cell(row=1, column=2, value="Status")
        ws_chk.cell(row=1, column=3, value="Details")
        ws_chk.cell(row=1, column=4, value="Timestamp")
        print("Added Checks headers")
    else:
        print(f"Checks headers exist: {ws_chk.cell(row=1,column=1).value}")
else:
    print(f"WARNING: Sheet '{CHECKS_SHEET}' not found!")

# ============================================================
# SAVE
# ============================================================
print(f"\n--- Saving ---")
wb.save(DEMO_PATH)
print(f"Saved: {DEMO_PATH}")
print(f"Backup at: {BACKUP_PATH}")
print("\nDone! Open the file in Excel and test:")
print("  1. Action 7 (Data Quality Scan) — should show findings")
print("  2. Action 46 (Variance Commentary) — should show narratives")
print("  3. Action 12 (Executive Dashboard) — should show real numbers")
print("  4. Action 3 (Reconciliation) — should show PASS/FAIL checks")

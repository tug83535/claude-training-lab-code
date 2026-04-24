"""
Fix Demo File Data v4 — Add real numeric values to product Revenue rows.

Problem: The product-level Revenue/Cost of Revenue cells contain formulas
that Excel calculates at runtime, but the chart-building VBA reads them as
formulas (strings) and gets 0. This makes all charts empty.

Fix: Overwrite formula cells with hardcoded numeric values that look realistic.
The charts will then have real data to plot.
"""

import openpyxl
from openpyxl.styles import Font
import os
import sys

DEMO_PATH = r"C:\Users\connor.atlee\RecTrial\DemoFile\ExcelDemoFile_adv.xlsm"

if not os.path.exists(DEMO_PATH):
    print(f"ERROR: Demo file not found at {DEMO_PATH}")
    sys.exit(1)

wb = openpyxl.load_workbook(DEMO_PATH, keep_vba=True)
ws = wb["P&L - Monthly Trend"]

print(f"Opened: {DEMO_PATH}")

# Check what's currently in the product revenue cells
print("\nCurrent cell values (before fix):")
for r in [7, 8, 9, 15, 16, 17, 23, 24, 25, 31, 32, 33, 39, 40, 41]:
    label = ws.cell(row=r, column=1).value
    jan_val = ws.cell(row=r, column=2).value
    feb_val = ws.cell(row=r, column=3).value
    is_formula = isinstance(jan_val, str) and str(jan_val).startswith("=")
    print(f"  Row {r}: {label:30s} | Jan={jan_val} | {'FORMULA' if is_formula else 'VALUE'}")

# Product revenue data (monthly, Jan-Dec 2025)
# Scale: iGO is ~55%, Affirm ~28%, InsureSight ~12%, DocFast ~5% of total
# Total Revenue per month: 9.25M -> 11.35M (growing)
product_data = {
    # iGO Revenue (row 15) — ~55% of total
    15: [5087500, 5214000, 5362500, 5456000, 5582500, 5709000,
         5610000, 5802500, 5896000, 5989500, 6077500, 6242500],
    # iGO Cost of Revenue (row 16) — ~40% of iGO revenue
    16: [2035000, 2085600, 2145000, 2182400, 2233000, 2283600,
         2244000, 2321000, 2358400, 2395800, 2431000, 2497000],
    # iGO Contribution Margin (row 17) — Revenue - CoR
    17: [3052500, 3128400, 3217500, 3273600, 3349500, 3425400,
         3366000, 3481500, 3537600, 3593700, 3646500, 3745500],

    # Affirm Revenue (row 23) — ~28% of total
    23: [2590000, 2654400, 2730000, 2777600, 2842000, 2906400,
         2856000, 2954000, 3001600, 3049200, 3094000, 3178000],
    # Affirm Cost of Revenue (row 24) — ~42% of Affirm revenue
    24: [1087800, 1114848, 1146600, 1166592, 1193640, 1220688,
         1199520, 1240680, 1260672, 1280664, 1299480, 1334760],
    # Affirm Contribution Margin (row 25)
    25: [1502200, 1539552, 1583400, 1611008, 1648360, 1685712,
         1656480, 1713320, 1740928, 1768536, 1794520, 1843240],

    # InsureSight Revenue (row 31) — ~12% of total
    31: [1110000, 1137600, 1170000, 1190400, 1218000, 1245600,
         1224000, 1266000, 1286400, 1306800, 1326000, 1362000],
    # InsureSight Cost of Revenue (row 32) — ~45% of InsureSight revenue
    32: [499500, 511920, 526500, 535680, 548100, 560520,
         550800, 569700, 578880, 588060, 596700, 612900],
    # InsureSight Contribution Margin (row 33)
    33: [610500, 625680, 643500, 654720, 669900, 685080,
         673200, 696300, 707520, 718740, 729300, 749100],

    # DocFast Revenue (row 39) — ~5% of total
    39: [462500, 474000, 487500, 496000, 507500, 519000,
         510000, 527500, 536000, 544500, 552500, 567500],
    # DocFast Cost of Revenue (row 40) — ~48% of DocFast revenue
    40: [222000, 227520, 234000, 238080, 243600, 249120,
         244800, 253200, 257280, 261360, 265200, 272400],
    # DocFast Contribution Margin (row 41)
    41: [240500, 246480, 253500, 257920, 263900, 269880,
         265200, 274300, 278720, 283140, 287300, 295100],

    # Consolidated Revenue (row 7) — sum of all products
    7: [9250000, 9480000, 9750000, 9920000, 10150000, 10380000,
        10200000, 10550000, 10720000, 10890000, 11050000, 11350000],
    # Consolidated Cost of Revenue (row 8)
    8: [3844300, 3939888, 4052100, 4122752, 4218340, 4313928,
        4239120, 4384580, 4455232, 4525884, 4592380, 4717060],
    # Consolidated Contribution Margin (row 9)
    9: [5405700, 5540112, 5697900, 5797248, 5931660, 6066072,
        5960880, 6165420, 6264768, 6364116, 6457620, 6632940],
}

# Contribution Margin % rows (10, 18, 20, 26, 28, 34, 36, 42, 44)
# These are percentages — CM$ / Revenue
cm_pct_rows = {
    10: (9, 7),    # Consolidated CM% = row 9 / row 7
    18: (17, 15),  # iGO CM%
    20: (17, 15),  # iGO CM+R&D% (use same as CM% for simplicity)
    26: (25, 23),  # Affirm CM%
    28: (25, 23),  # Affirm CM+R&D%
    34: (33, 31),  # InsureSight CM%
    36: (33, 31),  # InsureSight CM+R&D%
    42: (41, 39),  # DocFast CM%
    44: (41, 39),  # DocFast CM+R&D%
}

# Write product data (months in cols 2-13)
for row_num, monthly_vals in product_data.items():
    for i, val in enumerate(monthly_vals):
        c = i + 2  # Col 2 = Jan, Col 13 = Dec
        ws.cell(row=row_num, column=c, value=val)
        ws.cell(row=row_num, column=c).number_format = "#,##0"

    # Q1 (col 14) = Jan + Feb + Mar
    q1 = sum(monthly_vals[0:3])
    ws.cell(row=row_num, column=14, value=q1)
    ws.cell(row=row_num, column=14).number_format = "#,##0"

    # Q2 (col 15) = Apr + May + Jun
    q2 = sum(monthly_vals[3:6])
    ws.cell(row=row_num, column=15, value=q2)
    ws.cell(row=row_num, column=15).number_format = "#,##0"

    # Q3 (col 16) = Jul + Aug + Sep
    q3 = sum(monthly_vals[6:9])
    ws.cell(row=row_num, column=16, value=q3)
    ws.cell(row=row_num, column=16).number_format = "#,##0"

    # Q4 (col 17) = Oct + Nov + Dec
    q4 = sum(monthly_vals[9:12])
    ws.cell(row=row_num, column=17, value=q4)
    ws.cell(row=row_num, column=17).number_format = "#,##0"

    # FY Total (col 18) = sum of all months
    fy = sum(monthly_vals)
    ws.cell(row=row_num, column=18, value=fy)
    ws.cell(row=row_num, column=18).number_format = "#,##0"

    # Budget (col 19) — FY total +/- variance
    import random
    random.seed(42 + row_num)
    variance = random.uniform(-0.12, 0.08)
    budget = fy * (1 + variance)
    ws.cell(row=row_num, column=19, value=round(budget, 0))
    ws.cell(row=row_num, column=19).number_format = "#,##0"

print("\nProduct data written (rows 7-41, cols 2-19)")

# Write CM% rows
for pct_row, (cm_row, rev_row) in cm_pct_rows.items():
    for c in range(2, 19):
        cm_val = ws.cell(row=cm_row, column=c).value
        rev_val = ws.cell(row=rev_row, column=c).value
        if cm_val and rev_val and rev_val != 0:
            try:
                pct = float(cm_val) / float(rev_val)
                ws.cell(row=pct_row, column=c, value=round(pct, 4))
                ws.cell(row=pct_row, column=c).number_format = "0.0%"
            except (ValueError, TypeError, ZeroDivisionError):
                pass

print("CM% rows calculated")

# Also write CM+R&D rows (11, 19, 27, 35, 43) — slightly lower than CM
cmrd_rows = {
    11: (9, 7),    # Consolidated
    12: (9, 7),    # Consolidated CM+R&D %
    19: (17, 15),  # iGO
    20: (17, 15),  # iGO %
    27: (25, 23),  # Affirm
    28: (25, 23),  # Affirm %
    35: (33, 31),  # InsureSight
    36: (33, 31),  # InsureSight %
    43: (41, 39),  # DocFast
    44: (41, 39),  # DocFast %
}

for cmrd_row, (cm_row, rev_row) in cmrd_rows.items():
    label = ws.cell(row=cmrd_row, column=1).value
    if label and "%" in str(label):
        # Percentage row — already handled above
        continue
    elif label and "$" in str(label):
        # Dollar row — CM minus ~5% for R&D
        for c in range(2, 19):
            cm_val = ws.cell(row=cm_row, column=c).value
            if cm_val and isinstance(cm_val, (int, float)):
                ws.cell(row=cmrd_row, column=c, value=round(cm_val * 0.92, 0))
                ws.cell(row=cmrd_row, column=c).number_format = "#,##0"

print("CM+R&D rows calculated")

# Verify final state
print("\nFinal spot check:")
for r in [7, 15, 23, 31, 39, 48]:
    label = ws.cell(row=r, column=1).value
    jan = ws.cell(row=r, column=2).value
    dec = ws.cell(row=r, column=13).value
    bud = ws.cell(row=r, column=19).value
    jan_str = f"{jan:,.0f}" if isinstance(jan, (int, float)) else str(jan)
    dec_str = f"{dec:,.0f}" if isinstance(dec, (int, float)) else str(dec)
    bud_str = f"{bud:,.0f}" if isinstance(bud, (int, float)) and bud else "N/A"
    print(f"  Row {r}: {str(label):30s} | Jan={jan_str:>12s} | Dec={dec_str:>12s} | Budget={bud_str}")

# Save
wb.save(DEMO_PATH)
print(f"\nSaved: {DEMO_PATH}")
print("\nOpen in Excel, run CleanupAllOutputSheets, then test Action 12.")
print("Charts should now show real bars, lines, and pie slices.")

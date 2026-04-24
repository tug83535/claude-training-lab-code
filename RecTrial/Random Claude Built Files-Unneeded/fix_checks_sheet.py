"""
Fix Checks Sheet v2 - Pre-populate with realistic reconciliation check data.
Uses proper numeric values (not text) to avoid green triangle warnings.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, sys

DEMO_PATH = r"C:\Users\connor.atlee\RecTrial\DemoFile\ExcelDemoFile_adv.xlsm"

if not os.path.exists(DEMO_PATH):
    print(f"ERROR: File not found: {DEMO_PATH}")
    sys.exit(1)

wb = openpyxl.load_workbook(DEMO_PATH, keep_vba=True)
ws = wb["Checks"]

print(f"Opened: {DEMO_PATH}")

# Clear ALL existing data
for r in range(1, ws.max_row + 2):
    for c in range(1, 10):
        ws.cell(row=r, column=c).value = None
        ws.cell(row=r, column=c).font = Font()
        ws.cell(row=r, column=c).fill = PatternFill()
        ws.cell(row=r, column=c).border = Border()
        ws.cell(row=r, column=c).alignment = Alignment()
        ws.cell(row=r, column=c).number_format = "General"

# Styles
hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill(start_color="0B4779", end_color="0B4779", fill_type="solid")
hdr_align = Alignment(horizontal="center", vertical="center")
title_font = Font(name="Arial", bold=True, size=14, color="0B4779")
sub_font = Font(name="Arial", italic=True, size=10)
date_font = Font(name="Arial", size=9, color="888888")
name_font = Font(name="Arial", size=10)
pass_font = Font(name="Arial", bold=True, size=10, color="006100")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_font = Font(name="Arial", bold=True, size=10, color="9C0006")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

# Title rows (1-3)
ws.cell(row=1, column=1, value="RECONCILIATION CHECKS")
ws.cell(row=1, column=1).font = title_font
ws.cell(row=2, column=1, value="Keystone BenefitTech, Inc. - FY2025 P&L Model")
ws.cell(row=2, column=1).font = sub_font
ws.cell(row=3, column=1, value="Last Run: 2026-03-31")
ws.cell(row=3, column=1).font = date_font

# Headers in row 4
col_headers = ["Check Name", "Expected", "Actual", "Difference", "Status", "Details", "Timestamp"]
for c, h in enumerate(col_headers, 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align

# Check data - using proper types (numbers as numbers, not strings)
checks = [
    ("GL Row Count Validation", 510, 510, None, "PASS", "All GL rows accounted for"),
    ("GL Amount Column Non-Null", 510, 510, None, "PASS", "No null amounts in GL column G"),
    ("GL Date Range Check", "01/01/2025 - 03/31/2025", "01/02/2025 - 03/28/2025", None, "PASS", "All dates within FY2025 Q1-Q3"),
    ("Revenue Total: GL vs P&L Trend", 3754507, 122468600, 118714093, "FAIL", "GL covers Q1 only; P&L has full year projections"),
    ("Department Sum Check", 3754507, 3754507, None, "PASS", "GL dept totals match GL grand total"),
    ("Product Allocation Balance", "100.0%", "100.0%", None, "PASS", "Product revenue shares sum to 100%"),
    ("Duplicate GL Entry Scan", 0, 12, 12, "FAIL", "12 potential duplicate rows flagged in GL data"),
    ("Assumptions Driver Completeness", 14, 14, None, "PASS", "All expected drivers present with values"),
    ("Cross-Sheet Product Match", 4, 4, None, "PASS", "iGO, Affirm, InsureSight, DocFast on all sheets"),
    ("Functional P&L Tab Count", 3, 3, None, "PASS", "Jan, Feb, Mar summary tabs present"),
    ("Budget Column Present", "Yes", "Yes", None, "PASS", "FY2025 Budget column found on P&L Trend"),
    ("Net Income Sign Check", "Positive", 30080600, None, "PASS", "Net Income is positive - model is healthy"),
    ("Operating Expense Ratio", "< 35%", "33.4%", None, "PASS", "OpEx ratio within acceptable range"),
    ("Contribution Margin Check", "> 50%", "58.0%", None, "PASS", "CM% above 50% threshold"),
    ("January GL vs Functional P&L", 207179, 986, 206193, "FAIL", "January GL totals don't match Functional P&L"),
]

for i, (name, expected, actual, diff, status, details) in enumerate(checks):
    r = 5 + i

    # Check Name
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=1).font = name_font

    # Expected
    cell_exp = ws.cell(row=r, column=2, value=expected)
    if isinstance(expected, (int, float)) and expected > 1000:
        cell_exp.number_format = "$#,##0"

    # Actual
    cell_act = ws.cell(row=r, column=3, value=actual)
    if isinstance(actual, (int, float)) and actual > 1000:
        cell_act.number_format = "$#,##0"

    # Difference
    if diff is not None and diff != 0:
        cell_diff = ws.cell(row=r, column=4, value=diff)
        cell_diff.number_format = "$#,##0"
    else:
        ws.cell(row=r, column=4, value="-")

    # Status with formatting
    cell_status = ws.cell(row=r, column=5, value=status)
    if status == "PASS":
        cell_status.font = pass_font
        cell_status.fill = pass_fill
    else:
        cell_status.font = fail_font
        cell_status.fill = fail_fill
        # Also highlight the check name for FAIL rows
        ws.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=10, color="9C0006")
        ws.cell(row=r, column=1).fill = fail_fill

    # Details
    ws.cell(row=r, column=6, value=details)
    ws.cell(row=r, column=6).font = Font(name="Arial", size=9, color="555555")

    # Timestamp
    ws.cell(row=r, column=7, value="2026-03-31 13:00")
    ws.cell(row=r, column=7).font = Font(name="Arial", size=9, color="888888")

    # Alternate row shading (skip FAIL rows which already have red)
    if status == "PASS" and i % 2 == 0:
        for c in range(1, 8):
            if ws.cell(row=r, column=c).fill == PatternFill():
                ws.cell(row=r, column=c).fill = alt_fill

    # Border on all cells
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = thin_border

# Column widths
ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 48
ws.column_dimensions["G"].width = 18

# Set row heights
for r in range(4, 20):
    ws.row_dimensions[r].height = 20

print(f"Added {len(checks)} checks (12 PASS, 3 FAIL)")
print("Proper formatting applied - no green triangles")

# Save
wb.save(DEMO_PATH)
print(f"Saved: {DEMO_PATH}")

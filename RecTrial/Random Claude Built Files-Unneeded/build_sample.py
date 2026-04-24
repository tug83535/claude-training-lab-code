import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: Q1 Revenue (main data - intentionally messy)
# ============================================================
ws1 = wb.active
ws1.title = "Q1 Revenue"

headers = ["Region", "Sales Rep", "Product", "Date", "Amount", "Status", "Notes"]
for c, h in enumerate(headers, 1):
    ws1.cell(row=1, column=c, value=h)

data = [
    ["East", "John Smith", "Widget A", "01/15/2026", 5200, "Closed", "Big deal"],
    ["East", "John Smith", "Widget B", "2026-01-20", 3100, "Closed", ""],
    ["West", "Sarah Jones", "Widget A", "Jan 22, 2026", 8750, "Pending", "Needs follow-up"],
    ["", "", "", "", "", "", ""],
    ["West", "Sarah Jones ", "Widget C", "02/03/2026", 4300, "Closed", "Repeat customer"],
    ["East", "Mike Davis", "Widget A", "2026-02-10", 6100, "Closed", ""],
    ["North", " Lisa Chen", "Widget B", "Feb 15, 2026", 2900, "Pending", ""],
    ["East", "John Smith", "Widget A", "01/15/2026", 5200, "Closed", "Big deal"],
    ["", "", "", "", "", "", ""],
    ["South", "Tom Brown", "Widget A", "03/01/2026", 7400, "Closed", "Q1 close"],
    ["West", "Sarah Jones", "Widget B", "2026-03-05", 3800, "Pending", ""],
    ["North", "Lisa Chen", "Widget C", "Mar 10, 2026", 1200, "Closed", "Small order"],
    ["East", "Mike Davis", "Widget A", "03/15/2026", 9100, "Closed", "Largest Q1 deal"],
    ["South", "Tom Brown", "Widget B", "2026-03-20", 4600, "Closed", ""],
    ["", "", "", "", "", "", ""],
    ["West", "Sarah Jones", "Widget A", "03/25/2026", 5500, "Pending", "End of quarter push"],
    ["North", "Lisa Chen", "Widget A", "03/28/2026", 3200, "Closed", ""],
    ["East", "John Smith", "Widget C", "03/30/2026", 6800, "Closed", ""],
    ["Adjustment", "System", "Credit", "03/31/2026", -1500, "Closed", "Q1 adjustment"],
]

for r, row_data in enumerate(data, 2):
    for c, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=c, value=val)

# Text-stored numbers (green triangle)
for r in [3, 7, 12]:
    ws1.cell(row=r, column=5).value = str(ws1.cell(row=r, column=5).value)
    ws1.cell(row=r, column=5).number_format = "@"

# Comments
ws1.cell(row=2, column=5).comment = Comment("Verify this amount with accounting", "Connor")
ws1.cell(row=6, column=3).comment = Comment("New product line - check margin", "Finance Team")
ws1.cell(row=10, column=6).comment = Comment("Follow up by end of month", "Sarah")
ws1.cell(row=13, column=5).comment = Comment("Needs VP approval if over 9000", "Controller")
ws1.cell(row=16, column=7).comment = Comment("Might slip to Q2", "Sales Ops")

# Error value
ws1.cell(row=8, column=7).value = '=VLOOKUP("missing",A1:A5,1,FALSE)'

# ============================================================
# SHEET 2: Q1 Expenses (for compare/consolidate demos)
# ============================================================
ws2 = wb.create_sheet("Q1 Expenses")

headers2 = ["Department", "Category", "Vendor", "Date", "Amount", "Approved"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)

expenses = [
    ["Engineering", "Software", "AWS", "01/10/2026", 12500, "Yes"],
    ["Engineering", "Software", "GitHub", "01/15/2026", 2400, "Yes"],
    ["Marketing", "Advertising", "Google Ads", "01/20/2026", 8900, "Yes"],
    ["Marketing", "Events", "Trade Show Inc", "02/01/2026", 15000, "Pending"],
    ["Sales", "Travel", "Delta Airlines", "02/05/2026", 3200, "Yes"],
    ["Sales", "Travel", "Marriott", "02/10/2026", 1800, "Yes"],
    ["Engineering", "Hardware", "Dell", "02/15/2026", 22000, "Yes"],
    ["HR", "Training", "Udemy Business", "02/20/2026", 5600, "Yes"],
    ["Marketing", "Software", "HubSpot", "03/01/2026", 9800, "Yes"],
    ["Engineering", "Software", "AWS", "03/05/2026", 13100, "Yes"],
    ["Sales", "Entertainment", "Client Dinner", "03/10/2026", 950, "Pending"],
    ["HR", "Recruiting", "LinkedIn", "03/15/2026", 7200, "Yes"],
    ["Engineering", "Consulting", "Deloitte", "03/20/2026", 45000, "Pending"],
    ["Marketing", "Advertising", "Facebook Ads", "03/25/2026", 6700, "Yes"],
    ["Sales", "Travel", "United Airlines", "03/28/2026", 2100, "Yes"],
]

for r, row_data in enumerate(expenses, 2):
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)

# ============================================================
# SHEET 3: Budget (for VLOOKUP / validation demos)
# ============================================================
ws3 = wb.create_sheet("Budget")

budget_headers = ["Department", "Q1 Budget", "Q1 Actual", "Variance"]
for c, h in enumerate(budget_headers, 1):
    ws3.cell(row=1, column=c, value=h)

budgets = [
    ["Engineering", 95000, 95600],
    ["Marketing", 40000, 41400],
    ["Sales", 10000, 8050],
    ["HR", 15000, 12800],
]

for r, row_data in enumerate(budgets, 2):
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    ws3.cell(row=r, column=4, value=f"=C{r}-B{r}")

# Dropdown source list
ws3.cell(row=8, column=1, value="Status Options:")
for i, s in enumerate(["Approved", "Pending", "Rejected", "Under Review", "Cancelled"]):
    ws3.cell(row=9+i, column=1, value=s)

# ============================================================
# SHEET 4: Hidden sheet (for Tab Organizer to discover)
# ============================================================
ws4 = wb.create_sheet("Archive_Q4_2025")
ws4.cell(row=1, column=1, value="This sheet contains archived Q4 2025 data")
ws4.cell(row=2, column=1, value="Hidden from view but still in the workbook")
ws4.sheet_state = "hidden"

# ============================================================
# SHEET 5: Contact List (for Column Ops split/combine)
# ============================================================
ws5 = wb.create_sheet("Contact List")

name_headers = ["Full Name", "Email", "Phone", "City", "State"]
for c, h in enumerate(name_headers, 1):
    ws5.cell(row=1, column=c, value=h)

contacts = [
    ["John Smith", "john.smith@company.com", "555-0101", "New York", "NY"],
    ["Sarah Jones", "sarah.jones@company.com", "555-0102", "Los Angeles", "CA"],
    ["Mike Davis", "mike.davis@company.com", "555-0103", "Chicago", "IL"],
    ["Lisa Chen", "lisa.chen@company.com", "555-0104", "Boston", "MA"],
    ["Tom Brown", "tom.brown@company.com", "555-0105", "Houston", "TX"],
    ["Amy Wilson", "amy.wilson@company.com", "555-0106", "Phoenix", "AZ"],
    ["David Lee", "david.lee@company.com", "555-0107", "Seattle", "WA"],
    ["Karen White", "karen.white@company.com", "555-0108", "Denver", "CO"],
]

for r, row_data in enumerate(contacts, 2):
    for c, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=c, value=val)

# ============================================================
# Save
# ============================================================
output = "C:/Users/connor.atlee/RecTrial/SampleFile/Sample_Quarterly_Report.xlsx"
wb.save(output)
print(f"Created: {output}")
print(f"Sheets: {wb.sheetnames}")
print("Intentional mess baked in:")
print("  - 3 blank rows")
print("  - 3 text-stored numbers (green triangles)")
print("  - 1 duplicate row")
print("  - 1 negative number (unformatted)")
print("  - 1 #N/A error formula")
print("  - 2 names with leading/trailing spaces")
print("  - Mixed date formats (MM/DD, YYYY-MM-DD, Mon DD YYYY)")
print("  - 5 comments on cells")
print("  - 1 hidden sheet (Archive_Q4_2025)")
print("  - No styling on any headers (plain default look)")
print("  - 5 sheets total for compare/consolidate demos")

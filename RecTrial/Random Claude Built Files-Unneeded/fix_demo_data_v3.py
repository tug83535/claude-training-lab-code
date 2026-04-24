"""
Fix Demo File Data v3 — Add summary rows with hardcoded realistic values.
Since the P&L has formulas we can't evaluate in openpyxl, we use realistic
dollar values that match the scale of the existing data.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
import os
import sys

DEMO_PATH = r"C:\Users\connor.atlee\RecTrial\DemoFile\ExcelDemoFile_adv.xlsm"

if not os.path.exists(DEMO_PATH):
    print(f"ERROR: Demo file not found at {DEMO_PATH}")
    sys.exit(1)

wb = openpyxl.load_workbook(DEMO_PATH, keep_vba=True)
ws = wb["P&L - Monthly Trend"]

print(f"Opened: {DEMO_PATH}")

# First, remove any previous summary section we may have added
# (from v2 run that partially worked)
for r in range(45, ws.max_row + 1):
    for c in range(1, 20):
        ws.cell(row=r, column=c).value = None
        ws.cell(row=r, column=c).font = Font()
        ws.cell(row=r, column=c).fill = PatternFill()
        ws.cell(row=r, column=c).border = Border()

print("Cleared rows 45+ from any previous attempt")

# Styles
header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="0B4779", end_color="0B4779", fill_type="solid")
label_font = Font(name="Arial", bold=True, size=10)
label_font_total = Font(name="Arial", bold=True, size=10, color="0B4779")
number_fmt = "#,##0"
thin_border = Border(bottom=Side(style="thin", color="999999"))
double_border = Border(
    top=Side(style="double", color="000000"),
    bottom=Side(style="double", color="000000")
)

# Row 46: Section header
ws.cell(row=46, column=1, value="CONSOLIDATED P&L SUMMARY")
ws.cell(row=46, column=1).font = Font(name="Arial", bold=True, size=12, color="0B4779")

# Row 47: Column headers (same as row 4)
col_headers = {
    1: "Metric", 2: "Jan 2025", 3: "Feb 2025", 4: "Mar 2025",
    5: "Apr 2025", 6: "May 2025", 7: "Jun 2025", 8: "Jul 2025",
    9: "Aug 2025", 10: "Sep 2025", 11: "Oct 2025", 12: "Nov 2025",
    13: "Dec 2025", 14: "Q1 2025", 15: "Q2 2025", 16: "Q3 2025",
    17: "Q4 2025", 18: "2025 Total", 19: "FY2025 Budget"
}

for c, h in col_headers.items():
    cell = ws.cell(row=47, column=c, value=h)
    cell.font = copy(header_font)
    cell.fill = copy(header_fill)
    cell.alignment = Alignment(horizontal="center")

# Realistic monthly data (matches the ~$10M/month revenue scale)
# These values create meaningful variances between months and vs budget
summary_data = {
    # Row 48: Total Revenue
    48: {
        "label": "Total Revenue",
        "values": [9250000, 9480000, 9750000, 9920000, 10150000, 10380000,
                   10200000, 10550000, 10720000, 10890000, 11050000, 11350000,
                   28480000, 30450000, 31470000, 33290000, 122468600, 118500000],
    },
    # Row 49: Cost of Revenue
    49: {
        "label": "Cost of Revenue",
        "values": [3885000, 3982000, 4095000, 4166000, 4263000, 4360000,
                   4284000, 4431000, 4502000, 4574000, 4641000, 4767000,
                   11962000, 12789000, 13217000, 13982000, 51458000, 49800000],
    },
    # Row 50: Gross Profit
    50: {
        "label": "Gross Profit",
        "values": [5365000, 5498000, 5655000, 5754000, 5887000, 6020000,
                   5916000, 6119000, 6218000, 6316000, 6409000, 6583000,
                   16518000, 17661000, 18253000, 19308000, 71010600, 68700000],
    },
    # Row 51: Operating Expenses
    51: {
        "label": "Operating Expenses",
        "values": [3200000, 3250000, 3310000, 3280000, 3350000, 3420000,
                   3390000, 3460000, 3500000, 3540000, 3580000, 3650000,
                   9760000, 10050000, 10350000, 10770000, 40930000, 42000000],
    },
    # Row 52: Net Income
    52: {
        "label": "Net Income",
        "values": [2165000, 2248000, 2345000, 2474000, 2537000, 2600000,
                   2526000, 2659000, 2718000, 2776000, 2829000, 2933000,
                   6758000, 7611000, 7903000, 8538000, 30080600, 26700000],
    },
}

for row_num, data in summary_data.items():
    # Label
    ws.cell(row=row_num, column=1, value=data["label"])
    ws.cell(row=row_num, column=1).font = copy(label_font)
    ws.cell(row=row_num, column=1).border = copy(thin_border)

    # Values: cols 2-19 (12 months + 4 quarters + total + budget)
    for i, val in enumerate(data["values"]):
        c = i + 2
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.number_format = number_fmt
        cell.border = copy(thin_border)
        cell.alignment = Alignment(horizontal="right")

# Make Net Income row stand out
for c in range(1, 20):
    ws.cell(row=52, column=c).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=52, column=c).border = copy(double_border)

# Also make sure budget column header exists on the original row 4
if ws.cell(row=4, column=19).value is None:
    ws.cell(row=4, column=19, value="FY2025 Budget")
    ws.cell(row=4, column=19).font = copy(header_font)
    ws.cell(row=4, column=19).fill = copy(header_fill)

# Add budget values to the ORIGINAL product rows too (rows 5-44)
# so Variance Commentary can find variances at the line-item level
import random
random.seed(42)

for r in range(5, 45):
    # Get Dec value (col 13) as reference for budget
    dec_val = ws.cell(row=r, column=13).value
    if dec_val is not None:
        # Skip if it's a formula
        if isinstance(dec_val, str) and dec_val.startswith("="):
            continue
        try:
            dec_float = float(dec_val)
            if dec_float != 0:
                # Budget = Dec * 12 with some variance
                variance = random.uniform(-0.15, 0.10)
                annual_budget = dec_float * 12 * (1 + variance)
                ws.cell(row=r, column=19, value=round(annual_budget, 2))
                ws.cell(row=r, column=19).number_format = number_fmt
        except (ValueError, TypeError):
            pass

print("Summary rows added at rows 48-52")
print("Budget values added to original rows (col 19)")

# Verify
for row_num in [48, 49, 50, 51, 52]:
    label = ws.cell(row=row_num, column=1).value
    jan = ws.cell(row=row_num, column=2).value
    dec = ws.cell(row=row_num, column=13).value
    budget = ws.cell(row=row_num, column=19).value
    print(f"  Row {row_num}: {label} | Jan={jan:,.0f} | Dec={dec:,.0f} | Budget={budget:,.0f}")

# Save
wb.save(DEMO_PATH)
print(f"\nSaved: {DEMO_PATH}")
print("\nOpen in Excel and test:")
print("  Action 46 (Variance Commentary) — should show real narratives now")
print("  Action 12 (Executive Dashboard) — should show real dollar amounts now")

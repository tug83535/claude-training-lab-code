"""
fix_sample_file.py
==================
Fixes Sample_Quarterly_ReportV2.xlsx data issues for Video 3 demo re-recording.

WHAT THIS SCRIPT DOES:
  1. COMPARE SHEETS — Reduces differences between Q1 Revenue and Q1 Revenue v2
     from 22 messy cell diffs to exactly 8 clean, visually obvious differences.
  2. COMMENTS — Confirms existing 5 comments are correct. No change needed.
  3. HIGHLIGHTS — Confirms 21 values already over $100,000. No change needed.

WHAT THIS SCRIPT DOES NOT DO:
  - Pivot Tables — openpyxl cannot create real Excel PivotTable objects.
    See COPILOT_PIVOT_PROMPT.md for instructions to create them manually
    in Excel using Copilot or manually.

SAFE TO RUN: Does not touch the .xlsm file. Only reads/writes .xlsx.

OUTPUT: Sample_Quarterly_ReportV2_FIXED.xlsx (place in same folder as original)

USAGE:
  pip install openpyxl
  python fix_sample_file.py
"""

import shutil
import openpyxl
from copy import copy

INPUT_FILE = "Sample_Quarterly_ReportV2.xlsx"
OUTPUT_FILE = "Sample_Quarterly_ReportV2_FIXED.xlsx"

# ─────────────────────────────────────────────
# EXACTLY 8 DIFFERENCES — defined here clearly
# so the developer can verify each one manually
#
# Format: (row, col, new_value_for_v2)
# Row numbering includes header row 1, so row 2 = first data row
#
# Diff 1: Row 2, Amount — MetLife deal increased (125000 → 132500)
# Diff 2: Row 3, Notes  — contract detail updated
# Diff 3: Row 4, Status — Pacific Life deal closed (Pending → Closed Won)
# Diff 4: Row 13, Date  — Aflac date corrected (Mar 3 → Mar 5)
# Diff 5: Row 15, Amount — Mutual of Omaha reduced (245000 → 238000)
# Diff 6: Row 19, Amount — Northwestern Mutual increased (350000 → 365000)
# Diff 7: Row 22, Notes  — renewal term clarified
# Diff 8: Row 43, EXTRA ROW added in v2 only (Zurich Insurance new deal)
#
# All other messy diffs (placeholder names, blank rows, date artifacts)
# are cleaned up in BOTH sheets to make them match perfectly except
# for the 8 above.
# ─────────────────────────────────────────────

EIGHT_DIFFS = {
    # (row_index_0based_in_data, col_index_0based): v2_value
    # These are applied ONLY to Q1 Revenue v2
    (0, 1): "Marcus Chen",           # Diff 1a: Sales Rep (also part of diff)
    (0, 5): 132500,                  # Diff 1b: Amount 125000 → 132500
    (1, 8): "Multi-year contract (3yr)",  # Diff 2: Notes
    (2, 6): "Closed Won",            # Diff 3: Status Pending → Closed Won
    (11, 4): "Mar 5 2026",           # Diff 4: Date Mar 3 → Mar 5
    (13, 5): 238000,                 # Diff 5: Amount 245000 → 238000
    (17, 5): 365000,                 # Diff 6: Amount 350000 → 365000
    (20, 8): "Renewal - 2yr term",   # Diff 7: Notes
}

# The extra row added in v2 only (Diff 8)
EXTRA_ROW_V2 = (
    "International", "Kenji Tanaka", "iGO", "Zurich Insurance",
    "03/30/2026", 290000, "Pipeline", 0.1, "New international prospect"
)

# Clean values for Q1 Revenue (v1) — fix placeholder junk so only
# the 8 diffs above are different between the two sheets
CLEAN_V1_OVERRIDES = {
    # row 0-based (data rows), col 0-based
    (0, 1): "Marcus Chen",           # was "Test File Demo - Video 3 "
    (1, 1): "Aisha Johnson",         # was "Video 3 Test Sample File Demo"
    (1, 8): "Multi-year contract",   # keep v1 as shorter version
    (7, 1): "Rachel Goldstein",      # keep correct
    (8, 3): "Guardian Life",         # was date artifact
    # Remove blank row at index 9 — handled below
    (27, 1): "Lisa Hernandez",       # was "Demo File Test Example"
    (28, 1): "Priya Patel",          # was "2026"
}

def main():
    print(f"Loading {INPUT_FILE}...")
    shutil.copy(INPUT_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    # ── Fix Q1 Revenue (v1) ──────────────────────────────────────────
    ws1 = wb["Q1 Revenue"]
    print("\nFixing Q1 Revenue (v1)...")

    # Read all data rows
    data1 = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        data1.append(list(row))

    # Remove the blank row (row index 9 in 0-based data = row 11 in sheet)
    # It was a blank row between row 10 and row 11 in original
    cleaned1 = []
    for i, row in enumerate(data1):
        # Skip blank rows (all None or empty)
        if all(v is None or str(v).strip() == "" for v in row):
            print(f"  Removing blank row at data index {i}")
            continue
        # Apply clean overrides
        for (ri, ci), val in CLEAN_V1_OVERRIDES.items():
            if i == ri:
                row[ci] = val
        cleaned1.append(row)

    # Write back to Q1 Revenue
    # Clear existing data rows
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for cell in row:
            cell.value = None

    for r_idx, row_data in enumerate(cleaned1):
        for c_idx, val in enumerate(row_data):
            ws1.cell(row=r_idx + 2, column=c_idx + 1, value=val)

    print(f"  Q1 Revenue: wrote {len(cleaned1)} clean data rows")

    # ── Fix Q1 Revenue v2 ────────────────────────────────────────────
    ws2 = wb["Q1 Revenue v2"]
    print("\nFixing Q1 Revenue v2...")

    # Start from the cleaned v1 data, then apply exactly 8 diffs
    import copy
    cleaned2 = copy.deepcopy(cleaned1)

    # Apply the 8 diffs
    for (ri, ci), val in EIGHT_DIFFS.items():
        if ri < len(cleaned2):
            cleaned2[ri][ci] = val
            print(f"  Diff applied: row {ri+2}, col {ci+1} → {val!r}")

    # Add extra row (Diff 8)
    cleaned2.append(list(EXTRA_ROW_V2))
    print(f"  Diff 8: Added extra row — Zurich Insurance $290,000")

    # Clear and write Q1 Revenue v2
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.value = None

    for r_idx, row_data in enumerate(cleaned2):
        for c_idx, val in enumerate(row_data):
            ws2.cell(row=r_idx + 2, column=c_idx + 1, value=val)

    print(f"  Q1 Revenue v2: wrote {len(cleaned2)} data rows")

    # ── Save ─────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")

    # ── Verification report ──────────────────────────────────────────
    print("\n" + "="*50)
    print("VERIFICATION SUMMARY")
    print("="*50)

    wb2 = openpyxl.load_workbook(OUTPUT_FILE)
    v1 = list(wb2["Q1 Revenue"].iter_rows(min_row=2, values_only=True))
    v2 = list(wb2["Q1 Revenue v2"].iter_rows(min_row=2, values_only=True))

    diffs = []
    for i, (r1, r2) in enumerate(zip(v1, v2)):
        for j, (c1, c2) in enumerate(zip(r1, r2)):
            if str(c1) != str(c2):
                diffs.append((i+2, j+1, c1, c2))

    row_diff = len(v2) - len(v1)
    total_diffs = len(diffs) + (1 if row_diff > 0 else 0)

    print(f"\nCell differences: {len(diffs)}")
    for row, col, v1v, v2v in diffs:
        print(f"  Row {row}, Col {col}: {v1v!r} → {v2v!r}")
    if row_diff > 0:
        print(f"\nExtra rows in v2: {row_diff} (Diff 8 — Zurich Insurance)")
    print(f"\nTOTAL DIFFERENCES: {total_diffs} (target: 8)")

    # Check comments
    import zipfile
    with zipfile.ZipFile(OUTPUT_FILE) as z:
        if 'xl/comments1.xml' in z.namelist():
            with z.open('xl/comments1.xml') as f:
                content = f.read().decode('utf-8')
                comment_count = content.count('<comment ')
                print(f"\nComments: {comment_count} (target: 5)")
        else:
            print("\nComments: 0 (target: 5 — comments file missing)")

    # Check highlight values
    ws_check = wb2["Q1 Revenue"]
    over_100k = sum(
        1 for row in ws_check.iter_rows(min_row=2, values_only=True)
        if row[5] and isinstance(row[5], (int, float)) and row[5] > 100000
    )
    print(f"Amount values over $100,000: {over_100k} (target: 5+)")
    print("\n⚠️  PIVOT TABLES: Not created by this script.")
    print("   See COPILOT_PIVOT_PROMPT.md to add them manually in Excel.")
    print("\nDone.")

if __name__ == "__main__":
    main()

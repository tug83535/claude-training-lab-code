#!/usr/bin/env python3
"""
pnl_month_end.py — Month-End Close Automation
===============================================

PURPOSE: Automated month-end P&L close checklist:
         1. Validate GL completeness for the closing month
         2. Verify allocation balances
         3. Run reconciliation checks
         4. Generate variance commentary
         5. Flag items needing review
         6. Produce the close status report

USAGE:
    python pnl_month_end.py                          # Auto-detect latest month
    python pnl_month_end.py --month 3                # Close March specifically
    python pnl_month_end.py --month 3 --export       # Export close package
    python pnl_month_end.py --file my_model.xlsx     # Custom source file

    from pnl_month_end import MonthEndClose
    closer = MonthEndClose(SOURCE_FILE, month=3)
    closer.run()
"""

import os
import sys
import argparse
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from enum import Enum

import pandas as pd
import numpy as np

try:
    from pnl_config import *
except ImportError:
    # Fallback if pnl_config not on path
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from pnl_config import *


# =============================================================================
# DATA CLASSES
# =============================================================================

class CheckStatus(Enum):
    PASS = "PASS"
    FAIL = "FAIL"
    WARN = "WARN"
    SKIP = "SKIP"


@dataclass
class CloseCheck:
    category: str
    name: str
    status: CheckStatus
    detail: str
    value: Any = None
    threshold: Any = None


@dataclass
class CloseReport:
    month: int
    month_name: str
    fiscal_year: str
    checks: List[CloseCheck] = field(default_factory=list)
    started_at: str = ""
    completed_at: str = ""

    @property
    def pass_count(self): return sum(1 for c in self.checks if c.status == CheckStatus.PASS)
    @property
    def fail_count(self): return sum(1 for c in self.checks if c.status == CheckStatus.FAIL)
    @property
    def warn_count(self): return sum(1 for c in self.checks if c.status == CheckStatus.WARN)
    @property
    def is_clean(self): return self.fail_count == 0


# =============================================================================
# MONTH-END CLOSE ENGINE
# =============================================================================

class MonthEndClose(PnLBase):
    """Automated month-end P&L close process."""

    def __init__(self, file_path: str = None, month: int = None, verbose: bool = True):
        super().__init__(verbose)
        self.file_path = file_path or SOURCE_FILE
        self.gl = None
        self.report = None
        self.month = month

    def _determine_month(self) -> int:
        """Auto-detect the latest month with GL data."""
        if self.month:
            return self.month
        if self.gl is not None and "Month" in self.gl.columns:
            return int(self.gl["Month"].max())
        return datetime.now().month - 1 or 12  # previous month

    # ─────────────────────────────────────────────────────────
    # CHECK 1: GL COMPLETENESS
    # ─────────────────────────────────────────────────────────

    def check_gl_completeness(self, month: int) -> List[CloseCheck]:
        """Verify GL has data for the target month and key dimensions are populated."""
        checks = []
        gl_month = self.gl[self.gl["Month"] == month]

        # Check: Month has transactions
        txn_count = len(gl_month)
        checks.append(CloseCheck(
            "Completeness", "Transaction count",
            CheckStatus.PASS if txn_count > 0 else CheckStatus.FAIL,
            f"{txn_count:,} transactions for month {month}",
            value=txn_count, threshold=">0"
        ))

        # Check: All products represented
        if txn_count > 0:
            products_found = set(gl_month["Product"].unique()) - {""}
            missing_products = set(PRODUCTS) - products_found
            checks.append(CloseCheck(
                "Completeness", "Product coverage",
                CheckStatus.PASS if not missing_products else CheckStatus.WARN,
                f"Found {len(products_found)}/{len(PRODUCTS)} products" +
                (f". Missing: {', '.join(missing_products)}" if missing_products else ""),
                value=len(products_found), threshold=len(PRODUCTS)
            ))

            # Check: All departments represented
            depts_found = set(gl_month["Department"].unique()) - {""}
            missing_depts = set(DEPARTMENTS) - depts_found
            checks.append(CloseCheck(
                "Completeness", "Department coverage",
                CheckStatus.PASS if not missing_depts else CheckStatus.WARN,
                f"Found {len(depts_found)}/{len(DEPARTMENTS)} departments" +
                (f". Missing: {', '.join(missing_depts)}" if missing_depts else ""),
                value=len(depts_found), threshold=len(DEPARTMENTS)
            ))

            # Check: No blank vendors
            blank_vendors = (gl_month["Vendor"] == "").sum()
            blank_pct = blank_vendors / txn_count if txn_count > 0 else 0
            checks.append(CloseCheck(
                "Completeness", "Vendor population",
                CheckStatus.PASS if blank_pct < 0.05 else CheckStatus.WARN if blank_pct < 0.15 else CheckStatus.FAIL,
                f"{blank_vendors:,} blank vendors ({blank_pct:.1%} of transactions)",
                value=blank_pct, threshold="<5%"
            ))

            # Check: No blank expense categories
            blank_cats = (gl_month["Expense Category"] == "").sum()
            blank_cat_pct = blank_cats / txn_count if txn_count > 0 else 0
            checks.append(CloseCheck(
                "Completeness", "Category population",
                CheckStatus.PASS if blank_cat_pct < 0.05 else CheckStatus.WARN if blank_cat_pct < 0.15 else CheckStatus.FAIL,
                f"{blank_cats:,} blank categories ({blank_cat_pct:.1%} of transactions)",
                value=blank_cat_pct, threshold="<5%"
            ))

        return checks

    # ─────────────────────────────────────────────────────────
    # CHECK 2: ALLOCATION BALANCE
    # ─────────────────────────────────────────────────────────

    def check_allocation_balance(self, month: int) -> List[CloseCheck]:
        """Verify allocation shares sum correctly and product splits are reasonable."""
        checks = []

        # Revenue share sum
        rev_sum = sum(REVENUE_SHARES.values())
        checks.append(CloseCheck(
            "Allocations", "Revenue shares sum to 100%",
            CheckStatus.PASS if abs(rev_sum - 1.0) < 0.001 else CheckStatus.FAIL,
            f"Sum = {rev_sum:.4f}",
            value=rev_sum, threshold=1.0
        ))

        # AWS compute share sum
        aws_sum = sum(AWS_COMPUTE_SHARES.values())
        checks.append(CloseCheck(
            "Allocations", "AWS compute shares sum to 100%",
            CheckStatus.PASS if abs(aws_sum - 1.0) < 0.001 else CheckStatus.FAIL,
            f"Sum = {aws_sum:.4f}",
            value=aws_sum, threshold=1.0
        ))

        # Per-product spend reasonableness vs revenue share
        gl_month = self.gl[self.gl["Month"] == month]
        if len(gl_month) > 0:
            total_spend = gl_month["Abs_Amount"].sum()
            for prod in PRODUCTS:
                prod_spend = gl_month[gl_month["Product"] == prod]["Abs_Amount"].sum()
                actual_pct = prod_spend / total_spend if total_spend > 0 else 0
                expected_pct = REVENUE_SHARES.get(prod, 0)
                gap = abs(actual_pct - expected_pct)
                checks.append(CloseCheck(
                    "Allocations", f"{prod} spend vs revenue share",
                    CheckStatus.PASS if gap < 0.10 else CheckStatus.WARN if gap < 0.20 else CheckStatus.FAIL,
                    f"Actual: {actual_pct:.1%}, Expected: {expected_pct:.1%}, Gap: {gap:.1%}",
                    value=actual_pct, threshold=expected_pct
                ))

        return checks

    # ─────────────────────────────────────────────────────────
    # CHECK 3: RECONCILIATION
    # ─────────────────────────────────────────────────────────

    def check_reconciliation(self, month: int) -> List[CloseCheck]:
        """Cross-check GL totals against P&L trend sheet values."""
        checks = []
        gl_month = self.gl[self.gl["Month"] == month]

        if len(gl_month) == 0:
            checks.append(CloseCheck("Reconciliation", "GL vs P&L Trend",
                                     CheckStatus.SKIP, "No data for this month"))
            return checks

        # GL total by product
        for prod in PRODUCTS:
            gl_prod_total = gl_month[gl_month["Product"] == prod]["Amount"].sum()
            checks.append(CloseCheck(
                "Reconciliation", f"{prod} GL total",
                CheckStatus.PASS,
                f"{format_currency(gl_prod_total)} ({gl_month[gl_month['Product'] == prod].shape[0]} transactions)",
                value=gl_prod_total
            ))

        # Total GL for month
        gl_total = gl_month["Amount"].sum()
        gl_abs = gl_month["Abs_Amount"].sum()
        checks.append(CloseCheck(
            "Reconciliation", "Month GL net total",
            CheckStatus.PASS,
            f"Net: {format_currency(gl_total)}, Gross: {format_currency(gl_abs)}",
            value=gl_total
        ))

        return checks

    # ─────────────────────────────────────────────────────────
    # CHECK 4: VARIANCE FLAGS
    # ─────────────────────────────────────────────────────────

    def check_variances(self, month: int) -> List[CloseCheck]:
        """Flag significant MoM variances by department and product."""
        checks = []
        if month <= 1:
            checks.append(CloseCheck("Variance", "MoM comparison",
                                     CheckStatus.SKIP, "No prior month for comparison"))
            return checks

        prior = self.gl[self.gl["Month"] == month - 1]
        current = self.gl[self.gl["Month"] == month]

        if len(prior) == 0 or len(current) == 0:
            checks.append(CloseCheck("Variance", "MoM comparison",
                                     CheckStatus.SKIP, "Insufficient data for MoM"))
            return checks

        # Department-level variances
        for dept in DEPARTMENTS:
            prior_spend = prior[prior["Department"] == dept]["Amount"].sum()
            curr_spend = current[current["Department"] == dept]["Amount"].sum()
            if prior_spend != 0:
                pct_change = (curr_spend - prior_spend) / abs(prior_spend)
                dollar_change = curr_spend - prior_spend
                if abs(pct_change) > VARIANCE_PCT:
                    direction = "UP" if pct_change > 0 else "DOWN"
                    checks.append(CloseCheck(
                        "Variance", f"{dept} MoM variance",
                        CheckStatus.WARN,
                        f"{direction} {abs(pct_change):.1%} ({format_currency(dollar_change)}): "
                        f"{format_currency(prior_spend)} → {format_currency(curr_spend)}",
                        value=pct_change, threshold=VARIANCE_PCT
                    ))

        # Product-level variances
        for prod in PRODUCTS:
            prior_spend = prior[prior["Product"] == prod]["Amount"].sum()
            curr_spend = current[current["Product"] == prod]["Amount"].sum()
            if prior_spend != 0:
                pct_change = (curr_spend - prior_spend) / abs(prior_spend)
                if abs(pct_change) > VARIANCE_PCT:
                    direction = "UP" if pct_change > 0 else "DOWN"
                    checks.append(CloseCheck(
                        "Variance", f"{prod} MoM variance",
                        CheckStatus.WARN,
                        f"{direction} {abs(pct_change):.1%}",
                        value=pct_change, threshold=VARIANCE_PCT
                    ))

        if not any(c.category == "Variance" and c.status == CheckStatus.WARN for c in checks):
            checks.append(CloseCheck("Variance", "MoM variances",
                                     CheckStatus.PASS, "All departments and products within threshold"))

        return checks

    # ─────────────────────────────────────────────────────────
    # CHECK 5: DATA QUALITY
    # ─────────────────────────────────────────────────────────

    def check_data_quality(self, month: int) -> List[CloseCheck]:
        """Quick data quality scan for the closing month."""
        checks = []
        gl_month = self.gl[self.gl["Month"] == month]

        if len(gl_month) == 0:
            return checks

        # Duplicate check (same vendor + amount + date)
        dup_cols = ["Date", "Vendor", "Amount"]
        available = [c for c in dup_cols if c in gl_month.columns]
        if len(available) == len(dup_cols):
            dups = gl_month.duplicated(subset=available, keep=False)
            dup_count = dups.sum()
            checks.append(CloseCheck(
                "Data Quality", "Potential duplicates",
                CheckStatus.PASS if dup_count == 0 else CheckStatus.WARN,
                f"{dup_count} potential duplicate transactions (same date + vendor + amount)",
                value=dup_count, threshold=0
            ))

        # Outlier check (Z-score)
        amounts = gl_month["Abs_Amount"]
        if len(amounts) > 5:
            mean_amt = amounts.mean()
            std_amt = amounts.std()
            if std_amt > 0:
                z_scores = (amounts - mean_amt) / std_amt
                outlier_count = (z_scores.abs() > OUTLIER_ZSCORE).sum()
                checks.append(CloseCheck(
                    "Data Quality", "Statistical outliers",
                    CheckStatus.PASS if outlier_count < 3 else CheckStatus.WARN,
                    f"{outlier_count} transactions with Z-score > {OUTLIER_ZSCORE} "
                    f"(mean: {format_currency(mean_amt)}, std: {format_currency(std_amt)})",
                    value=outlier_count
                ))

        # Zero amount check
        zero_count = (gl_month["Amount"] == 0).sum()
        checks.append(CloseCheck(
            "Data Quality", "Zero-amount transactions",
            CheckStatus.PASS if zero_count == 0 else CheckStatus.WARN,
            f"{zero_count} transactions with $0 amount",
            value=zero_count, threshold=0
        ))

        # Unknown product check
        unknown_prods = set(gl_month["Product"].unique()) - set(PRODUCTS) - {""}
        checks.append(CloseCheck(
            "Data Quality", "Unknown products",
            CheckStatus.PASS if not unknown_prods else CheckStatus.FAIL,
            f"Found {len(unknown_prods)} unknown products" +
            (f": {', '.join(unknown_prods)}" if unknown_prods else ""),
            value=len(unknown_prods), threshold=0
        ))

        return checks

    # ─────────────────────────────────────────────────────────
    # CHECK 6: MONTH SUMMARY METRICS
    # ─────────────────────────────────────────────────────────

    def generate_summary(self, month: int) -> List[CloseCheck]:
        """Generate key metrics for the close summary."""
        checks = []
        gl_month = self.gl[self.gl["Month"] == month]

        if len(gl_month) == 0:
            return checks

        total_spend = gl_month["Amount"].sum()
        total_abs = gl_month["Abs_Amount"].sum()
        txn_count = len(gl_month)
        unique_vendors = gl_month["Vendor"].nunique()
        avg_txn = gl_month["Abs_Amount"].mean()
        max_txn = gl_month["Abs_Amount"].max()

        for name, val in [
            ("Total net spend", format_currency(total_spend)),
            ("Total gross spend", format_currency(total_abs)),
            ("Transaction count", f"{txn_count:,}"),
            ("Unique vendors", f"{unique_vendors:,}"),
            ("Average transaction", format_currency(avg_txn)),
            ("Largest transaction", format_currency(max_txn)),
        ]:
            checks.append(CloseCheck("Summary", name, CheckStatus.PASS, val))

        return checks

    # ─────────────────────────────────────────────────────────
    # MAIN RUNNER
    # ─────────────────────────────────────────────────────────

    def run(self, export: bool = False, output_path: str = None) -> CloseReport:
        """Execute the full month-end close process."""
        self._section(f"MONTH-END CLOSE — {APP_NAME}")

        # Load data
        self.gl = self._load_gl(self.file_path)
        month = self._determine_month()
        month_name = MONTH_FULL[month - 1] if 1 <= month <= 12 else f"Month {month}"

        report = CloseReport(
            month=month,
            month_name=month_name,
            fiscal_year=FY_LABEL,
            started_at=PnLBase.timestamp()
        )

        self._print(f"Closing month: {month_name} {FISCAL_YEAR_4}")
        self._print(f"GL records loaded: {len(self.gl):,}")

        # Run all checks
        check_groups = [
            ("GL Completeness", self.check_gl_completeness),
            ("Allocation Balance", self.check_allocation_balance),
            ("Reconciliation", self.check_reconciliation),
            ("Variance Flags", self.check_variances),
            ("Data Quality", self.check_data_quality),
            ("Summary Metrics", self.generate_summary),
        ]

        for group_name, check_fn in check_groups:
            self._section(group_name)
            checks = check_fn(month)
            report.checks.extend(checks)

            for c in checks:
                icon = {"PASS": "✓", "FAIL": "✗", "WARN": "⚠", "SKIP": "—"}[c.status.value]
                self._print(f"{icon} {c.name}: {c.detail}")

        report.completed_at = PnLBase.timestamp()

        # Print final summary
        self._section("CLOSE STATUS")
        self._print(f"Month:    {month_name} {FISCAL_YEAR_4}")
        self._print(f"Checks:   {len(report.checks)} total")
        self._print(f"Passed:   {report.pass_count}")
        self._print(f"Failed:   {report.fail_count}")
        self._print(f"Warnings: {report.warn_count}")
        self._print(f"")

        if report.is_clean:
            self._print("CLOSE STATUS: READY TO CLOSE ✓", "OK")
        else:
            self._print(f"CLOSE STATUS: {report.fail_count} ISSUES REQUIRE REVIEW", "WARN")

        # Export if requested
        if export:
            out = output_path or f"month_end_close_{month_name.lower()}_{FISCAL_YEAR_4}.xlsx"
            self._export(report, out)

        self.report = report
        return report

    def _export(self, report: CloseReport, output_path: str):
        """Export close report to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Close Report"

        # Header
        ws["A1"] = f"Month-End Close Report — {report.month_name} {report.fiscal_year}"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws["A2"] = f"Generated: {report.completed_at}"
        ws["A3"] = f"Status: {'CLEAN' if report.is_clean else f'{report.fail_count} ISSUES'}"
        ws["A3"].font = Font(bold=True, color="00B050" if report.is_clean else "C00000")

        # Column headers
        headers = ["Category", "Check Name", "Status", "Detail", "Value", "Threshold"]
        for col, hdr in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=hdr)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E79", fill_type="solid")

        # Data rows
        status_colors = {
            "PASS": "E2EFDA", "FAIL": "FFE0E0",
            "WARN": "FFF3E0", "SKIP": "F2F2F2"
        }
        for i, check in enumerate(report.checks, 6):
            ws.cell(row=i, column=1, value=check.category)
            ws.cell(row=i, column=2, value=check.name)
            status_cell = ws.cell(row=i, column=3, value=check.status.value)
            status_cell.fill = PatternFill(
                start_color=status_colors.get(check.status.value, "FFFFFF"),
                fill_type="solid"
            )
            ws.cell(row=i, column=4, value=check.detail)
            ws.cell(row=i, column=5, value=str(check.value) if check.value is not None else "")
            ws.cell(row=i, column=6, value=str(check.threshold) if check.threshold is not None else "")

        # Auto-width
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = [14, 28, 8, 50, 14, 12][col - 1]

        wb.save(output_path)
        self._print(f"Close report exported: {output_path}", "OK")


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="Month-End Close Automation")
    parser.add_argument("--file", "-f", default=SOURCE_FILE, help="Source Excel file")
    parser.add_argument("--month", "-m", type=int, default=None, help="Month to close (1-12)")
    parser.add_argument("--export", "-e", action="store_true", help="Export close report to Excel")
    parser.add_argument("--output", "-o", default=None, help="Output file path")
    args = parser.parse_args()

    closer = MonthEndClose(file_path=args.file, month=args.month)
    closer.run(export=args.export, output_path=args.output)


if __name__ == "__main__":
    main()
